from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from models import db, Vertical, Segment, Zone, Part, AffiliateLink, AffiliateNetwork, AffiliateCampaign, AffiliateStats, AIContent, SiteSettings, SocialChannel, VideoProject, VideoPublish, Article, Banner, Hotel, Attraction, Voucher, ArticleFeedback, ScheduledCSVImport, VoucherWidget, ContentEvent, AutoContentRule, ContentQueue, HotDeal, WardCommune, AccessTradeBanner, BacklinkKeyword, BacklinkInstance
from datetime import datetime, date, timedelta
import os, random, json

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///unilab.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'unilab-secret-2026'
# SQLite concurrency: WAL mode allows reads during writes, busy_timeout waits instead of failing
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'connect_args': {'timeout': 30},  # sqlite3 connect timeout (seconds)
}
db.init_app(app)

# Set SQLite pragmas for better concurrency (WAL + busy_timeout)
def _set_sqlite_pragmas(dbapi_conn, connection_record):
    try:
        cursor = dbapi_conn.cursor()
        cursor.execute("PRAGMA journal_mode=WAL")
        cursor.execute("PRAGMA busy_timeout=30000")
        cursor.execute("PRAGMA synchronous=NORMAL")
        cursor.close()
    except Exception:
        pass  # Ignore pragma errors on malformed/new databases

from sqlalchemy import event
with app.app_context():
    event.listen(db.engine, "connect", _set_sqlite_pragmas)

# ═══════════════════════════════════════════
# THEME STYLES CONFIG
# ═══════════════════════════════════════════
THEME_STYLES = {
    # ── Fallback / generic ──────────────────────────────────────
    'classic': {
        'font_primary': "'DM Sans', 'Inter', -apple-system, system-ui, sans-serif",
        'font_secondary': "'DM Mono', 'Inter', sans-serif",
        'bg_light': '#ffffff', 'bg_dark': '#1a1a1a',
        'surface_light': '#f8f9fa', 'surface_dark': '#2d2d2d',
        'border_light': '#dee2e6', 'border_dark': '#444444',
        'text_light': '#212529', 'text_dark': '#f8f9fa',
        'text_dim_light': '#6c757d', 'text_dim_dark': '#adb5bd',
        'accent': '#007bff', 'accent_hover': '#0056b3',
        'radius': '8px',
    },
    'modern': {
        'font_primary': "'Inter', -apple-system, system-ui, sans-serif",
        'font_secondary': "'DM Mono', monospace",
        'bg_light': '#fafafa', 'bg_dark': '#0a0a0a',
        'surface_light': '#ffffff', 'surface_dark': '#161616',
        'border_light': '#e5e5e5', 'border_dark': '#2a2a2a',
        'text_light': '#0a0a0a', 'text_dark': '#fafafa',
        'text_dim_light': '#737373', 'text_dim_dark': '#a3a3a3',
        'accent': '#111111', 'accent_hover': '#444444',
        'radius': '12px',
    },

    # ── Car  ·  Mạnh mẽ, cơ khí, premium ───────────────────────
    'car': {
        'font_primary': "'Barlow', 'Roboto', -apple-system, sans-serif",
        'font_secondary': "'Barlow Condensed', 'Roboto Condensed', sans-serif",
        'bg_light': '#f4f3f0', 'bg_dark': '#111111',
        'surface_light': '#ffffff', 'surface_dark': '#1c1c1c',
        'border_light': '#ddd9d0', 'border_dark': '#333333',
        'text_light': '#1a1a1a', 'text_dark': '#f0ede8',
        'text_dim_light': '#7a756b', 'text_dim_dark': '#9e9889',
        'accent': '#f39c12', 'accent_hover': '#d68910',
        'radius': '6px',
    },

    # ── Beauty  ·  Sang trọng, nữ tính, thanh lịch ─────────────
    'beauty': {
        'font_primary': "'Playfair Display', 'DM Sans', -apple-system, sans-serif",
        'font_secondary': "'Montserrat', sans-serif",
        'bg_light': '#fdf6f9', 'bg_dark': '#1a0f16',
        'surface_light': '#ffffff', 'surface_dark': '#2a1a24',
        'border_light': '#f2d5e0', 'border_dark': '#4a2d40',
        'text_light': '#2a1a24', 'text_dark': '#fdf6f9',
        'text_dim_light': '#957a8a', 'text_dim_dark': '#c5a3b5',
        'accent': '#e84393', 'accent_hover': '#d63384',
        'radius': '16px',
    },

    # ── Tech  ·  Sắc nét, tối giản, hi-tech ────────────────────
    'tech': {
        'font_primary': "'Space Grotesk', 'Inter', -apple-system, sans-serif",
        'font_secondary': "'JetBrains Mono', 'Fira Code', monospace",
        'bg_light': '#f0f1f5', 'bg_dark': '#09090b',
        'surface_light': '#ffffff', 'surface_dark': '#131318',
        'border_light': '#d4d4dc', 'border_dark': '#27272f',
        'text_light': '#09090b', 'text_dark': '#f0f1f5',
        'text_dim_light': '#62626b', 'text_dim_dark': '#9494a0',
        'accent': '#6c5ce7', 'accent_hover': '#5a4bd1',
        'radius': '4px',
    },

    # ── Pet  ·  Ấm áp, thân thiện, vui nhộn ────────────────────
    'pet': {
        'font_primary': "'Quicksand', 'Nunito', sans-serif",
        'font_secondary': "'Nunito', sans-serif",
        'bg_light': '#fef9f3', 'bg_dark': '#181410',
        'surface_light': '#ffffff', 'surface_dark': '#262018',
        'border_light': '#f0dfc8', 'border_dark': '#443a28',
        'text_light': '#2c2416', 'text_dark': '#fef9f3',
        'text_dim_light': '#8a7a64', 'text_dim_dark': '#b8a88e',
        'accent': '#e17055', 'accent_hover': '#d35a3f',
        'radius': '20px',
    },

    # ── Travel  ·  Phiêu lưu, thoáng đãng, tự do ──────────────
    'travel': {
        'font_primary': "'Outfit', 'Inter', -apple-system, sans-serif",
        'font_secondary': "'DM Sans', 'Inter', sans-serif",
        'bg_light': '#f4f8fc', 'bg_dark': '#0c1218',
        'surface_light': '#ffffff', 'surface_dark': '#161e28',
        'border_light': '#d0dfe8', 'border_dark': '#2a3a4a',
        'text_light': '#0c1218', 'text_dark': '#f4f8fc',
        'text_dim_light': '#5a7a90', 'text_dim_dark': '#7fa0b8',
        'accent': '#0984e3', 'accent_hover': '#0770c2',
        'radius': '12px',
    },

    # ── Bike  ·  Năng động, đô thị, trẻ trung ──────────────────
    'bike': {
        'font_primary': "'Archivo', 'Inter', -apple-system, sans-serif",
        'font_secondary': "'Archivo Narrow', 'Roboto Condensed', sans-serif",
        'bg_light': '#f2f8f8', 'bg_dark': '#0a1214',
        'surface_light': '#ffffff', 'surface_dark': '#141e22',
        'border_light': '#c8e2e2', 'border_dark': '#2a3e42',
        'text_light': '#0a1214', 'text_dark': '#f2f8f8',
        'text_dim_light': '#5a7a7e', 'text_dim_dark': '#7ea8ae',
        'accent': '#00cec9', 'accent_hover': '#00b3af',
        'radius': '8px',
    },

    # ── Sport  ·  Năng lượng, mạnh mẽ, tập trung ───────────────
    'sport': {
        'font_primary': "'Exo 2', 'Barlow', -apple-system, sans-serif",
        'font_secondary': "'Exo 2', 'Roboto', sans-serif",
        'bg_light': '#f3f7f4', 'bg_dark': '#0c110e',
        'surface_light': '#ffffff', 'surface_dark': '#161e1a',
        'border_light': '#c4dcc8', 'border_dark': '#2a3e30',
        'text_light': '#0c110e', 'text_dark': '#f3f7f4',
        'text_dim_light': '#5a7a62', 'text_dim_dark': '#7eaa88',
        'accent': '#00b894', 'accent_hover': '#009b7d',
        'radius': '6px',
    },
}

AVAILABLE_FONTS = [
    ("'Inter', -apple-system, system-ui, sans-serif", "Inter"),
    ("'DM Sans', 'Inter', sans-serif", "DM Sans"),
    ("'DM Mono', monospace", "DM Mono"),
    ("'Barlow', 'Roboto', sans-serif", "Barlow"),
    ("'Barlow Condensed', 'Roboto Condensed', sans-serif", "Barlow Condensed"),
    ("'Playfair Display', 'DM Sans', serif", "Playfair Display"),
    ("'Montserrat', sans-serif", "Montserrat"),
    ("'Space Grotesk', 'Inter', sans-serif", "Space Grotesk"),
    ("'JetBrains Mono', 'Fira Code', monospace", "JetBrains Mono"),
    ("'Quicksand', 'Nunito', sans-serif", "Quicksand"),
    ("'Nunito', sans-serif", "Nunito"),
    ("'Outfit', 'Inter', sans-serif", "Outfit"),
    ("'Archivo', 'Inter', sans-serif", "Archivo"),
    ("'Archivo Narrow', 'Roboto Condensed', sans-serif", "Archivo Narrow"),
    ("'Exo 2', 'Barlow', sans-serif", "Exo 2"),
]

def get_theme_styles():
    """Get theme styles — merge DB custom styles with hardcoded defaults"""
    styles = dict(THEME_STYLES)
    try:
        stored = SiteSettings.get('theme_styles_custom', '')
        if stored:
            custom = json.loads(stored)
            styles.update(custom)
    except:
        pass
    return styles

@app.context_processor
def inject_globals():
    try:
        site_mode = SiteSettings.get('site_mode', 'demo')
        logo_url = SiteSettings.get('logo_url', '')
        favicon_url = SiteSettings.get('favicon_url', '')

        # Voucher sidebar widget
        sidebar_vouchers = []
        vs_enabled = SiteSettings.get('voucher_sidebar_enabled', '1')
        if vs_enabled == '1':
            vs_count = int(SiteSettings.get('voucher_sidebar_count', '4'))
            sidebar_vouchers = Voucher.query.filter_by(
                is_active=True
            ).filter(
                Voucher.valid_to > datetime.utcnow()
            ).order_by(
                Voucher.is_featured.desc(),
                Voucher.discount_value.desc(),
                Voucher.created_at.desc()
            ).limit(vs_count).all()

        # Hot products widget (AccessTrade top products)
        hot_products = []
        hp_enabled = SiteSettings.get('hot_products_enabled', '0')
        hp_show_shop = SiteSettings.get('hot_products_show_shop', '1')
        hp_show_sidebar = SiteSettings.get('hot_products_show_sidebar', '1')
        if hp_enabled == '1':
            try:
                from accesstrade_integration import get_accesstrade_api
                hp_api = get_accesstrade_api()
                if hp_api:
                    hp_count = int(SiteSettings.get('hot_products_count', '8'))
                    result = hp_api.get_top_products()
                    hot_products = (result.get('data') or [])[:hp_count]
            except Exception:
                pass

        # Hot deals banner (uploaded from Excel)
        hotdeals_active = []
        hd_enabled = SiteSettings.get('hotdeal_enabled', '1')
        hd_show_shop = SiteSettings.get('hotdeal_show_shop', '1')
        hd_show_voucher = SiteSettings.get('hotdeal_show_voucher', '1')
        if hd_enabled == '1':
            now_utc = datetime.utcnow()
            hotdeals_active = HotDeal.query.filter(
                HotDeal.is_active == True,
                HotDeal.end_date > now_utc
            ).order_by(HotDeal.end_date.asc()).all()

        # AccessTrade auto-pull banners (dedup: each brand appears once)
        # Brand key = merchant + sub-brand from [...] in offer_name
        # e.g. "[M2store]-Giam 10%..." on SHOPEE → key "shopee_m2store"
        import re as _re
        def _at_brand_key(ab):
            merchant = (ab.merchant or '').strip().lower()
            offer = ab.offer_name or ''
            m = _re.search(r'\[([^\]]+)\]', offer)
            if m:
                return f"{merchant}_{m.group(1).strip().lower()}"
            return merchant

        at_banners_hotdeal = []
        at_banners_sidebar = []
        try:
            now_utc2 = datetime.utcnow()
            at_all = AccessTradeBanner.query.filter(
                AccessTradeBanner.is_active == True
            ).order_by(AccessTradeBanner.synced_at.desc()).all()
            hotdeal_brands = set()
            for ab in at_all:
                if ab.end_date and ab.end_date < now_utc2:
                    continue
                brand_key = _at_brand_key(ab)
                if ab.placement in ('hotdeal', 'both'):
                    if brand_key not in hotdeal_brands:
                        at_banners_hotdeal.append(ab)
                        hotdeal_brands.add(brand_key)
            for ab in at_all:
                if ab.end_date and ab.end_date < now_utc2:
                    continue
                brand_key = _at_brand_key(ab)
                if ab.placement in ('sidebar', 'both'):
                    if brand_key not in hotdeal_brands:
                        at_banners_sidebar.append(ab)
        except Exception:
            pass

        _now = datetime.utcnow()
        return {
            'sidebar_verticals': Vertical.query.order_by(Vertical.name).all(),
            'now': _now,
            'current_month': _now.month,
            'current_year': _now.year,
            'THEME_STYLES': get_theme_styles(),
            'site_mode': site_mode,
            'site_logo_url': logo_url,
            'site_favicon_url': favicon_url,
            'sidebar_vouchers': sidebar_vouchers,
            'voucher_sidebar_position': SiteSettings.get('voucher_sidebar_position', 'after_popular'),
            'custom_head_code': SiteSettings.get('custom_head_code', ''),
            'hot_products': hot_products,
            'hot_products_show_shop': hp_show_shop,
            'hot_products_show_sidebar': hp_show_sidebar,
            'hotdeals_active': hotdeals_active,
            'hotdeal_show_shop': hd_show_shop,
            'hotdeal_show_voucher': hd_show_voucher,
            'at_banners_hotdeal': at_banners_hotdeal,
            'at_banners_sidebar': at_banners_sidebar,
        }
    except:
        _fallback_now = datetime.utcnow()
        return {'sidebar_verticals': [], 'now': _fallback_now, 'current_month': _fallback_now.month, 'current_year': _fallback_now.year, 'THEME_STYLES': THEME_STYLES, 'site_mode': 'demo', 'site_logo_url': '', 'site_favicon_url': '', 'sidebar_vouchers': [], 'voucher_sidebar_position': 'after_popular', 'custom_head_code': '', 'hot_products': [], 'hot_products_show_shop': '1', 'hot_products_show_sidebar': '1', 'hotdeals_active': [], 'hotdeal_show_shop': '1', 'hotdeal_show_voucher': '1', 'at_banners_hotdeal': [], 'at_banners_sidebar': []}

def slugify(text):
    """Convert Vietnamese text to URL-friendly slug (no diacritics)"""
    import re, unicodedata

    # Vietnamese character mapping
    vietnamese_map = {
        'à': 'a', 'á': 'a', 'ạ': 'a', 'ả': 'a', 'ã': 'a',
        'â': 'a', 'ầ': 'a', 'ấ': 'a', 'ậ': 'a', 'ẩ': 'a', 'ẫ': 'a',
        'ă': 'a', 'ằ': 'a', 'ắ': 'a', 'ặ': 'a', 'ẳ': 'a', 'ẵ': 'a',
        'è': 'e', 'é': 'e', 'ẹ': 'e', 'ẻ': 'e', 'ẽ': 'e',
        'ê': 'e', 'ề': 'e', 'ế': 'e', 'ệ': 'e', 'ể': 'e', 'ễ': 'e',
        'ì': 'i', 'í': 'i', 'ĩ': 'i', 'ỉ': 'i', 'ị': 'i',
        'ò': 'o', 'ó': 'o', 'ọ': 'o', 'ỏ': 'o', 'õ': 'o',
        'ô': 'o', 'ồ': 'o', 'ố': 'o', 'ộ': 'o', 'ổ': 'o', 'ỗ': 'o',
        'ơ': 'o', 'ờ': 'o', 'ớ': 'o', 'ợ': 'o', 'ở': 'o', 'ỡ': 'o',
        'ù': 'u', 'ú': 'u', 'ụ': 'u', 'ủ': 'u', 'ũ': 'u',
        'ư': 'u', 'ừ': 'u', 'ứ': 'u', 'ự': 'u', 'ử': 'u', 'ữ': 'u',
        'ỳ': 'y', 'ý': 'y', 'ỵ': 'y', 'ỷ': 'y', 'ỹ': 'y',
        'đ': 'd',
        'À': 'A', 'Á': 'A', 'Ạ': 'A', 'Ả': 'A', 'Ã': 'A',
        'Â': 'A', 'Ầ': 'A', 'Ấ': 'A', 'Ậ': 'A', 'Ẩ': 'A', 'Ẫ': 'A',
        'Ă': 'A', 'Ằ': 'A', 'Ắ': 'A', 'Ặ': 'A', 'Ẳ': 'A', 'Ẵ': 'A',
        'È': 'E', 'É': 'E', 'Ẹ': 'E', 'Ẻ': 'E', 'Ẽ': 'E',
        'Ê': 'E', 'Ề': 'E', 'Ế': 'E', 'Ệ': 'E', 'Ể': 'E', 'Ễ': 'E',
        'Ì': 'I', 'Í': 'I', 'Ĩ': 'I', 'Ỉ': 'I', 'Ị': 'I',
        'Ò': 'O', 'Ó': 'O', 'Ọ': 'O', 'Ỏ': 'O', 'Õ': 'O',
        'Ô': 'O', 'Ồ': 'O', 'Ố': 'O', 'Ộ': 'O', 'Ổ': 'O', 'Ỗ': 'O',
        'Ơ': 'O', 'Ờ': 'O', 'Ớ': 'O', 'Ợ': 'O', 'Ở': 'O', 'Ỡ': 'O',
        'Ù': 'U', 'Ú': 'U', 'Ụ': 'U', 'Ủ': 'U', 'Ũ': 'U',
        'Ư': 'U', 'Ừ': 'U', 'Ứ': 'U', 'Ự': 'U', 'Ử': 'U', 'Ữ': 'U',
        'Ỳ': 'Y', 'Ý': 'Y', 'Ỵ': 'Y', 'Ỷ': 'Y', 'Ỹ': 'Y',
        'Đ': 'D'
    }

    # Replace Vietnamese characters
    for vn_char, latin_char in vietnamese_map.items():
        text = text.replace(vn_char, latin_char)

    # Normalize and remove remaining diacritics
    text = unicodedata.normalize('NFD', text)
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')

    # Remove special characters, keep only alphanumeric, spaces, and hyphens
    text = re.sub(r'[^\w\s-]', '', text.lower())

    # Replace spaces with hyphens
    text = re.sub(r'[-\s]+', '-', text)

    return text.strip('-')

# =============================================
# ADMIN — DASHBOARD
# =============================================
@app.route('/admin')
def admin_dashboard():
    verticals = Vertical.query.all()
    total_parts = Part.query.count()
    total_links = AffiliateLink.query.count()
    total_clicks = db.session.query(db.func.sum(AffiliateLink.clicks)).scalar() or 0
    total_conversions = db.session.query(db.func.sum(AffiliateLink.conversions)).scalar() or 0
    total_ai = AIContent.query.count()
    networks = AffiliateNetwork.query.all()

    # Fake analytics for demo
    demo_traffic = random.randint(1200, 5800)
    demo_revenue = random.randint(500000, 3500000)

    # Check if AccessTrade API is configured (no API calls here)
    from accesstrade_integration import get_accesstrade_api
    api = get_accesstrade_api()
    accesstrade_data = {'connected': api is not None}

    return render_template('admin/dashboard.html',
        verticals=verticals,
        total_parts=total_parts, total_links=total_links,
        total_clicks=total_clicks, total_conversions=total_conversions,
        total_ai=total_ai, networks=networks,
        demo_traffic=demo_traffic, demo_revenue=demo_revenue,
        accesstrade=accesstrade_data)

@app.route('/admin/api/accesstrade-data')
def admin_api_accesstrade_data():
    """Load AccessTrade data async — called via AJAX from dashboard"""
    from accesstrade_integration import get_accesstrade_api
    api = get_accesstrade_api()
    if not api:
        return jsonify({'connected': False})
    try:
        account = api.get_account_info()
        stats = api.get_statistics_summary(days=30)
        campaigns = api.get_campaigns(limit=10)
        offers = api.get_offers(limit=10)
        return jsonify({
            'connected': True,
            'account': account,
            'stats': stats,
            'campaigns': campaigns,
            'offers': offers
        })
    except Exception:
        return jsonify({'connected': False})

# =============================================
# ADMIN — VERTICALS CRUD
# =============================================
@app.route('/admin/verticals')
def admin_verticals():
    verticals = Vertical.query.order_by(Vertical.created_at.desc()).all()
    verticals_data = []
    for v in verticals:
        products_count = AffiliateLink.query.join(Part).join(Zone).join(Segment).filter(Segment.vertical_id == v.id).count()
        articles_count = Article.query.filter_by(vertical_slug=v.slug).count()
        verticals_data.append({
            'vertical': v,
            'products': products_count,
            'articles': articles_count
        })
    return render_template('admin/verticals.html', verticals=verticals, verticals_data=verticals_data)

@app.route('/admin/vertical/new', methods=['GET','POST'])
def admin_vertical_new():
    if request.method == 'POST':
        v = Vertical(
            name=request.form['name'], slug=slugify(request.form['name']),
            icon=request.form.get('icon',''), color=request.form.get('color','#6c5ce7'),
            description=request.form.get('description',''), status='draft',
            template=request.form.get('template','general'),
            style=request.form.get('style','classic'),
            shop_link=request.form.get('shop_link','')
        )
        db.session.add(v)
        db.session.commit()
        flash(f'Da tao: {v.name}', 'success')
        return redirect(url_for('admin_vertical_detail', vid=v.id))
    return render_template('admin/vertical_form.html', vertical=None)

@app.route('/admin/vertical/<int:vid>')
def admin_vertical_detail(vid):
    v = Vertical.query.get_or_404(vid)
    return render_template('admin/vertical_detail.html', vertical=v)

@app.route('/admin/vertical/<int:vid>/edit', methods=['GET','POST'])
def admin_vertical_edit(vid):
    v = Vertical.query.get_or_404(vid)
    if request.method == 'POST':
        v.name = request.form['name']
        v.icon = request.form.get('icon','')
        v.color = request.form.get('color','#6c5ce7')
        v.description = request.form.get('description','')
        v.default_mode = request.form.get('default_mode','minimal')
        v.template = request.form.get('template','general')
        v.style = request.form.get('style','classic')
        v.shop_link = request.form.get('shop_link','')
        db.session.commit()
        flash(f'Da cap nhat: {v.name}', 'success')
        return redirect(url_for('admin_vertical_detail', vid=v.id))
    return render_template('admin/vertical_form.html', vertical=v)

@app.route('/admin/vertical/<int:vid>/toggle', methods=['POST'])
def admin_vertical_toggle(vid):
    v = Vertical.query.get_or_404(vid)
    v.status = 'live' if v.status == 'draft' else 'draft'
    db.session.commit()
    flash(f'{v.name} -> {v.status.upper()}', 'success')
    return redirect(request.referrer or url_for('admin_verticals'))

@app.route('/admin/vertical/<int:vid>/delete', methods=['POST'])
def admin_vertical_delete(vid):
    v = Vertical.query.get_or_404(vid)
    db.session.delete(v)
    db.session.commit()
    flash(f'Da xoa: {v.name}', 'success')
    return redirect(url_for('admin_verticals'))

# SEGMENT CRUD
@app.route('/admin/segment/<int:sid>')
def admin_segment_detail(sid):
    s = Segment.query.get_or_404(sid)
    return render_template('admin/segment_detail.html', segment=s)

@app.route('/admin/segment/new/<int:vid>', methods=['GET','POST'])
def admin_segment_new(vid):
    v = Vertical.query.get_or_404(vid)
    if request.method == 'POST':
        s = Segment(vertical_id=v.id, name=request.form['name'], slug=slugify(request.form['name']),
                    icon=request.form.get('icon',''), description=request.form.get('description',''))
        db.session.add(s)
        db.session.commit()
        flash(f'Da tao segment: {s.name}', 'success')
        return redirect(url_for('admin_vertical_detail', vid=v.id))
    return render_template('admin/segment_form.html', vertical=v, segment=None)

@app.route('/admin/segment/<int:sid>/edit', methods=['GET','POST'])
def admin_segment_edit(sid):
    s = Segment.query.get_or_404(sid)
    if request.method == 'POST':
        s.name = request.form['name']
        s.icon = request.form.get('icon','')
        s.description = request.form.get('description','')
        db.session.commit()
        flash(f'Da cap nhat: {s.name}', 'success')
        return redirect(url_for('admin_segment_detail', sid=s.id))
    return render_template('admin/segment_form.html', vertical=s.vertical, segment=s)

@app.route('/admin/segment/<int:sid>/delete', methods=['POST'])
def admin_segment_delete(sid):
    s = Segment.query.get_or_404(sid)
    vid = s.vertical_id
    db.session.delete(s)
    db.session.commit()
    flash(f'Da xoa segment: {s.name}', 'success')
    return redirect(url_for('admin_vertical_detail', vid=vid))

# ZONE CRUD
@app.route('/admin/zone/<int:zid>')
def admin_zone_detail(zid):
    z = Zone.query.get_or_404(zid)
    return render_template('admin/zone_detail.html', zone=z)

@app.route('/admin/zone/new/<int:sid>', methods=['GET','POST'])
def admin_zone_new(sid):
    s = Segment.query.get_or_404(sid)
    if request.method == 'POST':
        z = Zone(segment_id=s.id, name=request.form['name'], slug=slugify(request.form['name']),
                 icon=request.form.get('icon',''), color=request.form.get('color','#fdcb6e'),
                 description=request.form.get('description',''))
        db.session.add(z)
        db.session.commit()
        flash(f'Da tao zone: {z.name}', 'success')
        return redirect(url_for('admin_segment_detail', sid=s.id))
    return render_template('admin/zone_form.html', segment=s, zone=None)

@app.route('/admin/zone/<int:zid>/edit', methods=['GET','POST'])
def admin_zone_edit(zid):
    z = Zone.query.get_or_404(zid)
    if request.method == 'POST':
        z.name = request.form['name']
        z.icon = request.form.get('icon','')
        z.color = request.form.get('color','#fdcb6e')
        z.description = request.form.get('description','')
        db.session.commit()
        flash(f'Da cap nhat: {z.name}', 'success')
        return redirect(url_for('admin_zone_detail', zid=z.id))
    return render_template('admin/zone_form.html', segment=z.segment, zone=z)

@app.route('/admin/zone/<int:zid>/delete', methods=['POST'])
def admin_zone_delete(zid):
    z = Zone.query.get_or_404(zid)
    sid = z.segment_id
    db.session.delete(z)
    db.session.commit()
    flash(f'Da xoa zone: {z.name}', 'success')
    return redirect(url_for('admin_segment_detail', sid=sid))

# PART CRUD
@app.route('/admin/part/new/<int:zid>', methods=['GET','POST'])
def admin_part_new(zid):
    z = Zone.query.get_or_404(zid)
    if request.method == 'POST':
        p = Part(zone_id=z.id, name_vi=request.form['name_vi'], name_en=request.form.get('name_en',''),
                 slug=slugify(request.form['name_vi']), description=request.form.get('description',''),
                 content=request.form.get('content',''), oem_code=request.form.get('oem_code',''),
                 tags=request.form.get('tags',''), auto_category=request.form.get('auto_category',''),
                 embed_code=request.form.get('embed_code',''))
        db.session.add(p)
        db.session.commit()
        flash(f'Da them: {p.name_vi}', 'success')
        return redirect(url_for('admin_zone_detail', zid=z.id))
    return render_template('admin/part_form.html', zone=z, part=None)

@app.route('/admin/part/<int:pid>', methods=['GET','POST'])
def admin_part_edit(pid):
    p = Part.query.get_or_404(pid)
    if request.method == 'POST':
        p.name_vi = request.form['name_vi']
        p.name_en = request.form.get('name_en','')
        p.description = request.form.get('description','')
        p.content = request.form.get('content','')
        p.oem_code = request.form.get('oem_code','')
        p.tags = request.form.get('tags','')
        p.auto_category = request.form.get('auto_category','')
        p.embed_code = request.form.get('embed_code','')
        db.session.commit()
        flash(f'Da cap nhat: {p.name_vi}', 'success')
        return redirect(url_for('admin_zone_detail', zid=p.zone_id))
    return render_template('admin/part_form.html', zone=p.zone, part=p)

@app.route('/admin/part/<int:pid>/delete', methods=['POST'])
def admin_part_delete(pid):
    p = Part.query.get_or_404(pid)
    zid = p.zone_id
    db.session.delete(p)
    db.session.commit()
    flash(f'Da xoa: {p.name_vi}', 'success')
    return redirect(url_for('admin_zone_detail', zid=zid))

@app.route('/admin/part/<int:pid>/add-link', methods=['POST'])
def admin_add_link(pid):
    p = Part.query.get_or_404(pid)
    al = AffiliateLink(part_id=p.id, network=request.form['network'],
        product_name=request.form.get('product_name',''), url=request.form['url'],
        price=float(request.form.get('price',0)))
    db.session.add(al)
    db.session.commit()
    flash('Da them affiliate link', 'success')
    return redirect(url_for('admin_part_edit', pid=p.id))

@app.route('/admin/link/<int:lid>/delete', methods=['POST'])
def admin_delete_link(lid):
    al = AffiliateLink.query.get_or_404(lid)
    pid = al.part_id
    db.session.delete(al)
    db.session.commit()
    flash('Da xoa link', 'success')
    return redirect(url_for('admin_part_edit', pid=pid))

# =============================================
# ADMIN — MONETIZATION HUB (unified)
# =============================================
@app.route('/admin/monetization')
def admin_monetization():
    """Redirect old monetization URL to products hub"""
    tab = request.args.get('tab', 'products')
    if tab in ('hotels',):
        return redirect(url_for('admin_hotels_hub'))
    elif tab in ('vouchers',):
        return redirect(url_for('admin_vouchers_hub'))
    return redirect(url_for('admin_products_hub'))


@app.route('/admin/products-hub')
def admin_products_hub():
    """Products Hub — Products, Affiliate"""
    from sqlalchemy import func as sqlfunc
    tab = request.args.get('tab', 'products')
    ctx = {'active_tab': tab}

    if tab == 'products':
        f_vertical = request.args.get('vertical', '')
        f_network = request.args.get('network', '')
        f_status = request.args.get('status', '')
        f_search = request.args.get('q', '')
        page = request.args.get('page', 1, type=int)
        per_page = 50

        q = db.session.query(AffiliateLink, Part, Zone, Segment, Vertical).join(
            Part, AffiliateLink.part_id == Part.id
        ).join(Zone, Part.zone_id == Zone.id
        ).join(Segment, Zone.segment_id == Segment.id
        ).join(Vertical, Segment.vertical_id == Vertical.id)

        if f_vertical:
            q = q.filter(Vertical.slug == f_vertical)
        if f_network:
            q = q.filter(AffiliateLink.network == f_network)
        if f_status == 'active':
            q = q.filter(AffiliateLink.is_active == True)
        elif f_status == 'inactive':
            q = q.filter(AffiliateLink.is_active == False)
        if f_search:
            q = q.filter(db.or_(
                AffiliateLink.product_name.ilike(f'%{f_search}%'),
                Part.name_vi.ilike(f'%{f_search}%'),
                AffiliateLink.url.ilike(f'%{f_search}%')
            ))

        pagination = q.order_by(AffiliateLink.id.desc()).paginate(page=page, per_page=per_page, error_out=False)
        products = pagination.items
        verticals = Vertical.query.order_by(Vertical.name).all()
        networks = db.session.query(AffiliateLink.network).distinct().all()

        total_all = AffiliateLink.query.count()
        active = AffiliateLink.query.filter_by(is_active=True).count()
        total_clicks = db.session.query(sqlfunc.sum(AffiliateLink.clicks)).scalar() or 0
        total_conv = db.session.query(sqlfunc.sum(AffiliateLink.conversions)).scalar() or 0

        ctx.update(products=products, verticals=verticals, networks=[n[0] for n in networks],
                   f_vertical=f_vertical, f_network=f_network, f_status=f_status, f_search=f_search,
                   total=total_all, active=active, total_clicks=total_clicks, total_conv=total_conv,
                   pagination=pagination, page=page)

    elif tab == 'affiliate':
        networks = db.session.query(
            AffiliateLink.network, sqlfunc.count(AffiliateLink.id),
            sqlfunc.sum(AffiliateLink.clicks), sqlfunc.sum(AffiliateLink.conversions)
        ).group_by(AffiliateLink.network).all()
        net_stats = [{'name': n[0], 'count': n[1], 'clicks': n[2] or 0, 'conv': n[3] or 0} for n in networks]
        total_clicks = sum(n['clicks'] for n in net_stats)
        total_conv = sum(n['conv'] for n in net_stats)
        total_links = sum(n['count'] for n in net_stats)
        ctx.update(net_stats=net_stats, total_clicks=total_clicks, total_conv=total_conv, total_links=total_links)

    elif tab == 'top_products':
        merchant = request.args.get('merchant', '')
        try:
            from accesstrade_integration import get_accesstrade_api
            api = get_accesstrade_api()
            if api:
                result = api.get_top_products(merchant=merchant or None)
                top_products = result.get('data', [])
                top_total = result.get('total', 0)
            else:
                top_products = []
                top_total = 0
                flash('AccessTrade API key not configured', 'error')
        except Exception as e:
            top_products = []
            top_total = 0
            flash(f'Error fetching top products: {str(e)}', 'error')
        ctx.update(top_products=top_products, top_total=top_total, f_merchant=merchant)

    return render_template('admin/products_hub.html', **ctx)


@app.route('/admin/hotels-hub')
def admin_hotels_hub():
    """Hotels Hub — Hotels, Hotel Sync, Attractions"""
    from sqlalchemy import func as sqlfunc
    tab = request.args.get('tab', 'hotels')
    ctx = {'active_tab': tab}

    if tab == 'hotels':
        f_dest = request.args.get('dest', '')
        f_stars = request.args.get('stars', '')
        f_status = request.args.get('status', '')
        f_image = request.args.get('image', '')
        page = request.args.get('page', 1, type=int)
        per_page = 30

        query = Hotel.query
        if f_dest:
            query = query.filter(Hotel.destination == f_dest)
        if f_stars:
            query = query.filter(Hotel.stars == int(f_stars))
        if f_status == 'active':
            query = query.filter(Hotel.is_active == True)
        elif f_status == 'inactive':
            query = query.filter(Hotel.is_active == False)
        if f_image == 'missing':
            query = query.filter(db.or_(Hotel.image_url == '', Hotel.image_url == None))
        elif f_image == 'has':
            query = query.filter(Hotel.image_url != '', Hotel.image_url != None)

        total_filtered = query.count()
        hotels = query.order_by(Hotel.is_featured.desc(), Hotel.rating.desc()).offset((page - 1) * per_page).limit(per_page).all()
        destinations = db.session.query(Hotel.destination, Hotel.destination_name).distinct().all()

        total = Hotel.query.count()
        active_h = Hotel.query.filter_by(is_active=True).count()
        no_image = Hotel.query.filter(db.or_(Hotel.image_url == '', Hotel.image_url == None)).count()
        total_clicks = db.session.query(sqlfunc.sum(Hotel.clicks)).scalar() or 0
        total_conv = db.session.query(sqlfunc.sum(Hotel.conversions)).scalar() or 0

        ctx.update(hotels=hotels, destinations=destinations,
                   f_dest=f_dest, f_stars=f_stars, f_status=f_status, f_image=f_image,
                   total=total, active_h=active_h, no_image=no_image,
                   total_clicks=total_clicks, total_conv=total_conv,
                   page=page, per_page=per_page, total_filtered=total_filtered,
                   total_pages=(total_filtered + per_page - 1) // per_page)

    elif tab == 'sync':
        from agoda_integration import get_agoda_api, VIETNAM_DESTINATIONS
        # Auto-populate default Agoda credentials if not saved yet
        _default_cid = '1959245'
        _default_key = '1959245:5669c3b3-2865-4591-ba56-1b02a3c04082'
        if not SiteSettings.get('agoda_cid', ''):
            SiteSettings.set_val('agoda_cid', _default_cid, 'api')
            SiteSettings.set_val('agoda_api_key', _default_key, 'api')
            SiteSettings.set_val('agoda_enabled', '1', 'general')
            import agoda_integration
            agoda_integration._api_instance = None
            db.session.commit()

        api = get_agoda_api()
        api_connected = api is not None
        total_hotels = Hotel.query.count()
        total_agoda = Hotel.query.filter_by(source='agoda_api').count()
        total_manual = Hotel.query.filter(Hotel.source != 'agoda_api').count()
        total_active = Hotel.query.filter_by(is_active=True).count()
        cid = SiteSettings.get('agoda_cid', '')
        has_key = bool(SiteSettings.get('agoda_api_key', ''))
        destinations_sync = [
            {'slug': slug, 'name': name, 'city_id': city_id, 'province_code': ''}
            for name, slug, city_id in VIETNAM_DESTINATIONS
        ]
        agoda_destinations = [
            {'slug': d['slug'], 'name': d['name'], 'city_id': d['city_id']}
            for d in destinations_sync if d['city_id']
        ]
        page_sync = request.args.get('page', 1, type=int)
        per_page_sync = 30
        f_image = request.args.get('image', '')
        rq = Hotel.query.filter_by(source='agoda_api')
        if f_image == 'missing':
            rq = rq.filter(db.or_(Hotel.image_url == '', Hotel.image_url == None))
        elif f_image == 'has':
            rq = rq.filter(Hotel.image_url != '', Hotel.image_url != None)
        agoda_no_image = Hotel.query.filter_by(source='agoda_api').filter(db.or_(Hotel.image_url == '', Hotel.image_url == None)).count()
        recent_total = rq.count()
        recent = rq.order_by(Hotel.id.desc()).offset((page_sync - 1) * per_page_sync).limit(per_page_sync).all()
        recent_pages = (recent_total + per_page_sync - 1) // per_page_sync
        ctx.update(api_connected=api_connected, cid=cid, has_key=has_key,
                   total_hotels=total_hotels, total_agoda=total_agoda,
                   total_manual=total_manual, total_active=total_active,
                   destinations=destinations_sync, agoda_destinations=agoda_destinations,
                   recent=recent, page=page_sync, recent_total=recent_total,
                   recent_pages=recent_pages, per_page=per_page_sync, f_image=f_image,
                   agoda_no_image=agoda_no_image)

    elif tab == 'attractions':
        f_dest = request.args.get('dest', 'all')
        f_cat = request.args.get('cat', 'all')

        query = Attraction.query
        if f_dest != 'all':
            query = query.filter_by(destination=f_dest)
        if f_cat != 'all':
            query = query.filter_by(category=f_cat)

        items = query.order_by(Attraction.id.desc()).all()
        destinations = db.session.query(Attraction.destination).distinct().all()
        cats = db.session.query(Attraction.category).distinct().all()
        total = Attraction.query.count()
        active_a = Attraction.query.filter_by(is_active=True).count()

        ctx.update(items=items, destinations=[d[0] for d in destinations if d[0]],
                   cats=[c[0] for c in cats if c[0]],
                   f_dest=f_dest, f_cat=f_cat, total=total, active_a=active_a)

    return render_template('admin/hotels_hub.html', **ctx)


@app.route('/admin/vouchers-hub')
def admin_vouchers_hub():
    """Standalone Voucher Hub with sub-tabs: vouchers, widgets, sync"""
    tab = request.args.get('tab', 'vouchers')
    ctx = dict(active_tab=tab)

    if tab == 'vouchers':
        from sqlalchemy import func as sqlfunc
        f_cat = request.args.get('cat', 'all')
        f_merchant = request.args.get('merchant', 'all')
        f_status = request.args.get('status', 'all')
        query = Voucher.query
        if f_cat != 'all':
            query = query.filter_by(category=f_cat)
        if f_merchant != 'all':
            query = query.filter_by(merchant=f_merchant)
        if f_status == 'active':
            query = query.filter_by(is_active=True)
        elif f_status == 'inactive':
            query = query.filter_by(is_active=False)
        items = query.order_by(Voucher.created_at.desc()).all()
        merchants_q = db.session.query(Voucher.merchant).distinct().all()
        cats_q = db.session.query(Voucher.category).distinct().all()
        ctx.update(items=items, merchants=[m[0] for m in merchants_q if m[0]],
                   cats=[c[0] for c in cats_q if c[0]],
                   f_cat=f_cat, f_merchant=f_merchant, f_status=f_status,
                   v_total=Voucher.query.count(),
                   v_active=Voucher.query.filter_by(is_active=True).count(),
                   v_clicks=db.session.query(sqlfunc.sum(Voucher.clicks)).scalar() or 0)

    elif tab == 'widgets':
        ctx.update(widgets=VoucherWidget.query.all())

    elif tab == 'sync':
        from accesstrade_integration import get_accesstrade_api
        api = get_accesstrade_api()
        api_connected = api is not None
        sync_total_synced = Voucher.query.filter_by(sync_mode='api').count()
        try:
            sync_total_active = _voucher_valid_filter(Voucher.query.filter_by(sync_mode='api')).count()
        except Exception:
            sync_total_active = Voucher.query.filter_by(sync_mode='api', is_active=True).count()
        ctx.update(
            api_connected=api_connected,
            sync_total_synced=sync_total_synced,
            sync_total_active=sync_total_active,
            sync_total_manual=Voucher.query.filter_by(sync_mode='manual').count(),
            auto_sync_enabled=SiteSettings.get('voucher_auto_sync', 'off') == 'on',
            sync_interval=SiteSettings.get('voucher_sync_interval', '6'),
            last_sync=SiteSettings.get('voucher_last_sync', ''),
            last_sync_count=SiteSettings.get('voucher_last_sync_count', '0'),
            last_sync_error=SiteSettings.get('voucher_last_sync_error', ''),
            recent_synced=Voucher.query.filter_by(sync_mode='api').order_by(Voucher.created_at.desc()).limit(20).all())

    elif tab == 'hotdeals':
        deals = HotDeal.query.order_by(HotDeal.created_at.desc()).all()
        ctx.update(deals=deals)

    elif tab == 'banners':
        at_banners = AccessTradeBanner.query.order_by(AccessTradeBanner.created_at.desc()).all()
        ctx.update(
            at_banners=at_banners,
            banner_auto_sync=SiteSettings.get('banner_auto_sync', 'on') == 'on',
            banner_sync_time=SiteSettings.get('banner_sync_time', '03:00'),
            banner_last_auto_sync=SiteSettings.get('banner_last_auto_sync', ''),
            banner_last_auto_sync_result=SiteSettings.get('banner_last_auto_sync_result', ''),
        )

    return render_template('admin/voucher_hub.html', **ctx)

# =============================================
# ADMIN — AFFILIATE HUB
# =============================================
@app.route('/admin/affiliate')
def admin_affiliate():
    networks = AffiliateNetwork.query.all()
    total_clicks = db.session.query(db.func.sum(AffiliateLink.clicks)).scalar() or 0
    total_conv = db.session.query(db.func.sum(AffiliateLink.conversions)).scalar() or 0
    total_links = AffiliateLink.query.count()
    return render_template('admin/affiliate.html', networks=networks,
        total_clicks=total_clicks, total_conv=total_conv, total_links=total_links)

@app.route('/admin/affiliate/network/<int:nid>')
def admin_affiliate_network(nid):
    n = AffiliateNetwork.query.get_or_404(nid)
    extra = {}
    if n.slug == 'agoda':
        extra = {
            'site_id': SiteSettings.get('agoda_site_id', ''),
            'cid': SiteSettings.get('agoda_cid', ''),
            'enabled': SiteSettings.get('agoda_enabled', '0') == '1',
        }
    return render_template('admin/affiliate_network.html', network=n, extra=extra)

@app.route('/admin/affiliate/network/<int:nid>/connect', methods=['POST'])
def admin_affiliate_connect(nid):
    n = AffiliateNetwork.query.get_or_404(nid)
    n.api_key = request.form.get('api_key','')
    n.status = 'connected' if n.api_key else 'disconnected'
    n.last_sync = datetime.utcnow()
    # Handle Agoda extra fields
    if n.slug == 'agoda':
        SiteSettings.set_val('agoda_api_key', n.api_key, 'api')
        SiteSettings.set_val('agoda_site_id', request.form.get('site_id', ''), 'api')
        SiteSettings.set_val('agoda_cid', request.form.get('cid', ''), 'api')
        enabled = '1' if request.form.get('enabled') else '0'
        SiteSettings.set_val('agoda_enabled', enabled, 'general')
        n.status = 'connected' if enabled == '1' and n.api_key else 'disconnected'
    db.session.commit()
    flash(f'{n.name} -> {"Connected" if n.status=="connected" else "Disconnected"}', 'success')
    return redirect(url_for('admin_affiliate_network', nid=n.id))

@app.route('/admin/affiliate/network/<int:nid>/sync', methods=['POST'])
def admin_affiliate_sync(nid):
    """Sync campaigns from affiliate network API"""
    n = AffiliateNetwork.query.get_or_404(nid)
    if not n.api_key:
        flash('API key required to sync', 'error')
        return redirect(url_for('admin_affiliate_network', nid=n.id))

    try:
        import requests as req
        synced = 0

        if n.slug == 'accesstrade':
            headers = {'Authorization': f'Token {n.api_key}', 'Content-Type': 'application/json'}
            cat_map = {'26':'Insurance','29':'Finance','35':'Banking','59':'E-Commerce',
                       '60':'Retail','63':'Travel','65':'Education','66':'Services',
                       '67':'Food & Beverage','68':'Beauty & Health','69':'Technology','71':'Telecom'}
            # Fetch all pages of campaigns
            all_campaigns = []
            page = 1
            while True:
                r = req.get(f'https://api.accesstrade.vn/v1/campaigns?page={page}', headers=headers, timeout=30)
                if r.status_code != 200:
                    break
                data = r.json()
                all_campaigns.extend(data.get('data', []))
                total_pages = int(data.get('total_page', 1))
                if page >= total_pages:
                    break
                page += 1

            # Clear old campaigns for this network
            AffiliateCampaign.query.filter_by(network_id=n.id).delete()

            # Insert fresh data
            for c in all_campaigns:
                cat = c.get('category', '')
                if cat.isdigit():
                    cat = cat_map.get(cat, cat)
                sub = c.get('sub_category', '')
                if sub and not sub.isdigit():
                    cat = sub.split(',')[0]
                camp = AffiliateCampaign(
                    network_id=n.id,
                    name=c.get('name', ''),
                    campaign_id_ext=str(c.get('id', '')),
                    commission=str(c.get('max_com', '')),
                    status='active' if c.get('status') == 1 else 'inactive',
                    category=cat,
                    url=c.get('url', '')
                )
                db.session.add(camp)
                synced += 1

            # Also sync offers/vouchers
            import re as re_mod
            try:
                r2 = req.get('https://api.accesstrade.vn/v1/offers_informations', headers=headers, timeout=60)
                if r2.status_code == 200:
                    offers = r2.json().get('data', [])
                    Voucher.query.filter_by(sync_mode='api', network='accesstrade').delete()
                    voucher_count = 0
                    for o in offers:
                        oname = o.get('name', '')
                        coupons = o.get('coupons', [])
                        coupon_code = coupons[0].get('coupon_code', '') if coupons else ''
                        coupon_desc = coupons[0].get('coupon_desc', '') if coupons else ''
                        desc_text = coupon_desc or oname
                        d_type, d_val, d_min, d_max = 'percentage', 0, 0, 0
                        pct = re_mod.search(r'Giảm (\d+)%', desc_text)
                        if pct:
                            d_val = float(pct.group(1))
                        fix = re_mod.search(r'Giảm ([\d,]+)\s*VNĐ', desc_text)
                        if fix and not pct:
                            d_type, d_val = 'fixed_amount', float(fix.group(1).replace(',', ''))
                        mx = re_mod.search(r'tối đa ([\d,]+)\s*VNĐ', desc_text)
                        if mx:
                            d_max = float(mx.group(1).replace(',', ''))
                        mn = re_mod.search(r'tối thiểu ([\d,]+)\s*VNĐ', desc_text)
                        if mn:
                            d_min = float(mn.group(1).replace(',', ''))
                        merchant = o.get('merchant', 'shopee')
                        bm = re_mod.match(r'\[(.+?)\]', oname)
                        if bm:
                            merchant = bm.group(1)
                        st = o.get('start_time', '')
                        et = o.get('end_time', '')
                        try:
                            vf = datetime.strptime(st, '%Y-%m-%d') if st else datetime.utcnow()
                            vt = datetime.strptime(et, '%Y-%m-%d') if et else datetime.utcnow()
                        except:
                            vf = vt = datetime.utcnow()
                        v = Voucher(code=coupon_code or o.get('id', ''), title=oname, description=coupon_desc,
                                    merchant=merchant, category='shopping', discount_type=d_type,
                                    discount_value=d_val, min_order=d_min, max_discount=d_max,
                                    valid_from=vf, valid_to=vt, network='accesstrade',
                                    affiliate_url=o.get('aff_link', ''), image_url=o.get('image', ''),
                                    is_active=True, sync_mode='api')
                        db.session.add(v)
                        voucher_count += 1
                    synced += voucher_count
            except:
                pass  # Voucher sync is optional, don't fail the whole sync

        n.last_sync = datetime.utcnow()
        db.session.commit()
        flash(f'Synced {synced} campaigns from {n.name}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Sync error: {str(e)}', 'error')

    return redirect(url_for('admin_affiliate_network', nid=n.id))

@app.route('/admin/affiliate/performance')
def admin_affiliate_performance():
    networks = AffiliateNetwork.query.all()
    stats = AffiliateStats.query.order_by(AffiliateStats.date.desc()).limit(30).all()
    return render_template('admin/affiliate_performance.html', networks=networks, stats=stats)

@app.route('/admin/affiliate/deeplinks')
def admin_deeplinks():
    """Manage deeplink templates for affiliate networks"""
    networks = AffiliateNetwork.query.all()
    return render_template('admin/deeplinks.html', networks=networks)

@app.route('/admin/affiliate/deeplink/<int:nid>/update', methods=['POST'])
def admin_deeplink_update(nid):
    """Update deeplink template for a network"""
    network = AffiliateNetwork.query.get_or_404(nid)
    network.deeplink_template = request.form.get('deeplink_template', '')
    db.session.commit()
    flash(f'Đã cập nhật deeplink template cho {network.name}', 'success')
    return redirect(url_for('admin_deeplinks'))

# =============================================
# ADMIN — AI CONTENT
# =============================================
@app.route('/admin/content')
def admin_content():
    contents = AIContent.query.order_by(AIContent.created_at.desc()).all()
    total_cost = db.session.query(db.func.sum(AIContent.cost_vnd)).scalar() or 0
    return render_template('admin/content.html', contents=contents, total_cost=total_cost)

@app.route('/admin/content/generate', methods=['GET','POST'])
def admin_content_generate():
    verticals = Vertical.query.all()
    if request.method == 'POST':
        title = request.form['title']
        prompt_text = request.form.get('prompt', '')
        ai_provider = request.form.get('ai_provider', 'openai')
        vertical_slug = request.form.get('vertical_slug', '')

        # Try real AI generation
        result_text = ''
        tokens_used = 0
        cost_vnd = 0
        try:
            from ai_service import generate_article, build_context_custom
            vertical_name = vertical_slug or 'General'
            for v in verticals:
                if v.slug == vertical_slug:
                    vertical_name = v.name
                    break
            context = build_context_custom(
                vertical_name=vertical_name,
                topic=title,
                keywords=prompt_text[:200] if prompt_text else title,
                tone='seo',
                word_count='1200',
            )
            gen = generate_article('custom', context)
            result_text = gen.get('content', '')
            tokens_used = gen.get('tokens_used', 0)
            cost_vnd = gen.get('cost_vnd', 0)
        except Exception as e:
            result_text = f'[AI generation failed: {str(e)}]\n\nPrompt was: {prompt_text}'
            tokens_used = 0
            cost_vnd = 0

        ai = AIContent(
            title=title, content_type=request.form.get('content_type','article'),
            ai_provider=ai_provider, prompt=prompt_text,
            result=result_text,
            status='draft', vertical_slug=vertical_slug,
            cost_tokens=tokens_used, cost_vnd=cost_vnd
        )
        db.session.add(ai)
        db.session.commit()
        flash(f'Da tao content: {ai.title}', 'success')
        return redirect(url_for('admin_content'))
    return render_template('admin/content_generate.html', verticals=verticals)

@app.route('/admin/content/<int:cid>')
def admin_content_detail(cid):
    c = AIContent.query.get_or_404(cid)
    return render_template('admin/content_detail.html', content=c)

@app.route('/admin/content/<int:cid>/delete', methods=['POST'])
def admin_content_delete(cid):
    c = AIContent.query.get_or_404(cid)
    db.session.delete(c)
    db.session.commit()
    flash('Da xoa content', 'success')
    return redirect(url_for('admin_content'))

@app.route('/admin/contents/bulk-delete', methods=['POST'])
def admin_contents_bulk_delete():
    """Bulk delete AI contents by IDs (JSON)"""
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify(ok=False, error='Khong co content nao'), 400
    deleted = AIContent.query.filter(AIContent.id.in_([int(i) for i in ids])).delete(synchronize_session=False)
    db.session.commit()
    return jsonify(ok=True, deleted=deleted)

# =============================================
# ADMIN — TOOLS HUB (unified page)
# =============================================
@app.route('/admin/tools')
def admin_tools():
    """Unified Tools Hub — Analytics, Video, Seed Data in tabs"""
    tab = request.args.get('tab', 'analytics')
    ctx = {'active_tab': tab}

    if tab == 'analytics':
        from sqlalchemy import func as sqlfunc
        verticals = Vertical.query.all()
        vertical_stats = []
        for v in verticals:
            products = db.session.query(AffiliateLink).join(Part).join(Zone).join(Segment).filter(Segment.vertical_id == v.id)
            clicks = sum(p.clicks or 0 for p in products.all())
            conversions = sum(p.conversions or 0 for p in products.all())
            articles = Article.query.filter_by(vertical_slug=v.slug).count()
            views = db.session.query(sqlfunc.sum(Article.views)).filter(Article.vertical_slug == v.slug).scalar() or 0
            vertical_stats.append({
                'vertical': v, 'clicks': clicks, 'conversions': conversions,
                'articles': articles, 'views': views
            })
        totals = {
            'clicks': sum(vs['clicks'] for vs in vertical_stats),
            'conversions': sum(vs['conversions'] for vs in vertical_stats),
            'articles': sum(vs['articles'] for vs in vertical_stats),
            'views': sum(vs['views'] for vs in vertical_stats),
        }
        ctx.update(verticals=verticals, vertical_stats=vertical_stats, totals=totals)

    elif tab == 'video':
        videos = VideoProject.query.order_by(VideoProject.created_at.desc()).limit(50).all()
        channels = SocialChannel.query.all()
        total_views = db.session.query(db.func.sum(VideoPublish.views)).scalar() or 0
        total_published = VideoPublish.query.filter_by(status='published').count()
        ctx.update(videos=videos, channels=channels, total_views=total_views, total_published=total_published)

    elif tab == 'vouchers':
        return redirect(url_for('admin_vouchers_hub'))

    elif tab == 'seed':
        verticals = Vertical.query.all()
        verticals_data = []
        for v in verticals:
            segments = Segment.query.filter_by(vertical_id=v.id).all()
            segs = []
            for s in segments:
                zones = Zone.query.filter_by(segment_id=s.id).all()
                zs = []
                for z in zones:
                    parts = Part.query.filter_by(zone_id=z.id).all()
                    zs.append({'zone': z, 'parts': parts})
                segs.append({'segment': s, 'zones': zs})
            verticals_data.append({'vertical': v, 'segments': segs})
        ctx.update(verticals_data=verticals_data)

    return render_template('admin/tools_hub.html', **ctx)

# =============================================
# ADMIN — ANALYTICS
# =============================================
@app.route('/admin/analytics')
def admin_analytics():
    verticals = Vertical.query.all()

    # Per-vertical stats
    vertical_stats = []
    for v in verticals:
        articles = Article.query.filter_by(vertical_slug=v.slug).all()
        article_count = len(articles)
        total_views = sum(a.views or 0 for a in articles)

        # Count parts via Vertical > Segments > Zones > Parts
        part_count = 0
        link_count = 0
        total_clicks = 0
        total_conversions = 0
        for seg in v.segments:
            for zone in seg.zones:
                part_count += len(zone.parts)
                for part in zone.parts:
                    link_count += len(part.affiliate_links)
                    total_clicks += sum(al.clicks or 0 for al in part.affiliate_links)
                    total_conversions += sum(al.conversions or 0 for al in part.affiliate_links)

        # Article tier breakdown
        tier_nganh = sum(1 for a in articles if a.tier == 'nganh')
        tier_chung = sum(1 for a in articles if a.tier == 'chung')
        tier_chitiet = sum(1 for a in articles if a.tier == 'chi-tiet')

        vertical_stats.append({
            'name': v.name, 'slug': v.slug, 'icon': v.icon, 'color': v.color,
            'articles': article_count, 'views': total_views,
            'parts': part_count, 'links': link_count,
            'clicks': total_clicks, 'conversions': total_conversions,
            'tier_nganh': tier_nganh, 'tier_chung': tier_chung, 'tier_chitiet': tier_chitiet,
        })

    # Totals
    totals = {
        'articles': sum(s['articles'] for s in vertical_stats),
        'views': sum(s['views'] for s in vertical_stats),
        'parts': sum(s['parts'] for s in vertical_stats),
        'links': sum(s['links'] for s in vertical_stats),
        'clicks': sum(s['clicks'] for s in vertical_stats),
        'conversions': sum(s['conversions'] for s in vertical_stats),
    }

    return render_template('admin/analytics.html',
        verticals=verticals, vertical_stats=vertical_stats, totals=totals)

# =============================================
# ADMIN — SETTINGS
# =============================================
@app.route('/admin/settings', methods=['GET','POST'])
def admin_settings():
    tab = request.args.get('tab', 'general')
    if request.method == 'POST':
        tab = request.form.get('_tab', 'general')
        # Only save keys relevant to current tab (avoid wiping other tabs)
        tab_keys = {
            'general': ['site_mode', 'site_name', 'default_mode', 'carousel_product_limit', 'logo_url', 'favicon_url',
                        'voucher_sidebar_enabled', 'voucher_sidebar_count', 'voucher_sidebar_position',
                        'shop_display_mode', 'custom_head_code',
                        'hot_products_enabled', 'hot_products_count',
                        'hot_products_show_shop', 'hot_products_show_sidebar',
                        'hotdeal_enabled', 'hotdeal_show_shop', 'hotdeal_show_voucher',
                        'redirect_404_target'],
            'api': ['openai_key', 'claude_key', 'dalle_key', 'deepl_key'],
        }
        if tab == 'robots':
            robots_path = os.path.join(app.root_path, 'robots.txt')
            content = request.form.get('robots_content', '')
            with open(robots_path, 'w', encoding='utf-8') as f:
                f.write(content)
            flash('robots.txt updated!', 'success')
            return redirect(url_for('admin_settings', tab='robots'))
        keys_to_save = tab_keys.get(tab, [])
        for key in keys_to_save:
            val = request.form.get(key, '')
            cat = 'api' if '_key' in key or '_id' in key or '_cid' in key else 'general'
            SiteSettings.set_val(key, val, cat)
        flash('Settings saved!', 'success')
        return redirect(url_for('admin_settings', tab=tab))
    settings = {s.key: s.value for s in SiteSettings.query.all()}
    styles = get_theme_styles()
    # Determine which are custom (editable) vs default
    custom_raw = SiteSettings.get('theme_styles_custom', '{}')
    try:
        custom_names = set(json.loads(custom_raw).keys())
    except:
        custom_names = set()
    db_info = _get_db_info() if tab == 'database' else None
    robots_content = ''
    if tab == 'robots':
        robots_path = os.path.join(app.root_path, 'robots.txt')
        if os.path.exists(robots_path):
            with open(robots_path, 'r', encoding='utf-8') as f:
                robots_content = f.read()
    return render_template('admin/settings.html', settings=settings, styles=styles,
                           active_tab=tab, custom_names=custom_names,
                           default_names=set(THEME_STYLES.keys()),
                           available_fonts=AVAILABLE_FONTS,
                           db_info=db_info, robots_content=robots_content)

@app.route('/admin/settings/styles', methods=['POST'])
def admin_settings_styles():
    action = request.form.get('action', 'save')
    name = request.form.get('style_name', '').strip().lower().replace(' ', '-')

    if action == 'delete' and name:
        custom = json.loads(SiteSettings.get('theme_styles_custom', '{}'))
        custom.pop(name, None)
        SiteSettings.set_val('theme_styles_custom', json.dumps(custom), 'styles')
        flash(f'Deleted style "{name}"', 'success')
    elif action == 'save' and name:
        style = {
            'font_primary': request.form.get('font_primary', "'Inter', sans-serif"),
            'font_secondary': request.form.get('font_secondary', "'DM Mono', monospace"),
            'bg_light': request.form.get('bg_light', '#ffffff'),
            'bg_dark': request.form.get('bg_dark', '#1a1a1a'),
            'surface_light': request.form.get('surface_light', '#f8f9fa'),
            'surface_dark': request.form.get('surface_dark', '#2d2d2d'),
            'border_light': request.form.get('border_light', '#dee2e6'),
            'border_dark': request.form.get('border_dark', '#444444'),
            'text_light': request.form.get('text_light', '#212529'),
            'text_dark': request.form.get('text_dark', '#f8f9fa'),
            'text_dim_light': request.form.get('text_dim_light', '#6c757d'),
            'text_dim_dark': request.form.get('text_dim_dark', '#adb5bd'),
            'accent': request.form.get('accent', '#007bff'),
            'accent_hover': request.form.get('accent_hover', '#0056b3'),
            'radius': request.form.get('radius', '8px'),
        }
        custom = json.loads(SiteSettings.get('theme_styles_custom', '{}'))
        custom[name] = style
        SiteSettings.set_val('theme_styles_custom', json.dumps(custom), 'styles')
        flash(f'Style "{name}" saved!', 'success')

    return redirect(url_for('admin_settings', tab='styles'))

# ═══════════════════════════════════════════
# DATABASE MANAGEMENT
# ═══════════════════════════════════════════
def _get_db_info():
    """Get database info for the settings page — includes health report."""
    info = {'size_display': '—', 'table_count': 0, 'tables': [], 'backups': [], 'health': None}

    # Health report from db_backup
    try:
        from db_backup import get_health_report
        health = get_health_report(app)
        info['health'] = health
        info['size_display'] = health['size_display']
        info['backups'] = health['backups']
    except:
        db_path = os.path.join(app.instance_path, 'unilab.db')
        if os.path.exists(db_path):
            size_bytes = os.path.getsize(db_path)
            if size_bytes < 1024:
                info['size_display'] = f'{size_bytes} B'
            elif size_bytes < 1024*1024:
                info['size_display'] = f'{size_bytes/1024:.1f} KB'
            else:
                info['size_display'] = f'{size_bytes/(1024*1024):.1f} MB'

    # Get all model table names from SQLAlchemy metadata
    model_tables = sorted(db.metadata.tables.keys())

    # Check each table
    for tname in model_tables:
        tinfo = {'name': tname, 'columns': 0, 'rows': 0, 'status': 'missing'}
        try:
            cols = _get_table_columns(tname)
            if cols:
                tinfo['columns'] = len(cols)
                tinfo['status'] = 'ok'
                row_count = db.session.execute(db.text(f"SELECT COUNT(*) FROM [{tname}]")).scalar()
                tinfo['rows'] = row_count or 0
        except Exception:
            tinfo['status'] = 'error'
        info['tables'].append(tinfo)

    info['table_count'] = sum(1 for t in info['tables'] if t['status'] == 'ok')

    return info

@app.route('/admin/settings/database', methods=['POST'])
def admin_settings_database():
    action = request.form.get('action', '')

    if action == 'update_schema':
        try:
            db.create_all()
            _run_schema_migration()
            flash('Schema da duoc cap nhat thanh cong!', 'success')
        except Exception as e:
            flash(f'Loi cap nhat schema: {e}', 'error')

    elif action == 'backup':
        from db_backup import create_backup
        result = create_backup(app, label='manual')
        if result:
            flash(f'Backup thanh cong: {result["name"]} ({result["size_display"]})', 'success')
        else:
            flash('Backup that bai!', 'error')

    elif action == 'integrity_check':
        from db_backup import check_integrity
        result = check_integrity(app)
        if result['ok']:
            flash(f'Database HEALTHY ({result["duration_ms"]}ms)', 'success')
        else:
            flash(f'Database co LOI: {"; ".join(result["errors"][:3])}', 'error')

    elif action == 'repair':
        from db_backup import repair_database, create_backup
        # Always backup before repair
        create_backup(app, label='pre_repair')
        result = repair_database(app)
        if result['success']:
            method = result['method']
            tables = result['tables_recovered']
            rows = result['rows_recovered']
            if method == 'wal_cleanup':
                flash('Sua chua thanh cong! Da xoa WAL/SHM files.', 'success')
            else:
                flash(f'Sua chua thanh cong qua {method}! Phuc hoi {tables} bang, {rows} dong du lieu.', 'success')
            # Re-run migration after repair
            try:
                db.create_all()
                _run_schema_migration()
            except:
                pass
        else:
            flash(f'Sua chua that bai: {"; ".join(result["errors"][:2])}. Thu restore tu backup.', 'error')

    elif action == 'restore':
        backup_name = request.form.get('backup_name', '')
        from db_backup import restore_from_backup, list_backups
        # Find the backup by name
        backup_path = None
        for b in list_backups(app):
            if b['name'] == backup_name:
                backup_path = b['path']
                break
        if backup_path:
            result = restore_from_backup(app, backup_path)
            if result['success']:
                # Re-run migration after restore
                try:
                    db.create_all()
                    _run_schema_migration()
                except:
                    pass
                flash(result['message'], 'success')
            else:
                flash(result['message'], 'error')
        else:
            flash(f'Khong tim thay backup: {backup_name}', 'error')

    elif action == 'vacuum':
        from db_backup import vacuum_database
        result = vacuum_database(app)
        flash(result['message'], 'success' if result['success'] else 'error')

    elif action == 'reset':
        import shutil
        from db_backup import create_backup
        db_path = os.path.join(app.instance_path, 'unilab.db')
        try:
            # Proper backup before reset
            create_backup(app, label='pre_reset')
            db.session.remove()
            db.engine.dispose()
            if os.path.exists(db_path):
                os.remove(db_path)
                for ext in ['-wal', '-shm']:
                    p = db_path + ext
                    if os.path.exists(p):
                        os.remove(p)
            db.create_all()
            _run_schema_migration()
            flash('Database da duoc reset! Data cu da backup trong thu muc backups/.', 'success')
        except Exception as e:
            flash(f'Loi reset database: {e}', 'error')

    return redirect(url_for('admin_settings', tab='database'))

@app.route('/admin/toggle-mode', methods=['POST'])
def admin_toggle_mode():
    """Quick toggle site mode between demo and live"""
    current = SiteSettings.get('site_mode', 'demo')
    new_mode = 'live' if current == 'demo' else 'demo'
    SiteSettings.set_val('site_mode', new_mode, 'general')
    flash(f'Site mode switched to {new_mode.upper()}', 'success')
    return redirect(request.referrer or url_for('admin_dashboard'))

@app.route('/admin/deployment')
def admin_deployment():
    """Deployment guide and workflow documentation"""
    return render_template('admin/deployment.html')

@app.route('/admin/robots')
def admin_robots():
    """Redirect to Settings > Robots.txt tab"""
    return redirect(url_for('admin_settings', tab='robots'))

# =============================================
# ADMIN — SEO BACKLINK ENGINE
# =============================================
import re as _re
from markupsafe import Markup

def _build_target_url(vertical_slug, target_type, target_slug, segment_slug='', zone_slug=''):
    """Build the frontend URL for a backlink target."""
    if target_type == 'article':
        return f'/{vertical_slug}/bai-viet/{target_slug}/'
    elif target_type == 'zone':
        if segment_slug:
            return f'/{vertical_slug}/{segment_slug}/{zone_slug or target_slug}/'
        return f'/{vertical_slug}/zone/{target_slug}/'
    elif target_type == 'part':
        if segment_slug and zone_slug:
            return f'/{vertical_slug}/{segment_slug}/{zone_slug}/{target_slug}/'
        return f'/{vertical_slug}/part/{target_slug}/'
    return '#'

def _inject_backlinks(html_content, vertical_slug, source_type, source_id, source_slug):
    """Inject internal backlinks into HTML content by replacing keyword matches with links.

    - Only replaces text outside of HTML tags and existing links
    - Respects max_per_page limit for each keyword
    - Sorted by priority (highest first) to link the most important keywords
    - Tracks clicks via BacklinkInstance
    """
    if not html_content:
        return html_content

    # Get active backlink instances for this source
    instances = BacklinkInstance.query.join(BacklinkKeyword).filter(
        BacklinkInstance.source_type == source_type,
        BacklinkInstance.source_id == source_id,
        BacklinkInstance.status == 'active',
        BacklinkKeyword.is_active == True
    ).all()

    if not instances:
        return html_content

    # Group instances by keyword, respect max_per_page
    keyword_data = {}
    for inst in instances:
        kw = inst.keyword
        if kw.id not in keyword_data:
            keyword_data[kw.id] = {
                'keyword': kw.keyword,
                'anchor': inst.anchor_text or kw.anchor_text or kw.keyword,
                'target_type': kw.target_type,
                'target_slug': kw.target_slug,
                'vertical_slug': kw.vertical_slug or vertical_slug,
                'priority': kw.priority,
                'max_per_page': kw.max_per_page,
                'count': 0,
                'instance_id': inst.id
            }

    # Sort by priority (highest first), then by keyword length (longest first for better matching)
    sorted_kw = sorted(keyword_data.values(), key=lambda x: (-x['priority'], -len(x['keyword'])))

    for kw_data in sorted_kw:
        keyword = kw_data['keyword']
        max_links = kw_data['max_per_page']

        # Build target URL
        target_url = _build_target_url(kw_data['vertical_slug'], kw_data['target_type'], kw_data['target_slug'])
        anchor = kw_data['anchor']
        inst_id = kw_data['instance_id']

        # Build the link HTML
        link_html = f'<a href="{target_url}" class="seo-backlink" data-bid="{inst_id}" title="{anchor}">{anchor}</a>'

        # Replace keyword in text content only (not inside HTML tags or existing links)
        # Strategy: split by tags, only replace in text nodes
        parts = _re.split(r'(<[^>]+>)', html_content)
        in_link = 0
        replaced = 0
        new_parts = []

        for part in parts:
            if part.startswith('<'):
                # It's a tag
                lower_part = part.lower()
                if lower_part.startswith('<a ') or lower_part.startswith('<a>'):
                    in_link += 1
                elif lower_part.startswith('</a'):
                    in_link = max(0, in_link - 1)
                new_parts.append(part)
            else:
                # It's text content
                if in_link == 0 and replaced < max_links and part.strip():
                    # Case-insensitive replacement, but preserve original case in anchor
                    pattern = _re.compile(_re.escape(keyword), _re.IGNORECASE)
                    match = pattern.search(part)
                    if match:
                        part = part[:match.start()] + link_html + part[match.end():]
                        replaced += 1
                new_parts.append(part)

        html_content = ''.join(new_parts)

    return html_content

@app.template_filter('backlinks')
def backlinks_filter(content, vertical_slug='', source_type='', source_id=0, source_slug=''):
    """Jinja filter to inject backlinks into article content."""
    result = _inject_backlinks(content, vertical_slug, source_type, source_id, source_slug)
    return Markup(result)

@app.route('/admin/seo-hub')
def admin_seo_hub():
    """Unified SEO Hub — Dashboard, Keywords, Instances, Suggestions in tabs"""
    tab = request.args.get('tab', 'dashboard')
    verticals = Vertical.query.order_by(Vertical.name).all()
    ctx = {'active_tab': tab, 'verticals': verticals}

    if tab == 'dashboard':
        from sqlalchemy import func as sqlfunc
        filter_v = request.args.get('vertical', 'all')
        total_keywords = BacklinkKeyword.query.count()
        active_keywords = BacklinkKeyword.query.filter_by(is_active=True).count()
        total_instances = BacklinkInstance.query.count()
        active_instances = BacklinkInstance.query.filter_by(status='active').count()
        total_clicks = db.session.query(sqlfunc.sum(BacklinkInstance.clicks)).scalar() or 0
        v_stats = []
        for v in verticals:
            kw_count = BacklinkKeyword.query.filter_by(vertical_slug=v.slug).count()
            inst_count = BacklinkInstance.query.join(BacklinkKeyword).filter(BacklinkKeyword.vertical_slug == v.slug).count()
            clicks = db.session.query(sqlfunc.sum(BacklinkInstance.clicks)).join(BacklinkKeyword).filter(BacklinkKeyword.vertical_slug == v.slug).scalar() or 0
            v_stats.append({'vertical': v, 'keywords': kw_count, 'instances': inst_count, 'clicks': clicks})
        recent = BacklinkInstance.query.order_by(BacklinkInstance.created_at.desc()).limit(20).all()
        ctx.update(filter_v=filter_v, total_keywords=total_keywords, active_keywords=active_keywords,
                   total_instances=total_instances, active_instances=active_instances,
                   total_clicks=total_clicks, v_stats=v_stats, recent=recent)

    elif tab == 'keywords':
        filter_v = request.args.get('vertical', 'all')
        filter_type = request.args.get('type', 'all')
        q = request.args.get('q', '')
        query = BacklinkKeyword.query
        if filter_v != 'all':
            query = query.filter_by(vertical_slug=filter_v)
        if filter_type != 'all':
            query = query.filter_by(target_type=filter_type)
        if q:
            query = query.filter(BacklinkKeyword.keyword.ilike(f'%{q}%'))
        keywords = query.order_by(BacklinkKeyword.priority.desc()).all()
        ctx.update(keywords=keywords, filter_v=filter_v, filter_type=filter_type, q=q)

    elif tab == 'instances':
        filter_v = request.args.get('vertical', 'all')
        filter_type = request.args.get('type', 'all')
        filter_status = request.args.get('status', 'all')
        query = BacklinkInstance.query.join(BacklinkKeyword)
        if filter_v != 'all':
            query = query.filter(BacklinkKeyword.vertical_slug == filter_v)
        if filter_type != 'all':
            query = query.filter(BacklinkInstance.link_type == filter_type)
        if filter_status != 'all':
            query = query.filter(BacklinkInstance.status == filter_status)
        instances = query.order_by(BacklinkInstance.created_at.desc()).all()
        ctx.update(instances=instances, filter_v=filter_v, filter_type=filter_type, filter_status=filter_status)

    elif tab == 'suggestions':
        filter_v = request.args.get('vertical', 'all')
        query = BacklinkInstance.query.filter_by(link_type='suggest')
        if filter_v != 'all':
            query = query.join(BacklinkKeyword).filter(BacklinkKeyword.vertical_slug == filter_v)
        suggestions = query.all()
        grouped = {}
        for s in suggestions:
            key = f"{s.source_type}:{s.source_id}"
            if key not in grouped:
                grouped[key] = {'source_title': s.source_title, 'source_slug': s.source_slug, 'source_type': s.source_type, 'items': []}
            grouped[key]['items'].append(s)
        ctx.update(suggestions=suggestions, grouped=grouped, filter_v=filter_v)

    elif tab == 'scan':
        pass  # Scan page is mostly client-side with form POSTs

    elif tab == 'auto':
        pass  # Auto generate page is mostly forms

    return render_template('admin/seo_hub.html', **ctx)

@app.route('/admin/seo')
def admin_seo_dashboard():
    """SEO Backlink Dashboard — overview of all backlink keywords & instances."""
    verticals = Vertical.query.order_by(Vertical.name).all()
    filter_v = request.args.get('vertical', '')

    # Stats
    total_keywords = BacklinkKeyword.query.count()
    active_keywords = BacklinkKeyword.query.filter_by(is_active=True).count()
    total_instances = BacklinkInstance.query.count()
    active_instances = BacklinkInstance.query.filter_by(status='active').count()
    total_clicks = db.session.query(db.func.coalesce(db.func.sum(BacklinkInstance.clicks), 0)).scalar()

    # Per-vertical stats
    v_stats = []
    for v in verticals:
        kw_count = BacklinkKeyword.query.filter_by(vertical_slug=v.slug, is_active=True).count()
        inst_count = BacklinkInstance.query.join(BacklinkKeyword).filter(
            BacklinkKeyword.vertical_slug == v.slug,
            BacklinkInstance.status == 'active'
        ).count()
        v_clicks = db.session.query(db.func.coalesce(db.func.sum(BacklinkInstance.clicks), 0)).join(BacklinkKeyword).filter(
            BacklinkKeyword.vertical_slug == v.slug
        ).scalar()
        v_stats.append({'vertical': v, 'keywords': kw_count, 'instances': inst_count, 'clicks': v_clicks})

    # Recent instances
    recent = BacklinkInstance.query.order_by(BacklinkInstance.created_at.desc()).limit(20).all()

    return render_template('admin/seo_dashboard.html',
        verticals=verticals, filter_v=filter_v,
        total_keywords=total_keywords, active_keywords=active_keywords,
        total_instances=total_instances, active_instances=active_instances,
        total_clicks=total_clicks, v_stats=v_stats, recent=recent)

@app.route('/admin/seo/keywords')
def admin_seo_keywords():
    """Manage backlink keywords."""
    filter_v = request.args.get('vertical', '')
    filter_type = request.args.get('type', '')
    q = request.args.get('q', '')

    query = BacklinkKeyword.query
    if filter_v:
        query = query.filter_by(vertical_slug=filter_v)
    if filter_type:
        query = query.filter_by(target_type=filter_type)
    if q:
        query = query.filter(BacklinkKeyword.keyword.ilike(f'%{q}%'))

    keywords = query.order_by(BacklinkKeyword.priority.desc(), BacklinkKeyword.created_at.desc()).all()
    verticals = Vertical.query.order_by(Vertical.name).all()

    return render_template('admin/seo_keywords.html',
        keywords=keywords, verticals=verticals,
        filter_v=filter_v, filter_type=filter_type, q=q)

@app.route('/admin/seo/keyword/new', methods=['GET', 'POST'])
def admin_seo_keyword_new():
    """Create a new backlink keyword."""
    verticals = Vertical.query.order_by(Vertical.name).all()
    if request.method == 'POST':
        kw = BacklinkKeyword(
            vertical_slug=request.form.get('vertical_slug', ''),
            keyword=request.form.get('keyword', '').strip(),
            target_type=request.form.get('target_type', 'article'),
            target_slug=request.form.get('target_slug', '').strip(),
            target_title=request.form.get('target_title', '').strip(),
            anchor_text=request.form.get('anchor_text', '').strip(),
            priority=int(request.form.get('priority', 5)),
            max_per_page=int(request.form.get('max_per_page', 1)),
            is_active=request.form.get('is_active') == 'on'
        )
        db.session.add(kw)
        db.session.commit()
        flash(f'Keyword "{kw.keyword}" created successfully!', 'success')
        return redirect(url_for('admin_seo_keywords'))
    return render_template('admin/seo_keyword_form.html', verticals=verticals, keyword=None)

@app.route('/admin/seo/keyword/<int:kid>/edit', methods=['GET', 'POST'])
def admin_seo_keyword_edit(kid):
    """Edit a backlink keyword."""
    kw = BacklinkKeyword.query.get_or_404(kid)
    verticals = Vertical.query.order_by(Vertical.name).all()
    if request.method == 'POST':
        kw.vertical_slug = request.form.get('vertical_slug', '')
        kw.keyword = request.form.get('keyword', '').strip()
        kw.target_type = request.form.get('target_type', 'article')
        kw.target_slug = request.form.get('target_slug', '').strip()
        kw.target_title = request.form.get('target_title', '').strip()
        kw.anchor_text = request.form.get('anchor_text', '').strip()
        kw.priority = int(request.form.get('priority', 5))
        kw.max_per_page = int(request.form.get('max_per_page', 1))
        kw.is_active = request.form.get('is_active') == 'on'
        db.session.commit()
        flash(f'Keyword "{kw.keyword}" updated!', 'success')
        return redirect(url_for('admin_seo_keywords'))
    return render_template('admin/seo_keyword_form.html', verticals=verticals, keyword=kw)

@app.route('/admin/seo/keyword/<int:kid>/delete', methods=['POST'])
def admin_seo_keyword_delete(kid):
    """Delete a backlink keyword and all its instances."""
    kw = BacklinkKeyword.query.get_or_404(kid)
    name = kw.keyword
    db.session.delete(kw)
    db.session.commit()
    flash(f'Keyword "{name}" deleted.', 'success')
    return redirect(url_for('admin_seo_keywords'))

@app.route('/admin/seo/keyword/<int:kid>/toggle', methods=['POST'])
def admin_seo_keyword_toggle(kid):
    """Toggle active/inactive for a keyword."""
    kw = BacklinkKeyword.query.get_or_404(kid)
    kw.is_active = not kw.is_active
    db.session.commit()
    return redirect(request.referrer or url_for('admin_seo_keywords'))

@app.route('/admin/seo/instances')
def admin_seo_instances():
    """View all backlink instances (where links are placed)."""
    filter_v = request.args.get('vertical', '')
    filter_type = request.args.get('link_type', '')
    filter_status = request.args.get('status', '')

    query = BacklinkInstance.query.join(BacklinkKeyword)
    if filter_v:
        query = query.filter(BacklinkKeyword.vertical_slug == filter_v)
    if filter_type:
        query = query.filter(BacklinkInstance.link_type == filter_type)
    if filter_status:
        query = query.filter(BacklinkInstance.status == filter_status)

    instances = query.order_by(BacklinkInstance.created_at.desc()).all()
    verticals = Vertical.query.order_by(Vertical.name).all()

    return render_template('admin/seo_instances.html',
        instances=instances, verticals=verticals,
        filter_v=filter_v, filter_type=filter_type, filter_status=filter_status)

@app.route('/admin/seo/instance/<int:iid>/remove', methods=['POST'])
def admin_seo_instance_remove(iid):
    """Mark a backlink instance as removed."""
    inst = BacklinkInstance.query.get_or_404(iid)
    inst.status = 'removed'
    db.session.commit()
    flash('Backlink instance removed.', 'success')
    return redirect(request.referrer or url_for('admin_seo_instances'))

@app.route('/admin/seo/instance/<int:iid>/delete', methods=['POST'])
def admin_seo_instance_delete(iid):
    """Delete a backlink instance permanently."""
    inst = BacklinkInstance.query.get_or_404(iid)
    db.session.delete(inst)
    db.session.commit()
    flash('Backlink instance deleted.', 'success')
    return redirect(request.referrer or url_for('admin_seo_instances'))

@app.route('/admin/seo/auto-generate', methods=['GET', 'POST'])
def admin_seo_auto_generate():
    """Auto-generate backlink keywords from existing articles/parts/zones."""
    verticals = Vertical.query.order_by(Vertical.name).all()

    if request.method == 'POST':
        target_vertical = request.form.get('vertical_slug', '')
        gen_from = request.form.get('generate_from', 'articles')  # articles / parts / zones / all
        overwrite = request.form.get('overwrite') == 'on'
        created = 0

        if not target_vertical:
            flash('Please select a vertical.', 'error')
            return redirect(url_for('admin_seo_auto_generate'))

        # Generate from articles
        if gen_from in ('articles', 'all'):
            articles = Article.query.filter_by(vertical_slug=target_vertical, status='published').all()
            for art in articles:
                # Use title as keyword, tags as additional keywords
                keywords_to_add = [art.title]
                if art.tags:
                    keywords_to_add.extend([t.strip() for t in art.tags.split(',') if t.strip()])

                for kw_text in keywords_to_add:
                    if len(kw_text) < 3 or len(kw_text) > 100:
                        continue
                    existing = BacklinkKeyword.query.filter_by(
                        vertical_slug=target_vertical, keyword=kw_text, target_type='article', target_slug=art.slug
                    ).first()
                    if existing and not overwrite:
                        continue
                    if existing and overwrite:
                        existing.is_active = True
                        existing.target_title = art.title
                    else:
                        bk = BacklinkKeyword(
                            vertical_slug=target_vertical, keyword=kw_text,
                            target_type='article', target_slug=art.slug,
                            target_title=art.title, priority=7 if kw_text == art.title else 4
                        )
                        db.session.add(bk)
                        created += 1

        # Generate from parts
        if gen_from in ('parts', 'all'):
            zones = Zone.query.join(Segment).join(Vertical).filter(Vertical.slug == target_vertical).all()
            for z in zones:
                for p in z.parts:
                    keywords_to_add = [p.name_vi]
                    if p.name_en:
                        keywords_to_add.append(p.name_en)
                    if p.tags:
                        keywords_to_add.extend([t.strip() for t in p.tags.split(',') if t.strip()])

                    for kw_text in keywords_to_add:
                        if len(kw_text) < 3 or len(kw_text) > 100:
                            continue
                        existing = BacklinkKeyword.query.filter_by(
                            vertical_slug=target_vertical, keyword=kw_text, target_type='part', target_slug=p.slug
                        ).first()
                        if existing and not overwrite:
                            continue
                        if existing and overwrite:
                            existing.is_active = True
                        else:
                            bk = BacklinkKeyword(
                                vertical_slug=target_vertical, keyword=kw_text,
                                target_type='part', target_slug=p.slug,
                                target_title=p.name_vi, priority=6 if kw_text == p.name_vi else 3
                            )
                            db.session.add(bk)
                            created += 1

        # Generate from zones
        if gen_from in ('zones', 'all'):
            zones = Zone.query.join(Segment).join(Vertical).filter(Vertical.slug == target_vertical).all()
            for z in zones:
                existing = BacklinkKeyword.query.filter_by(
                    vertical_slug=target_vertical, keyword=z.name, target_type='zone', target_slug=z.slug
                ).first()
                if existing and not overwrite:
                    continue
                if existing and overwrite:
                    existing.is_active = True
                else:
                    bk = BacklinkKeyword(
                        vertical_slug=target_vertical, keyword=z.name,
                        target_type='zone', target_slug=z.slug,
                        target_title=z.name, priority=8
                    )
                    db.session.add(bk)
                    created += 1

        db.session.commit()
        flash(f'Auto-generated {created} backlink keywords for "{target_vertical}".', 'success')
        return redirect(url_for('admin_seo_keywords', vertical=target_vertical))

    return render_template('admin/seo_auto_generate.html', verticals=verticals)

@app.route('/admin/seo/scan', methods=['GET', 'POST'])
def admin_seo_scan():
    """Scan articles/parts content and create backlink instances where keywords are found."""
    verticals = Vertical.query.order_by(Vertical.name).all()

    if request.method == 'POST':
        target_vertical = request.form.get('vertical_slug', '')
        scan_scope = request.form.get('scan_scope', 'articles')  # articles / parts / all
        clear_old = request.form.get('clear_old') == 'on'

        if not target_vertical:
            flash('Please select a vertical.', 'error')
            return redirect(url_for('admin_seo_scan'))

        # Get active keywords for this vertical
        keywords = BacklinkKeyword.query.filter_by(vertical_slug=target_vertical, is_active=True)\
            .order_by(BacklinkKeyword.priority.desc()).all()

        if not keywords:
            flash('No active keywords found for this vertical. Generate keywords first.', 'error')
            return redirect(url_for('admin_seo_scan'))

        # Clear old instances if requested
        if clear_old:
            old_ids = [inst.id for inst in BacklinkInstance.query.join(BacklinkKeyword).filter(
                BacklinkKeyword.vertical_slug == target_vertical
            ).all()]
            if old_ids:
                BacklinkInstance.query.filter(BacklinkInstance.id.in_(old_ids)).delete(synchronize_session=False)
                db.session.flush()

        created = 0
        scanned = 0

        # Scan articles
        if scan_scope in ('articles', 'all'):
            articles = Article.query.filter_by(vertical_slug=target_vertical, status='published').all()
            for art in articles:
                scanned += 1
                content_text = (art.content or '') + ' ' + (art.excerpt or '')
                content_lower = content_text.lower()

                for kw in keywords:
                    # Skip self-linking (article linking to itself)
                    if kw.target_type == 'article' and kw.target_slug == art.slug:
                        continue

                    kw_lower = kw.keyword.lower()
                    if kw_lower in content_lower:
                        # Check if instance already exists
                        existing = BacklinkInstance.query.filter_by(
                            keyword_id=kw.id, source_type='article',
                            source_id=art.id, status='active'
                        ).first()
                        if not existing:
                            inst = BacklinkInstance(
                                keyword_id=kw.id,
                                source_type='article', source_id=art.id,
                                source_slug=art.slug, source_title=art.title,
                                target_type=kw.target_type, target_slug=kw.target_slug,
                                link_type='intext',
                                anchor_text=kw.anchor_text or kw.keyword
                            )
                            db.session.add(inst)
                            created += 1

        # Auto-generate inline suggestions (article → article, same vertical)
        if scan_scope in ('articles', 'all'):
            all_articles = Article.query.filter_by(vertical_slug=target_vertical, status='published').all()
            for art in all_articles:
                # Find articles to suggest: same category or same tier, different article
                candidates = [c for c in all_articles if c.id != art.id and (
                    (c.category and c.category == art.category) or
                    (c.related_zone_slug and c.related_zone_slug == art.related_zone_slug) or
                    c.tier == art.tier
                )]
                # Pick up to 3, prioritize same category first
                same_cat = [c for c in candidates if c.category and c.category == art.category]
                same_zone = [c for c in candidates if c.related_zone_slug and c.related_zone_slug == art.related_zone_slug and c not in same_cat]
                rest = [c for c in candidates if c not in same_cat and c not in same_zone]
                picks = (same_cat + same_zone + rest)[:3]

                for target_art in picks:
                    existing = BacklinkInstance.query.filter_by(
                        source_type='article', source_id=art.id,
                        target_type='article', target_slug=target_art.slug,
                        link_type='suggest'
                    ).first()
                    if not existing:
                        # Use or create a generic keyword for this target
                        generic_kw = BacklinkKeyword.query.filter_by(
                            vertical_slug=target_vertical, target_type='article',
                            target_slug=target_art.slug, keyword=target_art.title
                        ).first()
                        if not generic_kw:
                            generic_kw = BacklinkKeyword(
                                vertical_slug=target_vertical, keyword=target_art.title,
                                target_type='article', target_slug=target_art.slug,
                                target_title=target_art.title, priority=5
                            )
                            db.session.add(generic_kw)
                            db.session.flush()

                        inst = BacklinkInstance(
                            keyword_id=generic_kw.id,
                            source_type='article', source_id=art.id,
                            source_slug=art.slug, source_title=art.title,
                            target_type='article', target_slug=target_art.slug,
                            link_type='suggest',
                            anchor_text=target_art.title
                        )
                        db.session.add(inst)
                        created += 1

        # Scan parts
        if scan_scope in ('parts', 'all'):
            zones = Zone.query.join(Segment).join(Vertical).filter(Vertical.slug == target_vertical).all()
            for z in zones:
                for p in z.parts:
                    scanned += 1
                    content_text = (p.content or '') + ' ' + (p.description or '')
                    content_lower = content_text.lower()

                    for kw in keywords:
                        # Skip self-linking
                        if kw.target_type == 'part' and kw.target_slug == p.slug:
                            continue

                        kw_lower = kw.keyword.lower()
                        if kw_lower in content_lower:
                            existing = BacklinkInstance.query.filter_by(
                                keyword_id=kw.id, source_type='part',
                                source_id=p.id, status='active'
                            ).first()
                            if not existing:
                                inst = BacklinkInstance(
                                    keyword_id=kw.id,
                                    source_type='part', source_id=p.id,
                                    source_slug=p.slug, source_title=p.name_vi,
                                    target_type=kw.target_type, target_slug=kw.target_slug,
                                    link_type='intext',
                                    anchor_text=kw.anchor_text or kw.keyword
                                )
                                db.session.add(inst)
                                created += 1

        db.session.commit()
        flash(f'Scan complete! Scanned {scanned} items, created {created} new backlink instances.', 'success')
        return redirect(url_for('admin_seo_instances', vertical=target_vertical))

    return render_template('admin/seo_scan.html', verticals=verticals)

@app.route('/admin/seo/suggestions')
def admin_seo_suggestions():
    """Manage inline article suggestions (quote-style blocks inside articles)."""
    filter_v = request.args.get('vertical', '')
    verticals = Vertical.query.order_by(Vertical.name).all()

    query = BacklinkInstance.query.filter_by(link_type='suggest')
    if filter_v:
        query = query.join(BacklinkKeyword).filter(BacklinkKeyword.vertical_slug == filter_v)

    suggestions = query.order_by(BacklinkInstance.created_at.desc()).all()

    # Group by source article for display
    grouped = {}
    for s in suggestions:
        key = f'{s.source_type}:{s.source_id}'
        if key not in grouped:
            grouped[key] = {'source_title': s.source_title, 'source_slug': s.source_slug, 'source_type': s.source_type, 'items': []}
        grouped[key]['items'].append(s)

    return render_template('admin/seo_suggestions.html',
        suggestions=suggestions, grouped=grouped,
        verticals=verticals, filter_v=filter_v)

@app.route('/admin/seo/suggestion/add', methods=['POST'])
def admin_seo_suggestion_add():
    """Manually add an inline suggestion: source article → target article."""
    vertical_slug = request.form.get('vertical_slug', '')
    source_slug = request.form.get('source_slug', '').strip()
    target_slug = request.form.get('target_slug', '').strip()

    if not source_slug or not target_slug or not vertical_slug:
        flash('All fields are required.', 'error')
        return redirect(url_for('admin_seo_suggestions'))

    source = Article.query.filter_by(slug=source_slug, vertical_slug=vertical_slug).first()
    target = Article.query.filter_by(slug=target_slug, vertical_slug=vertical_slug).first()

    if not source or not target:
        flash('Source or target article not found.', 'error')
        return redirect(url_for('admin_seo_suggestions'))

    if source.id == target.id:
        flash('Cannot suggest an article to itself.', 'error')
        return redirect(url_for('admin_seo_suggestions'))

    # Check duplicate
    existing = BacklinkInstance.query.filter_by(
        source_type='article', source_id=source.id,
        target_type='article', target_slug=target_slug,
        link_type='suggest'
    ).first()
    if existing:
        flash('This suggestion already exists.', 'error')
        return redirect(url_for('admin_seo_suggestions'))

    # Find or create keyword
    kw = BacklinkKeyword.query.filter_by(
        vertical_slug=vertical_slug, target_type='article',
        target_slug=target_slug, keyword=target.title
    ).first()
    if not kw:
        kw = BacklinkKeyword(
            vertical_slug=vertical_slug, keyword=target.title,
            target_type='article', target_slug=target_slug,
            target_title=target.title, priority=5
        )
        db.session.add(kw)
        db.session.flush()

    inst = BacklinkInstance(
        keyword_id=kw.id,
        source_type='article', source_id=source.id,
        source_slug=source.slug, source_title=source.title,
        target_type='article', target_slug=target.slug,
        link_type='suggest',
        anchor_text=target.title
    )
    db.session.add(inst)
    db.session.commit()
    flash(f'Suggestion added: "{source.title[:30]}..." → "{target.title[:30]}..."', 'success')
    return redirect(url_for('admin_seo_suggestions', vertical=vertical_slug))

@app.route('/admin/seo/suggestion/<int:sid>/delete', methods=['POST'])
def admin_seo_suggestion_delete(sid):
    """Delete an inline suggestion."""
    inst = BacklinkInstance.query.filter_by(id=sid, link_type='suggest').first_or_404()
    db.session.delete(inst)
    db.session.commit()
    flash('Suggestion removed.', 'success')
    return redirect(request.referrer or url_for('admin_seo_suggestions'))

@app.route('/admin/seo/suggestion/<int:sid>/toggle', methods=['POST'])
def admin_seo_suggestion_toggle(sid):
    """Toggle active/removed for a suggestion."""
    inst = BacklinkInstance.query.filter_by(id=sid, link_type='suggest').first_or_404()
    inst.status = 'removed' if inst.status == 'active' else 'active'
    db.session.commit()
    return redirect(request.referrer or url_for('admin_seo_suggestions'))

@app.route('/admin/seo/bulk-action', methods=['POST'])
def admin_seo_bulk_action():
    """Bulk actions on keywords or instances."""
    action = request.form.get('action', '')
    item_type = request.form.get('item_type', 'keyword')
    ids = request.form.getlist('ids')

    if not ids:
        flash('No items selected.', 'error')
        return redirect(request.referrer or url_for('admin_seo_dashboard'))

    id_list = [int(i) for i in ids]

    if item_type == 'keyword':
        if action == 'activate':
            BacklinkKeyword.query.filter(BacklinkKeyword.id.in_(id_list)).update({BacklinkKeyword.is_active: True}, synchronize_session=False)
        elif action == 'deactivate':
            BacklinkKeyword.query.filter(BacklinkKeyword.id.in_(id_list)).update({BacklinkKeyword.is_active: False}, synchronize_session=False)
        elif action == 'delete':
            BacklinkKeyword.query.filter(BacklinkKeyword.id.in_(id_list)).delete(synchronize_session=False)
    elif item_type == 'instance':
        if action == 'activate':
            BacklinkInstance.query.filter(BacklinkInstance.id.in_(id_list)).update({BacklinkInstance.status: 'active'}, synchronize_session=False)
        elif action == 'remove':
            BacklinkInstance.query.filter(BacklinkInstance.id.in_(id_list)).update({BacklinkInstance.status: 'removed'}, synchronize_session=False)
        elif action == 'delete':
            BacklinkInstance.query.filter(BacklinkInstance.id.in_(id_list)).delete(synchronize_session=False)

    db.session.commit()
    flash(f'Bulk action "{action}" applied to {len(id_list)} items.', 'success')
    return redirect(request.referrer or url_for('admin_seo_dashboard'))

# =============================================
# ADMIN — SEED DATA
# =============================================
@app.route('/admin/seed-data', methods=['GET', 'POST'])
def admin_seed_data():
    """Seed data management - manually trigger seeding"""
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'seed_car':
            from seed_data import seed, seed_articles, seed_networks, seed_products_pet_travel
            try:
                seed()
                seed_articles()
                seed_networks()
                flash('✅ Car vertical seeded successfully!', 'success')
            except Exception as e:
                flash(f'❌ Error seeding Car: {str(e)}', 'error')

        elif action == 'seed_pet':
            from seed_data import seed_pet, seed_pet_articles, seed_products_pet_travel, seed_pet_v2
            try:
                seed_pet()
                seed_pet_articles()
                seed_pet_v2()
                flash('✅ Pet vertical seeded successfully! (v2 included)', 'success')
            except Exception as e:
                flash(f'❌ Error seeding Pet: {str(e)}', 'error')

        elif action == 'seed_travel':
            from seed_data import seed_travel, seed_travel_articles, seed_hotels, seed_attractions, seed_products_pet_travel
            try:
                seed_travel()
                seed_travel_articles()
                seed_hotels()
                seed_attractions()
                flash('✅ Travel vertical seeded successfully!', 'success')
            except Exception as e:
                flash(f'❌ Error seeding Travel: {str(e)}', 'error')

        elif action == 'seed_bike':
            from seed_data import seed_bike
            try:
                seed_bike()
                flash('✅ Bike vertical seeded successfully!', 'success')
            except Exception as e:
                flash(f'❌ Error seeding Bike: {str(e)}', 'error')

        elif action == 'seed_beauty':
            from seed_data import seed_beauty, seed_beauty_articles, seed_products_beauty_tech
            try:
                seed_beauty()
                seed_beauty_articles()
                seed_products_beauty_tech()
                flash('✅ Beauty vertical seeded successfully!', 'success')
            except Exception as e:
                flash(f'❌ Error seeding Beauty: {str(e)}', 'error')

        elif action == 'seed_tech':
            from seed_data import seed_tech, seed_tech_articles, seed_products_beauty_tech
            try:
                seed_tech()
                seed_tech_articles()
                seed_products_beauty_tech()
                flash('✅ Tech vertical seeded successfully!', 'success')
            except Exception as e:
                flash(f'❌ Error seeding Tech: {str(e)}', 'error')

        elif action == 'seed_sport':
            from seed_data import seed_sport, seed_sport_articles, seed_products_sport
            try:
                seed_sport()
                seed_sport_articles()
                seed_products_sport()
                flash('✅ Sport vertical seeded successfully!', 'success')
            except Exception as e:
                flash(f'❌ Error seeding Sport: {str(e)}', 'error')

        elif action == 'seed_garden':
            from seed_data import seed_garden, seed_garden_articles, seed_products_garden
            try:
                seed_garden()
                seed_garden_articles()
                seed_products_garden()
                flash('✅ Garden vertical seeded successfully!', 'success')
            except Exception as e:
                flash(f'❌ Error seeding Garden: {str(e)}', 'error')

        elif action == 'seed_new_verticals':
            from seed_data import seed_new_verticals
            try:
                seed_new_verticals()
                flash('✅ 30 new verticals seeded successfully!', 'success')
            except Exception as e:
                flash(f'❌ Error seeding new verticals: {str(e)}', 'error')

        elif action.startswith('seed_') and action != 'seed_all' and action != 'seed_new_verticals':
            slug = action.replace('seed_', '', 1)
            from seed_data import seed_vertical_content
            try:
                seed_vertical_content(slug)
                flash(f'✅ {slug.title()} vertical content seeded!', 'success')
            except Exception as e:
                flash(f'❌ Error seeding {slug}: {str(e)}', 'error')

        elif action == 'seed_all':
            from seed_data import (seed, seed_articles, seed_networks, seed_video,
                seed_pet, seed_pet_articles, seed_pet_v2, seed_travel, seed_travel_articles,
                seed_products_pet_travel, seed_hotels, seed_attractions,
                seed_bike, seed_vouchers, seed_beauty, seed_beauty_articles,
                seed_tech, seed_tech_articles, seed_products_beauty_tech,
                seed_sport, seed_sport_articles, seed_products_sport,
                seed_garden, seed_garden_articles, seed_products_garden,
                seed_new_verticals, seed_all_new_verticals_content)
            try:
                seed()
                seed_articles()
                seed_networks()
                seed_pet()
                seed_pet_articles()
                seed_pet_v2()
                seed_travel()
                seed_travel_articles()
                seed_products_pet_travel()
                seed_hotels()
                seed_attractions()
                seed_bike()
                seed_vouchers()
                seed_beauty()
                seed_beauty_articles()
                seed_tech()
                seed_tech_articles()
                seed_products_beauty_tech()
                seed_sport()
                seed_sport_articles()
                seed_products_sport()
                seed_garden()
                seed_garden_articles()
                seed_products_garden()
                seed_new_verticals()
                seed_all_new_verticals_content()
                flash('✅ All verticals seeded successfully!', 'success')
            except Exception as e:
                flash(f'❌ Error seeding all: {str(e)}', 'error')

        return redirect(url_for('admin_seed_data'))

    # Get stats for each vertical
    verticals_data = []
    for v in Vertical.query.all():
        segments_count = len(v.segments)
        zones_count = Zone.query.join(Segment).filter(Segment.vertical_id == v.id).count()
        parts_count = Part.query.join(Zone).join(Segment).filter(Segment.vertical_id == v.id).count()
        articles_count = Article.query.filter_by(vertical_slug=v.slug).count()
        products_count = AffiliateLink.query.join(Part).join(Zone).join(Segment).filter(Segment.vertical_id == v.id).count()

        verticals_data.append({
            'vertical': v,
            'segments': segments_count,
            'zones': zones_count,
            'parts': parts_count,
            'articles': articles_count,
            'products': products_count
        })

    return render_template('admin/seed_data.html', verticals_data=verticals_data)

# =============================================
# ADMIN — AI CHAT ASSISTANT
# =============================================
@app.route('/admin/ai-chat', methods=['POST'])
def admin_ai_chat():
    """AI Assistant endpoint - handles chat messages and executes actions"""
    try:
        data = request.get_json()
        user_message = data.get('message', '')

        if not user_message:
            return jsonify({'error': 'No message provided'}), 400

        # Get OpenAI API key from settings
        openai_key = SiteSettings.get_val('openai_key')
        if not openai_key:
            return jsonify({
                'response': 'OpenAI API key chưa được cấu hình. Vui lòng thêm API key trong Settings.'
            })

        # Prepare system context
        verticals = Vertical.query.all()
        total_articles = Article.query.count()
        total_products = AffiliateLink.query.count()

        system_prompt = f"""You are an AI assistant for the Unilab admin panel. You help administrators manage content, products, and operations.

Current System State:
- Verticals: {', '.join([v.name for v in verticals])}
- Total Articles: {total_articles}
- Total Products: {total_products}

You can help with:
1. Writing articles (provide title, content, tier, vertical)
2. Adding products (provide product info)
3. Creating verticals
4. Analyzing data
5. General admin tasks

When the user requests an action, respond in Vietnamese with:
1. A confirmation message
2. If creating content, provide the structure in JSON format

Keep responses concise and actionable."""

        # Call OpenAI API
        from openai import OpenAI
        client = OpenAI(api_key=openai_key)

        response = client.chat.completions.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_message}
            ],
            max_tokens=1000,
            temperature=0.7
        )

        ai_response = response.choices[0].message.content

        # Parse response for actions
        action = None

        # Check if AI wants to create an article
        if 'tạo bài' in user_message.lower() or 'viết bài' in user_message.lower():
            # Suggest redirecting to article creation page
            action = {
                'type': 'redirect',
                'url': '/admin/article/new'
            }

        return jsonify({
            'response': ai_response,
            'action': action
        })

    except Exception as e:
        return jsonify({
            'response': f'Có lỗi xảy ra: {str(e)}'
        }), 500

# =============================================
# ADMIN — CONTENT HUB (Unified: Articles + Feedbacks + AI Content)
# =============================================
@app.route('/admin/content-hub')
def admin_content_hub():
    """Unified Content Hub — Articles, Feedbacks, AI Content in tabs"""
    tab = request.args.get('tab', 'articles')
    ctx = {'active_tab': tab}

    if tab == 'articles':
        # Copy logic from admin_articles
        vertical_filter = request.args.get('vertical', 'all')
        tier_filter = request.args.get('tier', 'all')
        search_q = request.args.get('q', '')
        sort_col = request.args.get('sort', 'created_at')
        sort_order = request.args.get('order', 'desc')

        query = Article.query
        if vertical_filter != 'all':
            query = query.filter_by(vertical_slug=vertical_filter)
        if tier_filter != 'all':
            query = query.filter_by(tier=tier_filter)
        if search_q:
            query = query.filter(db.or_(Article.title.ilike(f'%{search_q}%'), Article.slug.ilike(f'%{search_q}%')))

        sort_attr = getattr(Article, sort_col, Article.created_at)
        if sort_order == 'asc':
            query = query.order_by(sort_attr.asc())
        else:
            query = query.order_by(sort_attr.desc())

        articles = query.all()
        verticals = Vertical.query.order_by(Vertical.name).all()
        stats = {
            'total': Article.query.count(),
            'nganh': Article.query.filter_by(tier='nganh').count(),
            'chung': Article.query.filter_by(tier='chung').count(),
            'chi_tiet': Article.query.filter_by(tier='chi-tiet').count(),
        }
        ctx.update(articles=articles, verticals=verticals, stats=stats,
                   current_vertical=vertical_filter, current_tier=tier_filter,
                   current_search=search_q, current_sort=sort_col, current_order=sort_order)

    elif tab == 'feedbacks':
        status_filter = request.args.get('status', 'all')
        query = ArticleFeedback.query
        if status_filter != 'all':
            query = query.filter_by(status=status_filter)
        feedbacks = query.order_by(ArticleFeedback.created_at.desc()).all()
        ctx.update(feedbacks=feedbacks, status_filter=status_filter,
                   pending_count=ArticleFeedback.query.filter_by(status='pending').count(),
                   reviewed_count=ArticleFeedback.query.filter_by(status='reviewed').count(),
                   resolved_count=ArticleFeedback.query.filter_by(status='resolved').count(),
                   dismissed_count=ArticleFeedback.query.filter_by(status='dismissed').count())

    elif tab == 'ai_content':
        contents = AIContent.query.order_by(AIContent.created_at.desc()).all()
        from sqlalchemy import func as sqlfunc
        total_cost = db.session.query(sqlfunc.sum(AIContent.cost_vnd)).scalar() or 0
        ctx.update(contents=contents, total_cost=total_cost)

    return render_template('admin/content_hub.html', **ctx)

# =============================================
# ADMIN — ARTICLES (Knowledge Base)
# =============================================
@app.route('/admin/articles')
def admin_articles():
    """Redirect standalone articles page to Content Hub"""
    return redirect(url_for('admin_content_hub', tab='articles'))

@app.route('/admin/article/new', methods=['GET','POST'])
def admin_article_new():
    verticals = Vertical.query.all()
    if request.method == 'POST':
        a = Article(
            vertical_slug=request.form.get('vertical_slug',''),
            title=request.form['title'], slug=slugify(request.form['title']),
            excerpt=request.form.get('excerpt',''), content=request.form.get('content',''),
            image_url=request.form.get('image_url',''),
            tier=request.form.get('tier','chung'), category=request.form.get('category',''),
            tags=request.form.get('tags',''),
            related_segment_slug=request.form.get('related_segment_slug',''),
            related_zone_slug=request.form.get('related_zone_slug',''),
            embed_code=request.form.get('embed_code',''),
            ai_generated='ai_generated' in request.form,
            reading_time=int(request.form.get('reading_time',5)),
            status=request.form.get('status','published')
        )
        db.session.add(a)
        db.session.commit()
        flash(f'Da tao bai viet: {a.title}', 'success')
        return redirect(url_for('admin_content_hub', tab='articles'))
    return render_template('admin/article_form.html', article=None, verticals=verticals)

@app.route('/admin/article/<int:aid>/edit', methods=['GET','POST'])
def admin_article_edit(aid):
    a = Article.query.get_or_404(aid)
    verticals = Vertical.query.all()
    if request.method == 'POST':
        a.title = request.form['title']
        a.excerpt = request.form.get('excerpt','')
        a.content = request.form.get('content','')
        a.image_url = request.form.get('image_url','')
        a.tier = request.form.get('tier','chung')
        a.category = request.form.get('category','')
        a.tags = request.form.get('tags','')
        a.related_segment_slug = request.form.get('related_segment_slug','')
        a.related_zone_slug = request.form.get('related_zone_slug','')
        a.embed_code = request.form.get('embed_code','')
        a.ai_generated = 'ai_generated' in request.form
        a.reading_time = int(request.form.get('reading_time',5))
        a.status = request.form.get('status','published')
        db.session.commit()
        flash(f'Da cap nhat: {a.title}', 'success')
        return redirect(url_for('admin_content_hub', tab='articles'))
    # Related articles for preview in edit form
    related_articles = Article.query.filter(
        Article.id != a.id, Article.status=='published',
        db.or_(Article.vertical_slug==a.vertical_slug),
        db.or_(Article.category==a.category, Article.tier==a.tier)
    ).order_by(Article.views.desc()).limit(6).all()
    return render_template('admin/article_form.html', article=a, verticals=verticals, related_articles=related_articles)

@app.route('/admin/article/<int:aid>/delete', methods=['POST'])
def admin_article_delete(aid):
    a = Article.query.get_or_404(aid)
    db.session.delete(a)
    db.session.commit()
    flash('Da xoa bai viet', 'success')
    return redirect(url_for('admin_content_hub', tab='articles'))

@app.route('/admin/articles/bulk-delete-selected', methods=['POST'])
def admin_articles_bulk_delete_selected():
    """Delete selected articles by IDs (JSON)"""
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify(ok=False, error='Khong co bai viet nao duoc chon'), 400
    deleted = Article.query.filter(Article.id.in_([int(i) for i in ids])).delete(synchronize_session=False)
    db.session.commit()
    return jsonify(ok=True, deleted=deleted)

@app.route('/admin/articles/bulk-delete-all', methods=['POST'])
def admin_articles_bulk_delete_all():
    """Delete all articles matching current filters"""
    data = request.get_json(silent=True) or {}
    vertical = data.get('vertical', '')
    tier = data.get('tier', '')
    search = data.get('search', '')
    q = Article.query
    if vertical:
        q = q.filter_by(vertical_slug=vertical)
    if tier:
        q = q.filter_by(tier=tier)
    if search:
        q = q.filter(db.or_(Article.title.ilike(f'%{search}%'), Article.category.ilike(f'%{search}%'), Article.tags.ilike(f'%{search}%')))
    count = q.count()
    if count == 0:
        return jsonify(ok=False, error='Khong co bai viet nao de xoa'), 400
    q.delete(synchronize_session=False)
    db.session.commit()
    return jsonify(ok=True, deleted=count)

@app.route('/admin/api/zone-products')
def admin_api_zone_products():
    """API: return products from a zone by slug for article product attachment UI"""
    slug = request.args.get('slug', '').strip()
    selected_ids_str = request.args.get('selected', '')
    selected_ids = set()
    if selected_ids_str:
        try:
            selected_ids = {int(x) for x in selected_ids_str.split(',') if x.strip()}
        except (ValueError, TypeError):
            pass
    if not slug:
        return jsonify(products=[])
    z = Zone.query.filter_by(slug=slug).first()
    if not z:
        return jsonify(products=[])
    parts = Part.query.filter_by(zone_id=z.id, status='published').all()
    products = []
    for p in parts:
        for al in p.affiliate_links:
            if al.is_active:
                products.append({
                    'id': al.id,
                    'network': al.network,
                    'product_name': al.product_name,
                    'part_name': p.name_vi,
                    'price': al.price,
                    'image_url': al.image_url,
                    'selected': al.id in selected_ids
                })
    return jsonify(products=products)

# =============================================
# ADMIN — ARTICLE FEEDBACKS
# =============================================
@app.route('/admin/feedbacks')
def admin_feedbacks():
    """Redirect standalone feedbacks page to Content Hub"""
    return redirect(url_for('admin_content_hub', tab='feedbacks'))

@app.route('/admin/feedback/<int:fid>')
def admin_feedback_detail(fid):
    """View feedback detail"""
    feedback = ArticleFeedback.query.get_or_404(fid)
    article = Article.query.get(feedback.article_id)
    return render_template('admin/feedback_detail.html', feedback=feedback, article=article)

@app.route('/admin/feedback/<int:fid>/update-status', methods=['POST'])
def admin_feedback_update_status(fid):
    """Update feedback status"""
    feedback = ArticleFeedback.query.get_or_404(fid)
    feedback.status = request.form.get('status', 'pending')
    feedback.admin_note = request.form.get('admin_note', '')

    if feedback.status in ['reviewed', 'resolved', 'dismissed']:
        feedback.reviewed_at = datetime.utcnow()

    db.session.commit()
    flash('Đã cập nhật trạng thái phản hồi', 'success')
    return redirect(url_for('admin_content_hub', tab='feedbacks'))

@app.route('/admin/feedback/<int:fid>/delete', methods=['POST'])
def admin_feedback_delete(fid):
    """Delete feedback"""
    feedback = ArticleFeedback.query.get_or_404(fid)
    db.session.delete(feedback)
    db.session.commit()
    flash('Đã xóa phản hồi', 'success')
    return redirect(url_for('admin_content_hub', tab='feedbacks'))

@app.route('/admin/feedbacks/bulk-delete', methods=['POST'])
def admin_feedbacks_bulk_delete():
    """Bulk delete feedbacks by IDs (JSON)"""
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify(ok=False, error='Khong co feedback nao'), 400
    deleted = ArticleFeedback.query.filter(ArticleFeedback.id.in_([int(i) for i in ids])).delete(synchronize_session=False)
    db.session.commit()
    return jsonify(ok=True, deleted=deleted)

# =============================================
# ADMIN — PRODUCTS (Quản lý sản phẩm tập trung)
# =============================================
@app.route('/admin/products')
def admin_products():
    """Redirect standalone products page to Products Hub"""
    return redirect(url_for('admin_products_hub'))

@app.route('/admin/product/new', methods=['GET','POST'])
def admin_product_new():
    if request.method == 'POST':
        al = AffiliateLink(
            part_id=int(request.form['part_id']),
            network=request.form['network'],
            product_name=request.form.get('product_name',''),
            url=request.form['url'],
            price=float(request.form.get('price', 0)),
            image_url=request.form.get('image_url',''),
            is_active='is_active' in request.form
        )
        db.session.add(al)
        db.session.commit()
        flash(f'Da them san pham: {al.product_name}', 'success')
        return redirect(url_for('admin_products_hub'))
    # Get all parts grouped by vertical > segment > zone
    parts_tree = []
    for v in Vertical.query.all():
        for s in v.segments:
            for z in s.zones:
                for p in z.parts:
                    parts_tree.append({
                        'id': p.id,
                        'label': f'{v.icon} {v.name} › {s.name} › {z.name} › {p.name_vi}'
                    })
    return render_template('admin/product_form.html', product=None, parts_tree=parts_tree)

@app.route('/admin/product/<int:pid>/edit', methods=['GET','POST'])
def admin_product_edit(pid):
    al = AffiliateLink.query.get_or_404(pid)
    if request.method == 'POST':
        al.part_id = int(request.form['part_id'])
        al.network = request.form['network']
        al.product_name = request.form.get('product_name','')
        al.url = request.form['url']
        al.price = float(request.form.get('price', 0))
        al.image_url = request.form.get('image_url','')
        al.is_active = 'is_active' in request.form
        db.session.commit()
        flash(f'Da cap nhat: {al.product_name}', 'success')
        return redirect(url_for('admin_products_hub'))
    parts_tree = []
    for v in Vertical.query.all():
        for s in v.segments:
            for z in s.zones:
                for p in z.parts:
                    parts_tree.append({
                        'id': p.id,
                        'label': f'{v.icon} {v.name} › {s.name} › {z.name} › {p.name_vi}'
                    })
    return render_template('admin/product_form.html', product=al, parts_tree=parts_tree)

@app.route('/admin/product/<int:pid>/toggle', methods=['POST'])
def admin_product_toggle(pid):
    al = AffiliateLink.query.get_or_404(pid)
    al.is_active = not al.is_active
    db.session.commit()
    flash(f'{"Bat" if al.is_active else "Tat"}: {al.product_name}', 'success')
    return redirect(url_for('admin_products_hub'))

@app.route('/admin/product/<int:pid>/delete', methods=['POST'])
def admin_product_delete(pid):
    al = AffiliateLink.query.get_or_404(pid)
    # Check if product is attached to articles (via zone or embed_code product_ids)
    part = Part.query.get(al.part_id)
    warnings = []
    if part:
        zone = Zone.query.get(part.zone_id)
        if zone:
            linked_articles = Article.query.filter(Article.related_zone_slug==zone.slug, Article.status=='published').all()
            if linked_articles:
                titles = ', '.join([a.title[:40] for a in linked_articles[:3]])
                warnings.append(f'San pham nay thuoc zone "{zone.slug}" dang gan voi {len(linked_articles)} bai viet: {titles}')
    # Check embed_code product_ids references
    embed_articles = Article.query.filter(Article.embed_code.contains(f'product_ids:'), Article.embed_code.contains(str(al.id))).all()
    for ea in embed_articles:
        # Verify the ID is actually in the list
        ec = ea.embed_code or ''
        if ec.startswith('product_ids:'):
            ids = ec.replace('product_ids:', '').split(',')
            if str(al.id) in ids:
                warnings.append(f'San pham dang duoc gan truc tiep trong bai viet: {ea.title[:40]}')
    db.session.delete(al)
    db.session.commit()
    msg = 'Da xoa san pham'
    if warnings:
        msg += ' (Luu y: ' + '; '.join(warnings) + ')'
    flash(msg, 'success' if not warnings else 'warning')
    return redirect(url_for('admin_products_hub'))

@app.route('/admin/products/bulk-delete', methods=['POST'])
def admin_products_bulk_delete():
    """Delete products in bulk: by filter, by network, or all"""
    action = request.form.get('action', '')
    f_vertical = request.form.get('vertical', '')
    f_network = request.form.get('network', '')

    q = AffiliateLink.query
    label = 'tat ca'

    if action == 'delete_filtered':
        # Delete by current filters
        if f_vertical:
            q = q.join(Part).join(Zone).join(Segment).join(Vertical).filter(Vertical.slug == f_vertical)
            label = f'vertical "{f_vertical}"'
        if f_network:
            q = q.filter(AffiliateLink.network == f_network)
            label = f'network "{f_network}"' if not f_vertical else f'{label} + network "{f_network}"'
    elif action == 'delete_all':
        label = 'tat ca'
    else:
        flash('Action khong hop le', 'error')
        return redirect(url_for('admin_products_hub'))

    count = q.count()
    if count == 0:
        flash('Khong co san pham nao de xoa', 'warning')
        return redirect(url_for('admin_products_hub'))

    q.delete(synchronize_session=False)
    db.session.commit()
    flash(f'Da xoa {count:,} san pham ({label})', 'success')
    return redirect(url_for('admin_products_hub'))

@app.route('/admin/products/bulk-delete-selected', methods=['POST'])
def admin_products_bulk_delete_selected():
    """Delete selected products by IDs"""
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify(ok=False, error='No IDs'), 400
    deleted = AffiliateLink.query.filter(AffiliateLink.id.in_([int(i) for i in ids])).delete()
    db.session.commit()
    return jsonify(ok=True, deleted=deleted)

# =============================================
# AI CONTROL CENTER - Unified AI Operations Hub
# =============================================

@app.route('/admin/ai-engine')
def admin_ai_engine():
    """Unified AI Engine — all AI features in one tabbed page"""
    tab = request.args.get('tab', 'pipeline')
    verticals = Vertical.query.filter_by(status='published').order_by(Vertical.name).all()

    ctx = {'active_tab': tab, 'verticals': verticals}

    if tab == 'pipeline':
        # Reuse gap analysis logic
        analysis = []
        for v in verticals:
            segments = Segment.query.filter_by(vertical_id=v.id).all()
            total_zones = 0
            zones_with_content = 0
            empty_zones = []
            for seg in segments:
                zones = Zone.query.filter_by(segment_id=seg.id).all()
                for z in zones:
                    total_zones += 1
                    parts_count = Part.query.filter_by(zone_id=z.id).count()
                    has_seo = bool(z.seo_content and len(z.seo_content.strip()) > 50)
                    if parts_count > 0 or has_seo:
                        zones_with_content += 1
                    else:
                        empty_zones.append({'segment': seg.name, 'zone': z.name, 'zone_id': z.id})
            articles_count = Article.query.filter_by(vertical_slug=v.slug, status='published').count()
            queue_pending = ContentQueue.query.filter_by(vertical_id=v.id, status='pending').count()
            queue_review = ContentQueue.query.filter_by(vertical_id=v.id, status='review').count()
            rule = AutoContentRule.query.filter_by(vertical_id=v.id).first()
            coverage = round(zones_with_content / total_zones * 100) if total_zones > 0 else 0
            analysis.append({
                'vertical': v, 'total_zones': total_zones, 'zones_with_content': zones_with_content,
                'coverage': coverage, 'empty_zones': empty_zones[:5], 'articles_count': articles_count,
                'queue_pending': queue_pending, 'queue_review': queue_review,
                'has_rule': rule is not None and rule.is_active if rule else False,
            })
        queue_items = ContentQueue.query.order_by(ContentQueue.created_at.desc()).limit(50).all()
        ctx.update(analysis=analysis, queue_items=queue_items, queue_stats={
            'pending': ContentQueue.query.filter_by(status='pending').count(),
            'review': ContentQueue.query.filter_by(status='review').count(),
            'published': ContentQueue.query.filter_by(status='published').count(),
            'total': ContentQueue.query.count(),
        })

    elif tab == 'calendar':
        import calendar as cal_mod
        year = request.args.get('year', date.today().year, type=int)
        month = request.args.get('month', date.today().month, type=int)
        first_day = date(year, month, 1)
        days_in_month = cal_mod.monthrange(year, month)[1]
        last_day = date(year, month, days_in_month)
        events = ContentEvent.query.filter(
            ContentEvent.start_date <= last_day,
            db.or_(ContentEvent.end_date >= first_day, ContentEvent.end_date.is_(None)),
            ContentEvent.is_active == True
        ).all()
        cal_queue = ContentQueue.query.filter(
            ContentQueue.scheduled_at >= datetime(year, month, 1),
            ContentQueue.scheduled_at <= datetime(year, month, days_in_month, 23, 59, 59)
        ).all()
        cal = cal_mod.Calendar(firstweekday=0)
        weeks = cal.monthdatescalendar(year, month)
        date_events = {}
        for e in events:
            d = e.start_date
            end = e.end_date or e.start_date
            while d <= end and d <= last_day:
                if d >= first_day:
                    date_events.setdefault(d, []).append({'type': 'event', 'obj': e})
                d += timedelta(days=1)
        for q in cal_queue:
            if q.scheduled_at:
                qd = q.scheduled_at.date()
                date_events.setdefault(qd, []).append({'type': 'queue', 'obj': q})
        prev_month = month - 1 if month > 1 else 12
        prev_year = year if month > 1 else year - 1
        next_month = month + 1 if month < 12 else 1
        next_year = year if month < 12 else year + 1
        ctx.update(year=year, month=month, weeks=weeks, date_events=date_events,
                   events=events, cal_queue=cal_queue,
                   prev_year=prev_year, prev_month=prev_month,
                   next_year=next_year, next_month=next_month,
                   month_name=cal_mod.month_name[month], today=date.today())

    elif tab == 'rules':
        rules = AutoContentRule.query.all()
        ctx.update(rules_map={r.vertical_id: r for r in rules})

    elif tab == 'health':
        # AI Control Center health data
        from sqlalchemy import func
        openai_key = SiteSettings.get('openai_key')
        claude_key = SiteSettings.get('claude_key')
        total_products = AffiliateLink.query.count()
        active_products = AffiliateLink.query.filter_by(is_active=True).count()
        dup_sub = db.session.query(
            AffiliateLink.url, func.count(AffiliateLink.id).label('cnt')
        ).group_by(AffiliateLink.url).having(func.count(AffiliateLink.id) > 1).subquery()
        dup_urls = db.session.query(func.count()).select_from(dup_sub).scalar() or 0
        no_image = AffiliateLink.query.filter(
            db.or_(AffiliateLink.image_url == '', AffiliateLink.image_url == None)
        ).count()
        zero_price = AffiliateLink.query.filter(
            db.or_(AffiliateLink.price == 0, AffiliateLink.price == None)
        ).count()
        stale_products = AffiliateLink.query.filter(AffiliateLink.clicks == 0).count()
        suspect_links = AffiliateLink.query.filter(
            AffiliateLink.clicks > 10, AffiliateLink.conversions == 0
        ).count()
        inactive_products = AffiliateLink.query.filter_by(is_active=False).count()
        total_articles = Article.query.count()
        thin_articles = Article.query.filter(
            db.or_(func.length(Article.content) < 500, Article.content == '', Article.content == None)
        ).count()
        zero_view_articles = Article.query.filter(
            db.or_(Article.views == 0, Article.views == None)
        ).count()
        no_product_articles = Article.query.filter(
            db.or_(Article.embed_code == '', Article.embed_code == None)
        ).count()
        ai_articles = Article.query.filter_by(ai_generated=True).count()
        issues_total = dup_urls + no_image + zero_price + suspect_links + thin_articles
        health_score = max(0, 100 - min(issues_total, 100))

        vert_health = []
        for v in verticals:
            prod_count = db.session.query(AffiliateLink).join(Part).join(Zone).join(Segment).filter(
                Segment.vertical_id == v.id).count()
            art_count = Article.query.filter_by(vertical_slug=v.slug).count()
            part_count = db.session.query(Part).join(Zone).join(Segment).filter(
                Segment.vertical_id == v.id).count()
            parts_no_prod = db.session.query(Part).join(Zone).join(Segment).filter(
                Segment.vertical_id == v.id
            ).outerjoin(AffiliateLink).filter(AffiliateLink.id == None).count()
            vert_health.append({
                'id': v.id, 'name': v.name, 'icon': v.icon, 'slug': v.slug,
                'products': prod_count, 'articles': art_count,
                'parts': part_count, 'parts_empty': parts_no_prod,
                'status': v.status
            })
        networks = db.session.query(
            AffiliateLink.network, func.count(AffiliateLink.id),
            func.sum(AffiliateLink.clicks), func.sum(AffiliateLink.conversions)
        ).group_by(AffiliateLink.network).all()
        net_stats = [{'name': n[0], 'count': n[1], 'clicks': n[2] or 0, 'conv': n[3] or 0} for n in networks]

        ctx.update(has_ai=bool(openai_key or claude_key), health_score=health_score,
                   total_products=total_products, active_products=active_products,
                   dup_urls=dup_urls, no_image=no_image, zero_price=zero_price,
                   stale_products=stale_products, suspect_links=suspect_links,
                   inactive_products=inactive_products,
                   total_articles=total_articles, thin_articles=thin_articles,
                   zero_view_articles=zero_view_articles, no_product_articles=no_product_articles,
                   ai_articles=ai_articles, vert_health=vert_health, net_stats=net_stats)

    return render_template('admin/ai_engine.html', **ctx)


# Keep old routes as redirects
@app.route('/admin/ai-center')
def admin_ai_center():
    """Redirect to unified AI Engine — health tab"""
    return redirect(url_for('admin_ai_engine', tab='health'))

@app.route('/admin/ai-center-legacy')
def admin_ai_center_legacy():
    """Unified AI Control Center - all AI operations in one place"""
    from sqlalchemy import func

    openai_key = SiteSettings.get('openai_key')
    claude_key = SiteSettings.get('claude_key')
    has_ai = bool(openai_key or claude_key)

    # ── Product Health Stats ──
    total_products = AffiliateLink.query.count()
    active_products = AffiliateLink.query.filter_by(is_active=True).count()

    # Duplicate URLs
    dup_sub = db.session.query(
        AffiliateLink.url, func.count(AffiliateLink.id).label('cnt')
    ).group_by(AffiliateLink.url).having(func.count(AffiliateLink.id) > 1).subquery()
    dup_urls = db.session.query(func.count()).select_from(dup_sub).scalar() or 0

    # Missing images
    no_image = AffiliateLink.query.filter(
        db.or_(AffiliateLink.image_url == '', AffiliateLink.image_url == None)
    ).count()

    # Price issues (0 or null)
    zero_price = AffiliateLink.query.filter(
        db.or_(AffiliateLink.price == 0, AffiliateLink.price == None)
    ).count()

    # Stale products (0 clicks)
    stale_products = AffiliateLink.query.filter(AffiliateLink.clicks == 0).count()

    # High clicks but 0 conversions (potential link issues)
    suspect_links = AffiliateLink.query.filter(
        AffiliateLink.clicks > 10, AffiliateLink.conversions == 0
    ).count()

    # Inactive products
    inactive_products = AffiliateLink.query.filter_by(is_active=False).count()

    # ── Article Health Stats ──
    total_articles = Article.query.count()
    # Thin content (content < 500 chars ~ approx 100 words)
    thin_articles = Article.query.filter(
        db.or_(func.length(Article.content) < 500, Article.content == '', Article.content == None)
    ).count()
    # No views
    zero_view_articles = Article.query.filter(
        db.or_(Article.views == 0, Article.views == None)
    ).count()
    # Articles without embed code (no products linked)
    no_product_articles = Article.query.filter(
        db.or_(Article.embed_code == '', Article.embed_code == None)
    ).count()
    # AI-generated articles
    ai_articles = Article.query.filter_by(ai_generated=True).count()

    # ── Vertical Health ──
    verticals = Vertical.query.all()
    vert_health = []
    for v in verticals:
        prod_count = db.session.query(AffiliateLink).join(Part).join(Zone).join(Segment).filter(
            Segment.vertical_id == v.id).count()
        art_count = Article.query.filter_by(vertical_slug=v.slug).count()
        part_count = db.session.query(Part).join(Zone).join(Segment).filter(
            Segment.vertical_id == v.id).count()
        # Parts with no products
        parts_no_prod = db.session.query(Part).join(Zone).join(Segment).filter(
            Segment.vertical_id == v.id
        ).outerjoin(AffiliateLink).filter(AffiliateLink.id == None).count()
        vert_health.append({
            'id': v.id, 'name': v.name, 'icon': v.icon, 'slug': v.slug,
            'products': prod_count, 'articles': art_count,
            'parts': part_count, 'parts_empty': parts_no_prod,
            'status': v.status
        })

    # ── Network Health ──
    networks = db.session.query(
        AffiliateLink.network,
        func.count(AffiliateLink.id),
        func.sum(AffiliateLink.clicks),
        func.sum(AffiliateLink.conversions)
    ).group_by(AffiliateLink.network).all()
    net_stats = [{'name': n[0], 'count': n[1], 'clicks': n[2] or 0, 'conv': n[3] or 0} for n in networks]

    # Score: overall health (0-100)
    issues_total = dup_urls + no_image + zero_price + suspect_links + thin_articles
    health_score = max(0, 100 - min(issues_total, 100))

    return render_template('admin/ai_center.html',
        has_ai=has_ai, health_score=health_score,
        total_products=total_products, active_products=active_products,
        dup_urls=dup_urls, no_image=no_image, zero_price=zero_price,
        stale_products=stale_products, suspect_links=suspect_links,
        inactive_products=inactive_products,
        total_articles=total_articles, thin_articles=thin_articles,
        zero_view_articles=zero_view_articles, no_product_articles=no_product_articles,
        ai_articles=ai_articles,
        vert_health=vert_health, net_stats=net_stats,
        verticals=verticals)

# Keep old route as redirect for backwards compatibility
@app.route('/admin/products/ai-classify')
def admin_products_ai_classify():
    return redirect(url_for('admin_ai_center'))


@app.route('/admin/ai-center/health-detail', methods=['POST'])
def admin_ai_center_health_detail():
    """AJAX: Get detailed list for a specific health issue"""
    from sqlalchemy import func
    data = request.get_json()
    issue_type = data.get('type', '')
    page = data.get('page', 1)
    per_page = 50

    results = []

    if issue_type == 'dup_urls':
        dup_urls = db.session.query(AffiliateLink.url).group_by(
            AffiliateLink.url).having(func.count(AffiliateLink.id) > 1).all()
        dup_url_list = [u[0] for u in dup_urls]
        items = AffiliateLink.query.filter(AffiliateLink.url.in_(dup_url_list)).order_by(
            AffiliateLink.url).offset((page-1)*per_page).limit(per_page).all()
        for al in items:
            results.append({'id': al.id, 'name': al.product_name, 'detail': al.url[:80],
                           'network': al.network, 'price': al.price})

    elif issue_type == 'no_image':
        items = AffiliateLink.query.filter(
            db.or_(AffiliateLink.image_url == '', AffiliateLink.image_url == None)
        ).offset((page-1)*per_page).limit(per_page).all()
        for al in items:
            results.append({'id': al.id, 'name': al.product_name, 'detail': 'Thieu hinh anh',
                           'network': al.network, 'price': al.price})

    elif issue_type == 'zero_price':
        items = AffiliateLink.query.filter(
            db.or_(AffiliateLink.price == 0, AffiliateLink.price == None)
        ).offset((page-1)*per_page).limit(per_page).all()
        for al in items:
            results.append({'id': al.id, 'name': al.product_name, 'detail': 'Gia = 0',
                           'network': al.network, 'price': 0})

    elif issue_type == 'suspect_links':
        items = AffiliateLink.query.filter(
            AffiliateLink.clicks > 10, AffiliateLink.conversions == 0
        ).order_by(AffiliateLink.clicks.desc()).offset((page-1)*per_page).limit(per_page).all()
        for al in items:
            results.append({'id': al.id, 'name': al.product_name,
                           'detail': f'{al.clicks} clicks, 0 conv',
                           'network': al.network, 'price': al.price})

    elif issue_type == 'stale':
        items = AffiliateLink.query.filter(
            AffiliateLink.clicks == 0
        ).offset((page-1)*per_page).limit(per_page).all()
        for al in items:
            results.append({'id': al.id, 'name': al.product_name, 'detail': '0 clicks',
                           'network': al.network, 'price': al.price})

    elif issue_type == 'thin_articles':
        from sqlalchemy import func as fn
        items = Article.query.filter(
            db.or_(fn.length(Article.content) < 500, Article.content == '', Article.content == None)
        ).offset((page-1)*per_page).limit(per_page).all()
        for a in items:
            results.append({'id': a.id, 'name': a.title,
                           'detail': f'{len(a.content or "")} ky tu', 'network': a.vertical_slug})

    elif issue_type == 'no_product_articles':
        items = Article.query.filter(
            db.or_(Article.embed_code == '', Article.embed_code == None)
        ).offset((page-1)*per_page).limit(per_page).all()
        for a in items:
            results.append({'id': a.id, 'name': a.title,
                           'detail': 'Chua gan san pham', 'network': a.vertical_slug})

    elif issue_type == 'parts_empty':
        vid = data.get('vertical_id')
        q = db.session.query(Part, Zone, Segment, Vertical).join(
            Zone, Part.zone_id == Zone.id).join(
            Segment, Zone.segment_id == Segment.id).join(
            Vertical, Segment.vertical_id == Vertical.id
        ).outerjoin(AffiliateLink).filter(AffiliateLink.id == None)
        if vid:
            q = q.filter(Vertical.id == vid)
        items = q.offset((page-1)*per_page).limit(per_page).all()
        for part, zone, seg, vert in items:
            results.append({'id': part.id, 'name': part.name_vi,
                           'detail': f'{vert.icon} {vert.name} > {zone.name}', 'network': ''})

    return jsonify({'results': results, 'page': page})


@app.route('/admin/ai-center/bulk-action', methods=['POST'])
def admin_ai_center_bulk_action():
    """AJAX: Perform bulk actions on products/articles from AI Center"""
    data = request.get_json()
    action = data.get('action')
    item_type = data.get('item_type', 'product')  # product or article
    ids = data.get('ids', [])

    if not ids:
        return jsonify({'error': 'Khong co item nao duoc chon'}), 400

    if item_type == 'product':
        if action == 'delete':
            count = AffiliateLink.query.filter(AffiliateLink.id.in_(ids)).delete(synchronize_session=False)
            db.session.commit()
            return jsonify({'success': True, 'message': f'Da xoa {count} san pham'})
        elif action == 'deactivate':
            AffiliateLink.query.filter(AffiliateLink.id.in_(ids)).update(
                {'is_active': False}, synchronize_session=False)
            db.session.commit()
            return jsonify({'success': True, 'message': f'Da tat {len(ids)} san pham'})
    elif item_type == 'article':
        if action == 'delete':
            count = Article.query.filter(Article.id.in_(ids)).delete(synchronize_session=False)
            db.session.commit()
            return jsonify({'success': True, 'message': f'Da xoa {count} bai viet'})

    return jsonify({'error': 'Action khong hop le'}), 400


@app.route('/admin/ai-center/ai-scan', methods=['POST'])
def admin_ai_center_ai_scan():
    """AJAX: AI-powered scan for product classification"""
    import requests as req

    data = request.get_json()
    vertical_id = data.get('vertical_id')
    offset = data.get('offset', 0)
    batch_size = min(data.get('batch_size', 20), 30)

    q = db.session.query(AffiliateLink, Part, Zone, Segment, Vertical).join(
        Part, AffiliateLink.part_id == Part.id
    ).join(Zone, Part.zone_id == Zone.id
    ).join(Segment, Zone.segment_id == Segment.id
    ).join(Vertical, Segment.vertical_id == Vertical.id)

    if vertical_id:
        q = q.filter(Vertical.id == vertical_id)

    total = q.count()
    products = q.order_by(AffiliateLink.id).offset(offset).limit(batch_size).all()

    if not products:
        return jsonify({'done': True, 'results': [], 'total': total, 'processed': offset})

    # Build vertical + zone context
    all_verticals = Vertical.query.all()
    vert_zones = {}
    for v in all_verticals:
        zones = []
        for s in v.segments:
            for z in s.zones:
                zones.append(z.name)
        vert_zones[v.name] = zones

    product_list = []
    for al, part, zone, seg, vert in products:
        product_list.append({
            'id': al.id, 'name': al.product_name or part.name_vi,
            'current_vertical': vert.name, 'current_zone': zone.name,
            'current_part': part.name_vi, 'url': (al.url or '')[:100], 'price': al.price,
        })

    # Build AI prompt
    vert_context = ""
    for vname, zones in vert_zones.items():
        vert_context += f"\n- {vname}: {', '.join(zones[:15])}"

    prompt = f"""Ban la chuyen gia phan loai san pham. Phan tich danh sach san pham duoi day va xac dinh chung thuoc nganh hang (vertical) nao.

Danh sach Verticals va cac Zone:{vert_context}

San pham can phan loai:
"""
    for i, p in enumerate(product_list):
        prompt += f"\n{i+1}. [{p['id']}] \"{p['name']}\" (hien tai: {p['current_vertical']} > {p['current_zone']})"

    prompt += """

Tra loi CHINH XAC theo format JSON array:
[{"id": <product_id>, "correct_vertical": "<ten vertical dung>", "confidence": "high/medium/low", "reason": "<ly do ngan>"}]

Chi tra ve JSON array, KHONG giai thich them. Neu san pham dang o dung vertical thi correct_vertical = vertical hien tai."""

    ai_results = _call_ai_api(prompt)

    if not ai_results:
        err_detail = getattr(_call_ai_api, 'last_error', '')
        return jsonify({'error': f'AI API loi: {err_detail}' if err_detail else 'Khong the ket noi AI API. Kiem tra API key trong Settings.'}), 500

    # Compare and build results
    results = []
    for p in product_list:
        ai_item = next((r for r in ai_results if r.get('id') == p['id']), None)
        if ai_item:
            suggested = ai_item.get('correct_vertical', p['current_vertical'])
            is_mismatch = suggested.strip().lower() != p['current_vertical'].strip().lower()
            results.append({
                'id': p['id'], 'name': p['name'],
                'current_vertical': p['current_vertical'], 'current_zone': p['current_zone'],
                'suggested_vertical': suggested,
                'confidence': ai_item.get('confidence', 'low'),
                'reason': ai_item.get('reason', ''), 'mismatch': is_mismatch,
            })
        else:
            results.append({
                'id': p['id'], 'name': p['name'],
                'current_vertical': p['current_vertical'], 'current_zone': p['current_zone'],
                'suggested_vertical': p['current_vertical'],
                'confidence': 'low', 'reason': 'AI khong phan loai duoc', 'mismatch': False,
            })

    return jsonify({
        'done': offset + batch_size >= total,
        'results': results, 'total': total,
        'processed': min(offset + batch_size, total),
        'next_offset': offset + batch_size
    })


def _call_ai_api(prompt):
    """Unified AI API caller - tries Claude first, then OpenAI. Returns parsed JSON or None.
    Sets _call_ai_api.last_error with debug info if all attempts fail."""
    import requests as req
    _call_ai_api.last_error = ''
    claude_key = SiteSettings.get('claude_key')
    openai_key = SiteSettings.get('openai_key')

    if not claude_key and not openai_key:
        _call_ai_api.last_error = 'Chua co API key nao. Vao Settings > AI API Keys de them.'
        return None

    if claude_key:
        try:
            resp = req.post('https://api.anthropic.com/v1/messages',
                headers={'x-api-key': claude_key, 'anthropic-version': '2023-06-01',
                         'content-type': 'application/json'},
                json={'model': 'claude-sonnet-4-5-20250929', 'max_tokens': 4000,
                      'messages': [{'role': 'user', 'content': prompt}]},
                timeout=60)
            if resp.status_code == 200:
                result = _parse_ai_json_response(resp.json()['content'][0]['text'])
                if result:
                    return result
                _call_ai_api.last_error = 'Claude tra loi nhung khong parse duoc JSON'
            else:
                err_body = resp.text[:300]
                _call_ai_api.last_error = f'Claude API loi {resp.status_code}: {err_body}'
        except Exception as e:
            _call_ai_api.last_error = f'Claude exception: {str(e)}'

    if openai_key:
        try:
            resp = req.post('https://api.openai.com/v1/chat/completions',
                headers={'Authorization': f'Bearer {openai_key}', 'Content-Type': 'application/json'},
                json={'model': 'gpt-4o-mini',
                      'messages': [
                          {'role': 'system', 'content': 'You are a product classification expert. Always respond with valid JSON.'},
                          {'role': 'user', 'content': prompt}],
                      'max_tokens': 4000, 'temperature': 0.2},
                timeout=60)
            if resp.status_code == 200:
                result = _parse_ai_json_response(resp.json()['choices'][0]['message']['content'])
                if result:
                    return result
                _call_ai_api.last_error = 'OpenAI tra loi nhung khong parse duoc JSON'
            else:
                err_body = resp.text[:300]
                _call_ai_api.last_error = f'OpenAI API loi {resp.status_code}: {err_body}'
        except Exception as e:
            _call_ai_api.last_error = f'OpenAI exception: {str(e)}'
    return None


def _parse_ai_json_response(text):
    """Parse AI response, extract JSON array"""
    import json, re
    text = text.strip()
    text = re.sub(r'^```(?:json)?\s*', '', text)
    text = re.sub(r'\s*```$', '', text)
    text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r'\[[\s\S]*\]', text)
        if match:
            try:
                return json.loads(match.group())
            except json.JSONDecodeError:
                pass
    return None


# Keep old apply route as alias
@app.route('/admin/products/ai-classify/apply', methods=['POST'])
def admin_products_ai_classify_apply():
    return admin_ai_center_bulk_action()


def _build_part_keyword_index():
    """Build keyword index from ALL Parts across ALL verticals for auto-mapping.
    Returns list of {part_id, keywords: set(), zone_name, part_name, ...}
    """
    import unicodedata, re
    def normalize(text):
        """Lowercase + strip Vietnamese diacritics for fuzzy matching"""
        text = text.lower().strip()
        nfkd = unicodedata.normalize('NFKD', text)
        return ''.join(c for c in nfkd if not unicodedata.combining(c))

    q = db.session.query(Part, Zone, Segment, Vertical).join(
        Zone, Part.zone_id == Zone.id
    ).join(Segment, Zone.segment_id == Segment.id
    ).join(Vertical, Segment.vertical_id == Vertical.id)

    index = []
    for part, zone, seg, vert in q.all():
        keywords = set()
        keywords.add(normalize(part.name_vi))
        if part.name_en:
            keywords.add(normalize(part.name_en))
        if part.tags:
            for tag in part.tags.split(','):
                tag = tag.strip()
                if len(tag) >= 2:
                    keywords.add(normalize(tag))
        if part.oem_code:
            for code in re.findall(r'[A-Za-z0-9]+-[A-Za-z0-9]+', part.oem_code):
                keywords.add(code.lower())

        index.append({
            'part_id': part.id,
            'keywords': keywords,
            'zone_name': zone.name,
            'zone_slug': zone.slug,
            'part_name': part.name_vi,
            'seg_name': seg.name,
            'vert_name': vert.name,
            'vert_slug': vert.slug,
        })

    # Zone-level keyword map — ALL verticals (fallback when no Part matches)
    # Keywords are normalized (no diacritics, lowercase)
    zone_keywords = {
        # === CAR / OTO ===
        'he-thong-treo': ['phuoc', 'giam chan', 'lo xo', 'cao su', 'rotuy', 'rotuyn', 'thanh can bang',
                          'shock', 'absorber', 'spring', 'bushing', 'ball joint', 'suspension',
                          'giam xoc', 'nhun', 'stabilizer', 'sway bar', 'gat phuoc',
                          'cao su cang a', 'cao su thanh giang', 'thanh giang',
                          'bilstein', 'kayaba', 'kyb', 'monroe', 'sachs',
                          'chup mui', 'op bi', 'bi dien'],
        'he-thong-phanh': ['ma phanh', 'dia phanh', 'bau tro luc', 'caliper',
                           'brake pad', 'brake disc', 'dau phanh', 'tang bua',
                           'bo kep phanh', 'phanh tay', 'phanh dia', 'phanh tang trong',
                           'brembo', 'akebono', 'bendix', 'ferodo', 'bosch brake',
                           'ong dau phanh', 'tro luc phanh', 'bua phanh'],
        'dong-co': ['dong co', 'engine', 'bugi', 'spark plug', 'kim phun', 'turbo', 'block may',
                    'loc gio', 'loc dau', 'loc nhien lieu', 'dau may', 'oil filter', 'air filter',
                    'piston', 'xi lanh', 'truc khuyu', 'crankshaft', 'gasket',
                    'gioang', 'day curoa', 'bom nuoc', 'water pump', 'bom xang',
                    'injector', 'throttle', 'nap may', 'supap', 'cam', 'truc cam',
                    'dau nhot', 'nhot dong co', 'castrol', 'mobil 1', 'shell helix',
                    'total quartz', 'motul', 'liqui moly', 'denso', 'ngk',
                    'turbo tang ap', 'intercooler', 'supercharger',
                    'ron nap may', 'phot dau', 'gioang quy lat'],
        'he-thong-dien': ['ac quy', 'battery oto', 'may phat', 'alternator',
                          'bong den oto', 'day dien oto',
                          'cau chi', 'cam bien', 'sensor', 'ecu', 'relay',
                          'gat mua', 'wiper', 'starter', 'khoi dong',
                          'den pha', 'den hau', 'den xi nhan', 'den led oto',
                          'philips oto', 'osram', 'bosch oto', 'varta',
                          'cam bien oxy', 'cam bien nhiet do', 'cam bien toc do',
                          'hop den', 'day co', 'cuon danh lua'],
        'he-thong-lai': ['thuoc lai', 'tro luc lai', 'vo lang',
                         'power steering', 'rack', 'tie rod', 'tay lai',
                         'cao su ro tuyn', 'rotuy lai', 'bom tro luc lai',
                         'dau tro luc lai', 'tay lai xe'],
        'gam-xe': ['gam xe', 'khung gam', 'chassis', 'chan bun', 'underbody', 'che gam',
                   'lop oto', 'mam oto', 'bac dan', 'bearing', 'moay o',
                   'lop xe oto', 'michelin', 'bridgestone', 'goodyear', 'dunlop',
                   'continental', 'hankook', 'yokohama', 'kumho', 'maxxis',
                   'mam xe oto', 'la zang', 'lazang', 'bu long banh xe',
                   'bi moay o', 'phot moay o'],
        'noi-that': ['noi that oto', 'ghe oto', 'taplo', 'dashboard', 'tham oto',
                     'dieu hoa oto', 'guong chieu hau', 'boc ghe oto', 'boc vo lang',
                     'tham lot san oto', 'tham san oto', 'nem tua lung oto',
                     'camera hanh trinh', 'dashcam', 'dvd oto', 'man hinh oto',
                     'loc dieu hoa', 'loc gio lanh'],
        'ngoai-that': ['ngoai that oto', 'can oto', 'bumper', 'guong oto',
                       'kinh oto', 'windshield', 'capo', 'hood', 'cop xe',
                       'tem xe', 'decal', 'ong xa oto', 'exhaust',
                       'thanh gia noc', 'baga noc', 'can truoc', 'can sau',
                       'op ca lang', 'body kit', 'spoiler', 'cua gio oto',
                       'chay bam', 'phu kinh', 'khung bao ve'],
        # === PET ===
        'dinh-duong': ['thuc an cho', 'thuc an meo', 'dog food', 'cat food', 'hat cho', 'hat meo',
                       'royal canin', 'pedigree', 'whiskas', 'taste of the wild', 'nutrience',
                       'smartheart', 'ganador', 'natural core', 'snack cho', 'snack meo',
                       'pate cho', 'pate meo', 'sua cho', 'sua meo', 'puppy food',
                       'kitten food', 'dinh duong thu cung', 'pet food',
                       'thuc an cho cho', 'thuc an cho meo',
                       'pro plan', 'proplan', 'hills', 'science diet',
                       'acana', 'orijen', 'ziwi peak', 'wellness', 'nutro',
                       'brit care', 'monge', 'reflex', 'minino', 'catsrang',
                       'iskhan', 'jerhigh', 'inaba', 'ciao churu',
                       'treats cho', 'treats meo', 'banh thuong cho', 'banh thuong meo',
                       'thuc an hat', 'thuc an uot', 'thuc an kho'],
        'y-te': ['vaccine cho', 'vaccine meo', 'tay giun', 'thuoc ve', 'thuoc ran',
                 'frontline', 'nexgard', 'heartgard', 'revolution', 'thuoc nho gay',
                 'thuoc tri nam', 'thuoc tri ve', 'y te thu cung',
                 'thuoc cho cho', 'thuoc cho meo',
                 'advantage', 'advocate', 'bravecto', 'simparica', 'seresto',
                 'drontal', 'milbemax', 'panacur', 'vetrimec',
                 'sua tam cho', 'sua tam meo', 'dau goi cho', 'dau goi meo',
                 've sinh tai', 've sinh rang', 'kem danh rang cho',
                 'thuoc sat trung', 'biotic', 'probiotic thu cung',
                 'vitamin thu cung', 'canxi cho', 'canxi meo'],
        'huan-luyen': ['huan luyen cho', 'day cho', 'treat huan luyen',
                       'clicker', 'muzzle', 'day dan cho',
                       'ro mom cho', 'bang khen', 'huan luyen thu cung'],
        'do-dung': ['chuong cho', 'chuong meo', 'long cho', 'long meo', 'bat an', 'binh nuoc',
                    'vong co', 'day dat', 'leash', 'collar', 'ao cho', 'ao meo',
                    'nem cho', 'giuong cho', 'nha meo', 'cat tree', 'cay leo meo',
                    'do choi cho', 'do choi meo', 'balo thu cung', 'tui van chuyen',
                    'xe day thu cung', 'pet carrier', 'khay ve sinh', 'cat ve sinh',
                    'cat toan', 'bentonite', 'tofu cat', 'crystal cat',
                    'pet fountain', 'phu kien thu cung',
                    'harness', 'yem cho', 'dai yem cho',
                    'chuong van chuyen', 'long van chuyen',
                    'balo meo', 'tui meo', 'nha cho', 'o cho', 'o meo',
                    'qua bong cho', 'xuong gam', 'cau cao meo',
                    'mieng lot khay', 'ta cho', 'bim cho',
                    'mang bat', 'may loc nuoc thu cung',
                    'luoc chai long', 'keo cat long', 'may cat long',
                    'ban chai long', 'gang tam', 'may say long'],
        'hamster': ['hamster', 'chuot hamster', 'long hamster', 'thuc an hamster',
                    'banh xe hamster', 'cat lot hamster', 'chuot lang', 'guinea pig',
                    'bo hamster', 'binh nuoc hamster', 'nha hamster'],
        'ca-canh': ['ca canh', 'be ca', 'aquarium', 'loc nuoc be ca', 'thuc an ca',
                    'den be ca', 'co thuy sinh', 'nen be ca', 'ca betta', 'ca vang',
                    'may bom oxy', 'may sui', 'loc be ca', 'may loc be ca',
                    'da trang tri be ca', 'lu be ca', 'kinh be ca'],
        # === BEAUTY ===
        'lam-sach': ['sua rua mat', 'tay trang', 'dau tay trang', 'nuoc tay trang', 'cleansing',
                     'micellar', 'foam rua mat', 'gel rua mat', 'oil cleanser', 'lam sach da',
                     'cerave cleanser', 'la roche posay cleanser', 'cetaphil',
                     'bioderma', 'garnier micellar', 'simple cleanser',
                     'rua mat', 'kem rua mat', 'bot rua mat',
                     'tay te bao chet', 'scrub', 'exfoliate', 'peeling gel'],
        'toner-essence': ['toner', 'nuoc hoa hong', 'essence', 'tinh chat lot', 'lotion',
                          'nuoc can bang', 'first treatment',
                          'klairs toner', 'thayers', 'hada labo', 'laneige',
                          'some by mi toner', 'cosrx toner', 'innisfree toner',
                          'nuoc than', 'nuoc than thanh xuan'],
        'serum': ['serum', 'tinh chat', 'ampoule', 'vitamin c serum', 'retinol', 'niacinamide',
                  'hyaluronic', 'peptide', 'serum duong',
                  'the ordinary', 'cerave serum', 'la roche posay serum',
                  'obagi', 'skinceuticals', 'paula choice', 'drunk elephant',
                  'cosrx serum', 'some by mi serum', 'klairs serum',
                  'serum tri mun', 'serum trang da', 'serum chong lao hoa',
                  'tinh chat duong', 'dac tri'],
        'kem-duong': ['kem duong', 'moisturizer', 'cream', 'kem duong am', 'night cream',
                      'day cream', 'gel duong', 'emulsion', 'kem mat',
                      'cerave cream', 'la roche posay cream', 'laneige cream',
                      'innisfree cream', 'sulwhasoo', 'whoo', 'sk-ii',
                      'kem lot', 'primer', 'kem nen', 'foundation',
                      'kem duong da', 'kem duong mat', 'kem duong the',
                      'body lotion', 'sua duong the', 'kem tay'],
        'chong-nang': ['chong nang', 'sunscreen', 'kem chong nang', 'sun cream',
                       'uv protection', 'sunblock',
                       'anessa', 'skin aqua', 'biore uv', 'la roche posay uv',
                       'eucerin sun', 'neutrogena sun', 'innisfree sun',
                       'chong nang da mat', 'chong nang body', 'xit chong nang',
                       'gel chong nang', 'sua chong nang'],
        'mat-na': ['mat na', 'mask', 'sheet mask', 'sleeping mask', 'clay mask',
                   'mat na giay', 'mat na dat set', 'mat na ngu', 'peel off',
                   'mat na vita', 'mat na collagen', 'mat na hyaluron',
                   'mediheal', 'jayjun', 'dr jart', 'papa recipe',
                   'mat na duong', 'mat na trang da', 'mat na tri mun'],
        'trang-diem': ['son moi', 'lipstick', 'son kem', 'lip tint', 'son duong',
                       'phan ma', 'blush', 'phan mat', 'eyeshadow', 'mascara',
                       'ke mat', 'eyeliner', 'ke may', 'eyebrow', 'phan phu',
                       'cushion', 'bb cream', 'cc cream', 'highlight', 'contour',
                       'mac', 'maybelline', 'loreal', 'nyx', 'revlon',
                       'rom&nd', 'romand', 'peripera', '3ce', 'black rouge',
                       'clio', 'etude', 'missha', 'the saem', 'espoir',
                       'phan nuoc', 'trang diem', 'makeup', 'kem nen'],
        'nuoc-hoa': ['nuoc hoa', 'parfum', 'perfume', 'eau de toilette', 'edt', 'edp',
                     'cologne', 'body mist', 'xit thom',
                     'chanel', 'dior perfume', 'gucci perfume', 'versace',
                     'calvin klein', 'narciso', 'jo malone', 'tom ford'],
        'cham-soc-toc': ['dau goi', 'dau xa', 'shampoo', 'conditioner', 'kem u toc',
                         'dau duong toc', 'hair serum', 'hair mask',
                         'nhuom toc', 'uon toc', 'duoi toc', 'keo noi toc',
                         'pantene', 'dove', 'tresemme', 'loreal hair',
                         'moroccanoil', 'olaplex', 'kerastase',
                         'gel vuot toc', 'sap vuot toc', 'keo xit toc',
                         'may say toc', 'may uon toc', 'may la toc',
                         'luoc', 'luoc go', 'luoc chai'],
        # === TECH ===
        'dien-thoai': ['dien thoai', 'smartphone', 'iphone', 'samsung galaxy', 'xiaomi',
                       'oppo', 'vivo', 'realme', 'poco', 'oneplus', 'google pixel',
                       'huawei', 'nokia', 'asus rog phone', 'nothing phone',
                       'iphone 15', 'iphone 16', 'galaxy s24', 'galaxy s25',
                       'redmi', 'redmi note'],
        'laptop': ['laptop', 'macbook', 'thinkpad', 'dell xps', 'asus zenbook',
                   'hp envy', 'hp pavilion', 'acer swift', 'lenovo ideapad',
                   'surface', 'chromebook', 'gaming laptop',
                   'asus vivobook', 'msi gaming', 'acer nitro', 'legion',
                   'may tinh xach tay', 'ultrabook'],
        'tablet': ['tablet', 'ipad', 'galaxy tab', 'xiaomi pad',
                   'may tinh bang', 'kindle', 'but cam ung', 'apple pencil'],
        'man-hinh': ['man hinh', 'monitor', 'man hinh may tinh', 'man hinh gaming',
                     'cuong luc', 'dan man hinh', 'tempered glass',
                     'man hinh 4k', 'man hinh cong', 'ultrawide'],
        'pin-sac': ['power bank', 'sac nhanh', 'sac khong day',
                    'wireless charging', 'cap sac', 'adapter', 'magsafe', 'pin du phong',
                    'anker', 'baseus', 'ugreen', 'belkin', 'aukey',
                    'cap type c', 'cap lightning', 'cap usb', 'cu sac',
                    'sac iphone', 'sac samsung', 'sac xe hoi'],
        'tai-nghe': ['tai nghe', 'headphone', 'earphone', 'earbuds', 'airpods',
                     'galaxy buds', 'sony wh', 'sony wf', 'jabra', 'jbl',
                     'bose', 'sennheiser', 'beats', 'marshall',
                     'tai nghe bluetooth', 'tai nghe khong day', 'tai nghe gaming',
                     'true wireless', 'chong on', 'noise cancelling'],
        'loa': ['loa bluetooth', 'loa di dong', 'loa thong minh', 'soundbar',
                'jbl speaker', 'harman kardon', 'marshall speaker', 'bose speaker',
                'loa karaoke', 'loa keo', 'loa sub', 'amply'],
        'dong-ho-thong-minh': ['smartwatch', 'dong ho thong minh', 'apple watch',
                               'galaxy watch', 'amazfit', 'huawei watch', 'mi band',
                               'fitbit', 'garmin venu', 'vong deo tay thong minh',
                               'day dong ho', 'day apple watch'],
        'phu-kien-tech': ['op lung', 'case', 'bao da', 'mieng dan',
                          'gia do dien thoai', 'tripod', 'selfie stick',
                          'the nho', 'sd card', 'usb', 'ssd', 'o cung',
                          'hub usb', 'dock', 'chuot', 'ban phim',
                          'webcam', 'microphone', 'ring light',
                          'tui chong soc', 'balo laptop'],
        # === BIKE ===
        'khung-xe': ['khung xe dap', 'frame', 'suon xe dap', 'carbon frame', 'nhom frame',
                     'giant', 'trek', 'specialized', 'cannondale', 'merida',
                     'xe dap dua', 'xe dap dia hinh', 'xe dap the thao',
                     'xe dap gap', 'xe dap touring', 'xe dap fixed gear'],
        'he-thong-truyen-dong': ['truyen dong', 'shimano', 'sram', 'xich xe dap',
                                  'bat dia', 'tay de', 'shifter', 'derailleur', 'groupset',
                                  'cassette', 'pedal', 'ban dap',
                                  'shimano deore', 'shimano 105', 'shimano ultegra',
                                  'shimano dura ace', 'sram eagle', 'sram rival',
                                  'campagnolo', 'truc giua', 'bottom bracket'],
        'banh-xe': ['banh xe dap', 'lop xe dap', 'vanh xe dap', 'nan hoa', 'hub xe dap',
                    'sam xe dap', 'tire bike', 'wheelset', 'tubeless',
                    'continental tire', 'schwalbe', 'vittoria', 'maxxis tire',
                    'mavic', 'dt swiss', 'fulcrum', 'lop 700c', 'lop 26', 'lop 29'],
        'yen-va-tay-lai': ['yen xe dap', 'tay lai xe dap', 'ghi dong', 'stem', 'saddle',
                           'handlebar', 'bar tape', 'grip',
                           'fizik', 'selle italia', 'brooks saddle', 'ergon',
                           'phuoc xe dap', 'fox', 'rockshox'],
        'phu-kien-xe-dap': ['mu bao hiem xe dap', 'helmet bike', 'gang tay dap xe', 'kinh dap xe',
                            'den xe dap', 'binh nuoc xe dap', 'bom xe dap', 'dong ho xe dap',
                            'garmin edge', 'wahoo', 'khoa xe dap',
                            'ao dap xe', 'quan dap xe', 'giay dap xe',
                            'tui xe dap', 'gac xe dap', 'baga xe dap',
                            'chuong xe dap', 'ke xe dap', 'bao tay xe dap'],
        # === SPORT ===
        'giay-chay': ['giay chay', 'running shoes', 'nike running', 'adidas running',
                      'asics', 'hoka', 'new balance running', 'saucony', 'brooks',
                      'giay chay bo', 'giay the thao',
                      'nike pegasus', 'nike vaporfly', 'adidas ultraboost',
                      'asics gel', 'asics nimbus', 'asics kayano',
                      'hoka clifton', 'hoka bondi', 'hoka speedgoat',
                      'on running', 'on cloud', 'mizuno', 'under armour'],
        'do-chay': ['do chay bo', 'quan chay', 'ao chay', 'running wear',
                    'shorts chay', 'tights', 'compression',
                    'ao tank', 'ao singlet', 'quan short chay',
                    'nike dri fit', 'adidas climalite', 'under armour heatgear',
                    'ao giu nhiet', 'ao chong nang chay'],
        'dong-ho-gps': ['dong ho gps', 'garmin forerunner', 'garmin fenix', 'coros',
                        'apple watch ultra', 'suunto', 'polar', 'dong ho the thao',
                        'coros pace', 'coros apex', 'coros vertix',
                        'garmin 255', 'garmin 265', 'garmin 965',
                        'polar vantage', 'polar pacer', 'suunto race'],
        'phu-kien-chay': ['dai deo dien thoai', 'running belt', 'ba lo chay',
                          'vo chay', 'running socks', 'headband', 'arm sleeve',
                          'mu chay', 'kinh chay', 'gang tay chay',
                          'dai deo nguc', 'heart rate monitor', 'hrm',
                          'flip belt', 'nathan', 'salomon vest'],
        'dinh-duong-chay': ['gel nang luong', 'energy gel', 'nuoc uong the thao',
                            'electrolyte', 'bcaa', 'whey protein', 'energy bar',
                            'gu energy', 'maurten', 'spring energy', 'tailwind',
                            'nuoc tang luc', 'binh nuoc chay', 'binh nau'],
        # === TRAVEL ===
        'mien-bac': ['ha noi', 'sa pa', 'ha long', 'ninh binh', 'ha giang', 'cao bang',
                     'moc chau', 'mai chau', 'tam dao', 'cat ba', 'du lich mien bac',
                     'fansipan', 'dong van', 'ban gioc', 'bac son', 'mu cang chai',
                     'yen bai', 'lang son', 'bac kan', 'thai nguyen', 'tuyen quang'],
        'mien-trung': ['da nang', 'hoi an', 'hue', 'nha trang', 'da lat', 'quy nhon',
                       'phu yen', 'quang binh', 'phong nha', 'du lich mien trung',
                       'son tra', 'ba na hills', 'cu lao cham', 'eo gio',
                       'ninh thuan', 'binh thuan', 'ly son', 'kon tum'],
        'mien-nam': ['sai gon', 'ho chi minh', 'vung tau', 'can tho', 'phu quoc',
                     'con dao', 'mekong', 'du lich mien nam',
                     'mui ne', 'long hai', 'ho tram', 'binh chau',
                     'ben tre', 'tien giang', 'an giang', 'chau doc'],
        'dong-nam-a': ['thai lan', 'singapore', 'malaysia', 'bali', 'indonesia',
                       'philippines', 'cambodia', 'myanmar', 'lao',
                       'bangkok', 'phuket', 'boracay', 'siem reap', 'luang prabang'],
        'dong-a': ['nhat ban', 'han quoc', 'dai loan', 'trung quoc', 'hong kong',
                   'tokyo', 'osaka', 'kyoto', 'seoul', 'busan', 'jeju'],
        'budget': ['hostel', 'homestay', 'nha nghi', 'budget hotel', 'backpacker',
                   'airbnb', 'phong tro', 'nha dan', 'camping', 'glamping'],
        'resort': ['resort', 'khach san', 'hotel 5 sao', 'hotel 4 sao', 'villa',
                   'spa resort', 'beach resort',
                   'vinpearl', 'melia', 'pullman', 'intercontinental',
                   'marriott', 'hilton', 'accor', 'hyatt', 'six senses',
                   'amanoi', 'jw marriott'],
        # === FASHION ===
        'ao-nam': ['ao polo', 'ao thun nam', 'ao so mi nam', 'ao khoac nam',
                   'ao vest', 'blazer nam', 'ao hoodie', 'ao len nam',
                   'ao dai tay nam', 'ao ngan tay nam'],
        'ao-nu': ['ao thun nu', 'ao so mi nu', 'ao khoac nu', 'ao dai',
                  'ao croptop', 'ao kiem', 'ao hai day', 'cardigan',
                  'ao len nu', 'ao choang'],
        'quan': ['quan jeans', 'quan tay', 'quan kaki', 'quan short',
                 'quan jogger', 'quan the thao', 'quan dai',
                 'quan au', 'quan ong rong', 'quan ong dung'],
        'vay-dam': ['dam', 'vay', 'dam lien', 'dam xoe', 'dam suong',
                    'chan vay', 'vay midi', 'vay maxi', 'jumpsuit',
                    'dam du tiec', 'dam cong so'],
        'giay-dep': ['giay nam', 'giay nu', 'giay sneaker', 'giay boot',
                     'giay cao got', 'giay the thao', 'dep', 'sandal',
                     'nike', 'adidas', 'converse', 'vans', 'puma',
                     'giay da', 'giay luoi', 'giay tay'],
        'tui-xach': ['tui xach', 'balo', 'vi', 'clutch', 'tui deo cheo',
                     'tui tote', 'tui deo vai', 'vi nam', 'vi nu',
                     'charles keith', 'pedro', 'coach', 'michael kors'],
        'phu-kien-thoi-trang': ['kinh mat', 'dong ho', 'vong tay', 'day chuyen',
                                'nhan', 'bong tai', 'khan', 'that lung',
                                'mu non', 'non ket', 'non luoi trai',
                                'trang suc', 'phu kien toc'],
        # === HOME & KITCHEN ===
        'nha-bep': ['noi', 'chao', 'xoong', 'noi com dien', 'lo vi song',
                    'may xay sinh to', 'may ep', 'binh dun nuoc',
                    'noi chien khong dau', 'air fryer', 'lo nuong',
                    'bep tu', 'bep gas', 'bep dien',
                    'dao', 'thot', 'muong', 'dia', 'chen', 'bat',
                    'tupperware', 'lock lock', 'tefal', 'sunhouse',
                    'supor', 'kangaroo', 'electrolux', 'philips'],
        'noi-that-nha': ['ban', 'ghe', 'tu', 'giuong', 'nem', 'goi',
                         'ga giuong', 'chan ga', 'rem cua', 'tham',
                         'den trang tri', 'tranh treo tuong',
                         'ke sach', 'tu quan ao', 'ban lam viec',
                         'sofa', 'ghe van phong', 'tu giay'],
        've-sinh': ['may hut bui', 'robot hut bui', 'may lau nha',
                    'nuoc lau san', 'nuoc rua chen', 'nuoc giat',
                    'nuoc xa', 'bot giat', 'vim', 'sunlight',
                    'comfort', 'downy', 'ariel', 'omo',
                    'giay ve sinh', 'khan giay', 'tui rac'],
        'dien-gia-dung': ['may giat', 'tu lanh', 'may lanh', 'dieu hoa',
                          'quat', 'may loc khong khi', 'may hut am',
                          'may loc nuoc', 'binh nong lanh',
                          'samsung', 'lg', 'panasonic', 'toshiba',
                          'sharp', 'daikin', 'mitsubishi'],
        # === HEALTH ===
        'thuc-pham-chuc-nang': ['vitamin', 'omega 3', 'dau ca', 'collagen',
                                'canxi', 'sat', 'kem', 'magie',
                                'vitamin c', 'vitamin d', 'vitamin e',
                                'thuc pham chuc nang', 'tpcn', 'bao',
                                'dhc', 'blackmores', 'swisse', 'nature made',
                                'kirkland', 'centrum', 'solgar', 'now foods',
                                'bo sung dinh duong', 'vien uong',
                                'tang can', 'giam can', 'detox'],
        'dung-cu-y-te': ['nhiet ke', 'may do huyet ap', 'may do duong huyet',
                         'may xong', 'may xong khi dung', 'khau trang',
                         'bang ca nhan', 'on dinh huyet ap',
                         'omron', 'microlife', 'beurer', 'yuwell'],
        # === BABY & KIDS ===
        'do-tre-em': ['ta em be', 'bim', 'sua bot', 'binh sua',
                      'xe day tre em', 'ghe an dam', 'noi em be',
                      'do choi tre em', 'lego', 'xe day', 'carseat',
                      'ta dan', 'ta quan', 'bobby', 'huggies',
                      'merries', 'moony', 'pampers', 'mamypoko',
                      'sua nan', 'sua ensure', 'sua abbott', 'similac',
                      'vinamilk', 'nutifood', 'friso', 'meiji'],
    }

    # Power keywords: highly distinctive brands/terms — single hit is enough
    # These map directly to zone slugs and give bonus score
    power_keywords = {
        # PET brands → dinh-duong
        'royal canin': 'dinh-duong', 'pedigree': 'dinh-duong', 'whiskas': 'dinh-duong',
        'taste of the wild': 'dinh-duong', 'acana': 'dinh-duong', 'orijen': 'dinh-duong',
        'smartheart': 'dinh-duong', 'ganador': 'dinh-duong', 'nutrience': 'dinh-duong',
        'pro plan': 'dinh-duong', 'hills science diet': 'dinh-duong', 'catsrang': 'dinh-duong',
        'minino': 'dinh-duong', 'ziwi peak': 'dinh-duong', 'jerhigh': 'dinh-duong',
        'ciao churu': 'dinh-duong', 'inaba': 'dinh-duong',
        # PET health → y-te
        'frontline': 'y-te', 'nexgard': 'y-te', 'heartgard': 'y-te',
        'bravecto': 'y-te', 'simparica': 'y-te', 'seresto': 'y-te',
        'advocate': 'y-te', 'revolution': 'y-te', 'drontal': 'y-te',
        # PET accessories → do-dung
        'bentonite': 'do-dung', 'tofu cat': 'do-dung', 'crystal cat': 'do-dung',
        'cat tree': 'do-dung', 'pet carrier': 'do-dung', 'pet fountain': 'do-dung',
        # BEAUTY brands
        'cerave': 'lam-sach', 'la roche posay': 'lam-sach', 'cetaphil': 'lam-sach',
        'bioderma': 'lam-sach', 'the ordinary': 'serum', 'obagi': 'serum',
        'skinceuticals': 'serum', 'paula choice': 'serum',
        'anessa': 'chong-nang', 'skin aqua': 'chong-nang', 'biore uv': 'chong-nang',
        'mediheal': 'mat-na', 'dr jart': 'mat-na',
        'romand': 'trang-diem', 'peripera': 'trang-diem', '3ce': 'trang-diem',
        'black rouge': 'trang-diem', 'maybelline': 'trang-diem',
        'moroccanoil': 'cham-soc-toc', 'olaplex': 'cham-soc-toc', 'kerastase': 'cham-soc-toc',
        # TECH brands
        'iphone': 'dien-thoai', 'samsung galaxy': 'dien-thoai', 'xiaomi': 'dien-thoai',
        'macbook': 'laptop', 'thinkpad': 'laptop', 'dell xps': 'laptop',
        'airpods': 'tai-nghe', 'galaxy buds': 'tai-nghe', 'sony wh': 'tai-nghe',
        'jbl speaker': 'loa', 'harman kardon': 'loa', 'marshall speaker': 'loa',
        'apple watch': 'dong-ho-thong-minh', 'galaxy watch': 'dong-ho-thong-minh',
        'amazfit': 'dong-ho-thong-minh',
        'anker': 'pin-sac', 'baseus': 'pin-sac', 'ugreen': 'pin-sac',
        # CAR brands
        'bilstein': 'he-thong-treo', 'kayaba': 'he-thong-treo', 'kyb': 'he-thong-treo',
        'brembo': 'he-thong-phanh', 'akebono': 'he-thong-phanh',
        'castrol': 'dong-co', 'mobil 1': 'dong-co', 'shell helix': 'dong-co',
        'liqui moly': 'dong-co', 'motul': 'dong-co',
        'michelin': 'gam-xe', 'bridgestone': 'gam-xe', 'goodyear': 'gam-xe',
        'continental': 'gam-xe', 'hankook': 'gam-xe', 'yokohama': 'gam-xe',
        # BIKE brands
        'shimano': 'he-thong-truyen-dong', 'sram': 'he-thong-truyen-dong',
        'campagnolo': 'he-thong-truyen-dong',
        'giant': 'khung-xe', 'trek': 'khung-xe', 'specialized': 'khung-xe',
        'rockshox': 'yen-va-tay-lai',
        # SPORT brands
        'nike pegasus': 'giay-chay', 'nike vaporfly': 'giay-chay',
        'adidas ultraboost': 'giay-chay', 'hoka clifton': 'giay-chay',
        'hoka bondi': 'giay-chay', 'on running': 'giay-chay',
        'garmin forerunner': 'dong-ho-gps', 'garmin fenix': 'dong-ho-gps',
        'coros pace': 'dong-ho-gps', 'suunto race': 'dong-ho-gps',
        'maurten': 'dinh-duong-chay', 'gu energy': 'dinh-duong-chay',
        # BABY brands
        'huggies': 'do-tre-em', 'pampers': 'do-tre-em', 'merries': 'do-tre-em',
        'moony': 'do-tre-em', 'bobby': 'do-tre-em', 'mamypoko': 'do-tre-em',
        'similac': 'do-tre-em', 'friso': 'do-tre-em',
        # HOME brands
        'tefal': 'nha-bep', 'sunhouse': 'nha-bep', 'lock lock': 'nha-bep',
        'air fryer': 'nha-bep', 'noi chien khong dau': 'nha-bep',
        'robot hut bui': 've-sinh', 'may hut bui': 've-sinh',
        'daikin': 'dien-gia-dung', 'mitsubishi': 'dien-gia-dung',
        # HEALTH brands
        'blackmores': 'thuc-pham-chuc-nang', 'swisse': 'thuc-pham-chuc-nang',
        'nature made': 'thuc-pham-chuc-nang', 'now foods': 'thuc-pham-chuc-nang',
        'centrum': 'thuc-pham-chuc-nang', 'kirkland': 'thuc-pham-chuc-nang',
        'omron': 'dung-cu-y-te', 'microlife': 'dung-cu-y-te',
        # TRAVEL brands
        'vinpearl': 'resort', 'six senses': 'resort', 'jw marriott': 'resort',
    }
    return index, zone_keywords, power_keywords, normalize

def _kw_match(kw, text, words):
    """Match a keyword against text with word-boundary awareness.
    Short keywords (< 5 chars): whole-word match only.
    Long keywords (>= 5 chars): substring match OK (more specific).
    Multi-word keywords: all words must be present.
    Returns score (0 = no match)."""
    if not kw or len(kw) < 2:
        return 0
    kw_words = kw.split()
    if len(kw_words) > 1:
        # Multi-word keyword: all words must be present as whole words
        if all(w in words for w in kw_words if len(w) >= 2):
            return len(kw) + 5
        return 0
    # Single-word keyword
    if len(kw) < 5:
        # Short: whole-word match only (avoid "day" matching "day dien")
        if kw in words:
            return len(kw) + 2
        return 0
    else:
        # Long (>= 5 chars): substring OK, more specific
        if kw in text:
            return len(kw) + 4
        return 0

def _match_product_to_part(product_name, category, part_index, zone_keywords, normalize_fn, power_keywords=None):
    """Match a product to best Part across all verticals.
    Returns (part_id, detected_category) tuple. part_id can be None.
    Uses 4-phase strategy: Power keywords → Part keywords → Zone keywords → CSV category."""
    text = normalize_fn(f"{product_name} {category}")
    words = set(text.split())

    # Phase 0: Power keywords — brand names so distinctive 1 hit is enough
    if power_keywords:
        best_pw_zone = None
        best_pw_score = 0
        for pw, zone_slug in power_keywords.items():
            s = _kw_match(pw, text, words)
            if s > 0 and s > best_pw_score:
                best_pw_score = s
                best_pw_zone = zone_slug
        if best_pw_zone and best_pw_score >= 6:
            first_part = Part.query.join(Zone).filter(Zone.slug == best_pw_zone).first()
            if first_part:
                return first_part.id, best_pw_zone

    best_part_id = None
    best_score = 0
    best_hits = 0
    best_entry = None

    # Phase 1: match against Part keywords
    for entry in part_index:
        score = 0
        hits = 0
        for kw in entry['keywords']:
            s = _kw_match(kw, text, words)
            if s > 0:
                score += s
                hits += 1
        if score > best_score:
            best_score = score
            best_hits = hits
            best_part_id = entry['part_id']
            best_entry = entry

    # Require score >= 10 AND at least 2 keyword hits for confidence
    if best_score >= 10 and best_hits >= 2:
        cat = best_entry['zone_slug'] if best_entry else ''
        return best_part_id, cat

    # Phase 2: fallback to Zone-level keywords
    best_zone_slug = None
    best_zone_score = 0
    best_zone_hits = 0
    for zone_slug, kws in zone_keywords.items():
        score = 0
        hits = 0
        for kw in kws:
            s = _kw_match(kw, text, words)
            if s > 0:
                score += s
                hits += 1
        if score > best_zone_score:
            best_zone_score = score
            best_zone_hits = hits
            best_zone_slug = zone_slug

    # Require score >= 8 AND at least 1 hit (relaxed for zone-level)
    if best_zone_slug and best_zone_score >= 8 and best_zone_hits >= 1:
        first_part = Part.query.join(Zone).filter(Zone.slug == best_zone_slug).first()
        if first_part:
            return first_part.id, best_zone_slug
        return None, best_zone_slug

    # Phase 3: use CSV category directly as fallback (exact zone slug match only)
    if category:
        cat_norm = normalize_fn(category)
        cat_slug = cat_norm.replace(' ', '-')
        for zone_slug in zone_keywords:
            if cat_slug == zone_slug or cat_norm == zone_slug.replace('-', ' '):
                first_part = Part.query.join(Zone).filter(Zone.slug == zone_slug).first()
                if first_part:
                    return first_part.id, zone_slug

    return None, ''


def _clean_csv_value(val):
    """Clean CSV cell value — handle nan/null/None from pandas/Excel exports."""
    if val is None:
        return ''
    s = str(val).strip()
    if s.lower() in ('nan', 'null', 'none', 'n/a', 'na', '#n/a', ''):
        return ''
    return s

def _get_or_create_hub_part(category_name):
    """Auto-create Hub vertical + Zone + Part for unmatched products.
    Products go into Hub first, can be reassigned to proper verticals later.
    Uses category_name from CSV to create/find a Zone."""
    cat = _clean_csv_value(category_name)
    if not cat:
        cat = 'Chua phan loai'
    cat_slug = slugify(cat)
    if not cat_slug:
        cat_slug = 'chua-phan-loai'

    # Find or create the Hub vertical
    hub = Vertical.query.filter_by(slug='hub').first()
    if not hub:
        hub = Vertical(name='Hub', slug='hub', icon='🛒', color='#10b981',
                        description='Kho san pham tong — chua phan loai vao vertical', status='published')
        db.session.add(hub)
        db.session.flush()

    # Find or create default segment
    seg = Segment.query.filter_by(vertical_id=hub.id, slug='san-pham').first()
    if not seg:
        seg = Segment(vertical_id=hub.id, name='San pham', slug='san-pham',
                       icon='📦', description='San pham import tu CSV', order=0)
        db.session.add(seg)
        db.session.flush()

    # Find or create Zone from category
    zone = Zone.query.filter_by(segment_id=seg.id, slug=cat_slug).first()
    if not zone:
        zone = Zone(segment_id=seg.id, name=cat[:100], slug=cat_slug,
                     icon='📁', description=f'Tu dong tao tu CSV category: {cat}', order=0)
        db.session.add(zone)
        db.session.flush()

    # Find or create a generic Part in this Zone
    part = Part.query.filter_by(zone_id=zone.id).first()
    if not part:
        part = Part(zone_id=zone.id, name_vi=cat[:200], slug=cat_slug,
                     description=f'San pham {cat}', status='published')
        db.session.add(part)
        db.session.flush()

    return part.id

@app.route('/admin/products/import-csv', methods=['GET', 'POST'])
def admin_products_import_csv():
    """Import products from CSV — all products go into hub, auto-create categories if needed"""
    if request.method == 'POST':
        if 'csv_file' not in request.files:
            flash('Không có file CSV', 'error')
            return redirect(request.url)

        file = request.files['csv_file']
        if file.filename == '':
            flash('Chưa chọn file', 'error')
            return redirect(request.url)

        if not file.filename.endswith('.csv'):
            flash('File phải có định dạng CSV', 'error')
            return redirect(request.url)

        # Read CSV
        import csv
        import io

        stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_reader = csv.DictReader(stream)
        # Normalize fieldnames to lowercase for case-insensitive matching
        if csv_reader.fieldnames:
            csv_reader.fieldnames = [f.strip().lower() for f in csv_reader.fieldnames]

        # Get form data
        mapping_mode = request.form.get('mapping_mode', 'auto')  # auto or manual
        part_id = request.form.get('part_id')
        network = request.form.get('network', 'shopee')
        apply_deeplink = 'apply_deeplink' in request.form

        if mapping_mode == 'manual' and not part_id:
            flash('Chưa chọn Part để gắn sản phẩm', 'error')
            return redirect(request.url)

        # Build auto-mapping index
        part_index = None
        zone_kw = None
        normalize_fn = None
        if mapping_mode == 'auto':
            part_index, zone_kw, power_kw, normalize_fn = _build_part_keyword_index()

        # Get network for deeplink
        network_obj = None
        if apply_deeplink:
            network_obj = AffiliateNetwork.query.filter_by(slug=network).first()

        count = 0
        skipped = 0
        auto_created = 0
        mapped_zones = {}  # Track mapping stats: zone_name -> count
        hub_zones = {}  # Track auto-created hub categories
        for row in csv_reader:
            # CSV format: sku, name, url, price, discount, image, desc, category
            # Clean all values (handle nan/null from pandas exports)
            product_name = _clean_csv_value(row.get('name', ''))[:200]
            url = _clean_csv_value(row.get('url', ''))
            category = _clean_csv_value(row.get('category', ''))

            if not product_name and not url:
                skipped += 1
                continue

            # Apply deeplink if enabled and template exists
            if apply_deeplink and network_obj and network_obj.deeplink_template and url:
                import urllib.parse
                url = network_obj.deeplink_template.replace('{url}', urllib.parse.quote(url))

            price_val = 0
            price_raw = _clean_csv_value(row.get('price', ''))
            try:
                price_val = float(price_raw) if price_raw else 0
            except (ValueError, TypeError):
                pass
            image_url = _clean_csv_value(row.get('image', ''))

            if mapping_mode == 'auto':
                # Step 1: ALWAYS create in Hub (from CSV category)
                hub_part_id = _get_or_create_hub_part(category)
                cat_label = category or 'Chua phan loai'
                hub_zones[cat_label] = hub_zones.get(cat_label, 0) + 1

                try:
                    al_hub = AffiliateLink(
                        part_id=hub_part_id,
                        network=network,
                        product_name=product_name,
                        url=url,
                        price=price_val,
                        image_url=image_url,
                        is_active=True,
                        category=slugify(category) if category else 'chua-phan-loai',
                    )
                    db.session.add(al_hub)
                    count += 1
                except (ValueError, TypeError):
                    skipped += 1
                    continue

                # Step 2: ALSO match to vertical (if keywords match)
                if part_index:
                    matched_part_id, detected_category = _match_product_to_part(
                        product_name, category, part_index, zone_kw, normalize_fn, power_kw)
                    if matched_part_id:
                        al_vert = AffiliateLink(
                            part_id=matched_part_id,
                            network=network,
                            product_name=product_name,
                            url=url,
                            price=price_val,
                            image_url=image_url,
                            is_active=True,
                            category=detected_category,
                        )
                        db.session.add(al_vert)
                        auto_created += 1
                        # Track vertical mapping stats
                        for entry in (part_index or []):
                            if entry['part_id'] == matched_part_id:
                                zone_label = f"{entry['vert_name']} › {entry['zone_name']} › {entry['part_name']}"
                                mapped_zones[zone_label] = mapped_zones.get(zone_label, 0) + 1
                                break
            else:
                # Manual mode: single Part assignment
                try:
                    al = AffiliateLink(
                        part_id=int(part_id),
                        network=network,
                        product_name=product_name,
                        url=url,
                        price=price_val,
                        image_url=image_url,
                        is_active=True,
                        category=category[:100],
                    )
                    db.session.add(al)
                    count += 1
                except (ValueError, TypeError):
                    skipped += 1

            # Batch commit every 200 rows to reduce lock duration
            if (count + skipped) % 200 == 0:
                db.session.commit()

        db.session.commit()

        # Build result message
        msg = f'Import {count} sản phẩm vào Hub thành công!'
        if auto_created:
            msg += f' + {auto_created} cũng match vào vertical'
        if skipped:
            msg += f' ({skipped} bỏ qua - thiếu tên/url)'
        if hub_zones:
            top_hub = sorted(hub_zones.items(), key=lambda x: -x[1])[:5]
            detail = ', '.join(f'{name}: {cnt}' for name, cnt in top_hub)
            msg += f' | Hub categories: {detail}'
        if mapped_zones:
            top_vert = sorted(mapped_zones.items(), key=lambda x: -x[1])[:5]
            detail = ', '.join(f'{name}: {cnt}' for name, cnt in top_vert)
            msg += f' | Vertical match: {detail}'
        flash(msg, 'success')
        return redirect(url_for('admin_products_hub'))

    # GET request - show form
    parts_tree = []
    for v in Vertical.query.all():
        for s in v.segments:
            for z in s.zones:
                for p in z.parts:
                    parts_tree.append({
                        'id': p.id,
                        'label': f'{v.icon} {v.name} › {s.name} › {z.name} › {p.name_vi}'
                    })

    networks = AffiliateNetwork.query.all()
    return render_template('admin/products_import_csv.html',
        parts_tree=parts_tree, networks=networks)

@app.route('/admin/products/datafeeds', methods=['GET', 'POST'])
def admin_products_datafeeds():
    """Browse & import products from AccessTrade Datafeeds API"""
    from accesstrade_integration import get_accesstrade_api

    api = get_accesstrade_api()
    if not api:
        flash('Chua cau hinh AccessTrade API key. Vao Affiliate Network de thiet lap.', 'error')
        return redirect(url_for('admin_products_hub'))

    # ── POST: Import selected products ──
    if request.method == 'POST':
        selected = request.form.getlist('selected_products')
        if not selected:
            flash('Chua chon san pham nao de import', 'warning')
            return redirect(request.url)

        import json as _json
        network = 'accesstrade'
        apply_deeplink = 'apply_deeplink' in request.form
        network_obj = AffiliateNetwork.query.filter_by(slug='accesstrade').first()

        # Build auto-mapping index
        part_index, zone_kw, power_kw, normalize_fn = _build_part_keyword_index()

        count = 0
        matched = 0
        for item_json in selected:
            try:
                item = _json.loads(item_json)
            except (ValueError, TypeError):
                continue

            product_name = (item.get('name') or '')[:200]
            url = item.get('url') or item.get('deep_link') or ''
            image_url = item.get('image') or ''
            category = item.get('category') or ''

            if not product_name and not url:
                continue

            # Use deep_link (affiliate link) if available, else apply deeplink template
            aff_url = item.get('deep_link') or url
            if apply_deeplink and network_obj and network_obj.deeplink_template and url and not item.get('deep_link'):
                import urllib.parse
                aff_url = network_obj.deeplink_template.replace('{url}', urllib.parse.quote(url))

            price_val = 0
            try:
                price_val = float(item.get('price') or item.get('promotion_price') or 0)
            except (ValueError, TypeError):
                pass

            # Always create in Hub
            hub_part_id = _get_or_create_hub_part(category)
            al_hub = AffiliateLink(
                part_id=hub_part_id,
                network=network,
                product_name=product_name,
                url=aff_url,
                price=price_val,
                image_url=image_url,
                is_active=True,
                category=slugify(category) if category else 'chua-phan-loai',
            )
            db.session.add(al_hub)
            count += 1

            # Also match to vertical
            if part_index:
                matched_part_id, detected_cat = _match_product_to_part(
                    product_name, category, part_index, zone_kw, normalize_fn, power_kw)
                if matched_part_id:
                    al_vert = AffiliateLink(
                        part_id=matched_part_id,
                        network=network,
                        product_name=product_name,
                        url=aff_url,
                        price=price_val,
                        image_url=image_url,
                        is_active=True,
                        category=detected_cat,
                    )
                    db.session.add(al_vert)
                    matched += 1

        db.session.commit()
        msg = f'Import {count} san pham tu Datafeeds thanh cong!'
        if matched:
            msg += f' + {matched} match vao vertical'
        flash(msg, 'success')
        return redirect(url_for('admin_products_hub'))

    # ── GET: Browse datafeeds ──
    keyword = request.args.get('keyword', '')
    domain = request.args.get('domain', '')
    campaign = request.args.get('campaign', '')
    page = request.args.get('page', 1, type=int)
    limit = 50

    products = []
    total = 0

    # Only call API if there's a search query
    if keyword or domain or campaign:
        result = api.get_datafeeds(
            keyword=keyword or None,
            domain=domain or None,
            campaign=campaign or None,
            page=page,
            limit=limit,
        )
        products = result.get('data', [])
        total = result.get('total', 0)

    # Get campaigns for dropdown filter
    campaigns = api.get_campaigns(limit=100, status=1)

    total_pages = (total + limit - 1) // limit if total > 0 else 0

    return render_template('admin/products_datafeeds.html',
        products=products, total=total, campaigns=campaigns,
        keyword=keyword, domain=domain, campaign=campaign,
        page=page, total_pages=total_pages)


@app.route('/admin/api/create-tracking-link', methods=['POST'])
def admin_api_create_tracking_link():
    """API endpoint: convert product URLs to affiliate tracking links."""
    from accesstrade_integration import get_accesstrade_api
    api = get_accesstrade_api()
    if not api:
        return jsonify({'success': False, 'error': 'AccessTrade API not configured'}), 400

    data = request.get_json() or {}
    campaign_id = data.get('campaign_id', '')
    urls = data.get('urls', [])
    if not campaign_id:
        return jsonify({'success': False, 'error': 'campaign_id is required'}), 400
    if not urls:
        return jsonify({'success': False, 'error': 'urls list is required'}), 400

    result = api.create_tracking_link(
        campaign_id=campaign_id,
        urls=urls,
        utm_source=data.get('utm_source'),
        utm_campaign=data.get('utm_campaign'),
    )
    return jsonify(result)


@app.route('/admin/api/order-list')
def admin_api_order_list():
    """API endpoint: get AccessTrade order list v2."""
    from accesstrade_integration import get_accesstrade_api
    api = get_accesstrade_api()
    if not api:
        return jsonify({'data': [], 'total': 0, 'error': 'AccessTrade API not configured'}), 400

    since = request.args.get('since', '')
    until = request.args.get('until', '')
    if not since or not until:
        return jsonify({'data': [], 'total': 0, 'error': 'since and until are required'}), 400

    result = api.get_order_list(
        since=since,
        until=until,
        page=request.args.get('page', 1, type=int),
        limit=request.args.get('limit', 30, type=int),
        status=request.args.get('status', None, type=int),
        merchant=request.args.get('merchant', None),
    )
    return jsonify(result)


@app.route('/admin/api/tiktok-search')
def admin_api_tiktok_search():
    """API endpoint: search TikTok Shop products."""
    from accesstrade_integration import get_accesstrade_api
    api = get_accesstrade_api()
    if not api:
        return jsonify({'products': [], 'total_count': 0, 'error': 'API not configured'}), 400

    keywords = request.args.get('q', '')
    if not keywords:
        return jsonify({'products': [], 'total_count': 0, 'error': 'q is required'}), 400

    result = api.tiktok_search_products(
        title_keywords=keywords,
        sort_field=request.args.get('sort', 'RECOMMENDED'),
        limit=request.args.get('limit', 20, type=int),
        page_token=request.args.get('page_token', None),
    )
    return jsonify(result)


@app.route('/admin/api/tiktok-create-link', methods=['POST'])
def admin_api_tiktok_create_link():
    """API endpoint: create TikTok Shop affiliate link."""
    from accesstrade_integration import get_accesstrade_api
    api = get_accesstrade_api()
    if not api:
        return jsonify({'status': False, 'error': 'API not configured'}), 400

    data = request.get_json() or {}
    product_url = data.get('product_url', '')
    if not product_url:
        return jsonify({'status': False, 'error': 'product_url is required'}), 400

    result = api.tiktok_create_link(
        product_url=product_url,
        product_id=data.get('product_id'),
        utm_source=data.get('utm_source'),
    )
    return jsonify(result)


@app.route('/admin/products/reclassify-hub', methods=['POST'])
def admin_products_reclassify_hub():
    """Re-run matching on Hub products → move matched ones to verticals.
    Called after creating new verticals/parts to classify existing Hub products."""
    hub = Vertical.query.filter_by(slug='hub').first()
    if not hub:
        flash('Chưa có vertical Hub nào.', 'warning')
        return redirect(url_for('admin_products_hub'))

    # Get all Hub segment/zone/part IDs
    hub_part_ids = set()
    for seg in hub.segments:
        for zone in seg.zones:
            for part in zone.parts:
                hub_part_ids.add(part.id)

    if not hub_part_ids:
        flash('Hub không có sản phẩm nào.', 'warning')
        return redirect(url_for('admin_products_hub'))

    # Get all products in Hub
    hub_products = AffiliateLink.query.filter(AffiliateLink.part_id.in_(hub_part_ids)).all()
    if not hub_products:
        flash('Hub không có sản phẩm nào cần phân loại.', 'info')
        return redirect(url_for('admin_products_hub'))

    # Build matching index (excluding Hub parts)
    part_index, zone_kw, power_kw, normalize_fn = _build_part_keyword_index()
    # Filter out Hub parts from index
    part_index = [e for e in part_index if e['vert_slug'] != 'hub']

    if not part_index and not zone_kw:
        flash('Chưa có vertical/zone nào để match. Hãy tạo vertical trước.', 'warning')
        return redirect(url_for('admin_products_hub'))

    matched = 0
    mapped_zones = {}
    for al in hub_products:
        matched_part_id, detected_category = _match_product_to_part(
            al.product_name, al.category or '', part_index, zone_kw, normalize_fn, power_kw)
        if matched_part_id:
            # Create a copy in the matched vertical (keep original in Hub)
            existing = AffiliateLink.query.filter_by(
                part_id=matched_part_id, url=al.url).first()
            if not existing:
                al_vert = AffiliateLink(
                    part_id=matched_part_id,
                    network=al.network,
                    product_name=al.product_name,
                    url=al.url,
                    price=al.price,
                    image_url=al.image_url,
                    is_active=True,
                    category=detected_category,
                )
                db.session.add(al_vert)
                matched += 1
                # Track stats
                for entry in part_index:
                    if entry['part_id'] == matched_part_id:
                        label = f"{entry['vert_name']} › {entry['zone_name']} › {entry['part_name']}"
                        mapped_zones[label] = mapped_zones.get(label, 0) + 1
                        break

    db.session.commit()

    msg = f'Re-classify: {matched}/{len(hub_products)} sản phẩm Hub được match vào vertical.'
    if mapped_zones:
        top = sorted(mapped_zones.items(), key=lambda x: -x[1])[:8]
        detail = ', '.join(f'{name}: {cnt}' for name, cnt in top)
        msg += f' | {detail}'
    flash(msg, 'success' if matched > 0 else 'info')
    return redirect(url_for('admin_products_hub'))

@app.route('/admin/scheduled-imports')
def admin_scheduled_imports():
    """List all scheduled CSV imports"""
    jobs = ScheduledCSVImport.query.order_by(ScheduledCSVImport.created_at.desc()).all()
    return render_template('admin/scheduled_imports.html', jobs=jobs)

@app.route('/admin/scheduled-import/create', methods=['POST'])
def admin_scheduled_import_create():
    """Create a new scheduled CSV import job"""
    from datetime import timedelta

    job = ScheduledCSVImport(
        name=request.form['name'],
        csv_url=request.form['csv_url'],
        part_id=int(request.form['part_id']),
        network=request.form.get('network', 'shopee'),
        apply_deeplink='apply_deeplink' in request.form,
        update_interval_days=int(request.form.get('update_interval_days', 7)),
        is_active=True
    )

    # Set next import time
    job.next_import_at = datetime.utcnow() + timedelta(days=job.update_interval_days)

    db.session.add(job)
    db.session.commit()

    # Run import now if requested
    if 'run_now' in request.form:
        return redirect(url_for('admin_scheduled_import_run', job_id=job.id))

    flash(f'Đã tạo scheduled import: {job.name}', 'success')
    return redirect(url_for('admin_scheduled_imports'))

@app.route('/admin/scheduled-import/<int:job_id>/run')
def admin_scheduled_import_run(job_id):
    """Manually trigger a scheduled import"""
    from datetime import timedelta
    import urllib.request
    import csv
    import io

    job = ScheduledCSVImport.query.get_or_404(job_id)

    try:
        # Fetch CSV from URL
        req = urllib.request.Request(job.csv_url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=30) as response:
            csv_content = response.read().decode('utf-8')

        # Parse CSV
        stream = io.StringIO(csv_content)
        csv_reader = csv.DictReader(stream)

        # Get network for deeplink
        network_obj = None
        if job.apply_deeplink:
            network_obj = AffiliateNetwork.query.filter_by(slug=job.network).first()

        # Delete old products for this part from this network (optional - for refresh)
        # AffiliateLink.query.filter_by(part_id=job.part_id, network=job.network).delete()

        count = 0
        for row in csv_reader:
            url = row.get('url', '')

            # Apply deeplink if enabled and template exists
            if job.apply_deeplink and network_obj and network_obj.deeplink_template and url:
                import urllib.parse
                url = network_obj.deeplink_template.replace('{url}', urllib.parse.quote(url))

            # Check if product already exists (by URL)
            existing = AffiliateLink.query.filter_by(part_id=job.part_id, url=url).first()
            if existing:
                # Update existing product
                existing.product_name = row.get('name', '')[:200]
                existing.price = float(row.get('price', 0))
                existing.image_url = row.get('image', '')
            else:
                # Create new product
                al = AffiliateLink(
                    part_id=job.part_id,
                    network=job.network,
                    product_name=row.get('name', '')[:200],
                    url=url,
                    price=float(row.get('price', 0)),
                    image_url=row.get('image', ''),
                    is_active=True
                )
                db.session.add(al)
            count += 1

        # Update job status
        job.last_import_at = datetime.utcnow()
        job.last_import_count = count
        job.last_import_status = 'success'
        job.last_error = ''
        job.next_import_at = datetime.utcnow() + timedelta(days=job.update_interval_days)

        db.session.commit()
        flash(f'✓ Import thành công {count} sản phẩm từ {job.name}', 'success')

    except Exception as e:
        job.last_import_status = 'error'
        job.last_error = str(e)
        db.session.commit()
        flash(f'Lỗi khi import: {str(e)}', 'error')

    return redirect(url_for('admin_scheduled_imports'))

@app.route('/admin/scheduled-import/<int:job_id>/toggle', methods=['POST'])
def admin_scheduled_import_toggle(job_id):
    """Toggle scheduled import active status"""
    job = ScheduledCSVImport.query.get_or_404(job_id)
    job.is_active = not job.is_active
    db.session.commit()
    flash(f'Đã {"kích hoạt" if job.is_active else "tắt"} scheduled import: {job.name}', 'success')
    return redirect(url_for('admin_scheduled_imports'))

@app.route('/admin/scheduled-import/<int:job_id>/delete', methods=['POST'])
def admin_scheduled_import_delete(job_id):
    """Delete a scheduled import job"""
    job = ScheduledCSVImport.query.get_or_404(job_id)
    name = job.name
    db.session.delete(job)
    db.session.commit()
    flash(f'Đã xóa scheduled import: {name}', 'success')
    return redirect(url_for('admin_scheduled_imports'))

# =============================================
# ADMIN — VIDEO PRODUCTION
# =============================================
@app.route('/admin/video')
def admin_video():
    """Redirect standalone video page to Tools Hub"""
    return redirect(url_for('admin_tools', tab='video'))

@app.route('/admin/video/channels')
def admin_video_channels():
    channels = SocialChannel.query.all()
    verticals = Vertical.query.all()
    return render_template('admin/video_channels.html', channels=channels, verticals=verticals)

@app.route('/admin/video/channel/add', methods=['POST'])
def admin_video_channel_add():
    ch = SocialChannel(
        vertical_id=int(request.form['vertical_id']),
        platform=request.form['platform'],
        channel_name=request.form['channel_name'],
        channel_url=request.form.get('channel_url',''),
        api_key=request.form.get('api_key',''),
        status='connected' if request.form.get('api_key') else 'disconnected'
    )
    db.session.add(ch)
    db.session.commit()
    flash(f'Da them channel: {ch.channel_name}', 'success')
    return redirect(url_for('admin_video_channels'))

@app.route('/admin/video/channel/<int:cid>/delete', methods=['POST'])
def admin_video_channel_delete(cid):
    ch = SocialChannel.query.get_or_404(cid)
    db.session.delete(ch)
    db.session.commit()
    flash('Da xoa channel', 'success')
    return redirect(url_for('admin_video_channels'))

@app.route('/admin/video/channel/<int:cid>/connect', methods=['POST'])
def admin_video_channel_connect(cid):
    ch = SocialChannel.query.get_or_404(cid)
    ch.api_key = request.form.get('api_key','')
    ch.status = 'connected' if ch.api_key else 'disconnected'
    db.session.commit()
    flash(f'{ch.channel_name} -> {ch.status}', 'success')
    return redirect(url_for('admin_video_channels'))

@app.route('/admin/video/create', methods=['GET','POST'])
def admin_video_create():
    verticals = Vertical.query.all()
    parts = Part.query.order_by(Part.name_vi).all()
    if request.method == 'POST':
        vp = VideoProject(
            title=request.form['title'],
            vertical_slug=request.form.get('vertical_slug',''),
            part_id=int(request.form['part_id']) if request.form.get('part_id') else None,
            video_type=request.form.get('video_type','short'),
            duration=request.form.get('duration','60s'),
            script=request.form.get('script',''),
            voiceover_text=request.form.get('voiceover_text',''),
            ai_provider=request.form.get('ai_provider','openai'),
            caption=request.form.get('caption',''),
            hashtags=request.form.get('hashtags',''),
            status='draft'
        )
        db.session.add(vp)
        db.session.commit()
        flash(f'Da tao video project: {vp.title}', 'success')
        return redirect(url_for('admin_video_detail', vid=vp.id))
    return render_template('admin/video_create.html', verticals=verticals, parts=parts)

@app.route('/admin/video/<int:vid>')
def admin_video_detail(vid):
    v = VideoProject.query.get_or_404(vid)
    channels = SocialChannel.query.filter_by(status='connected').all()
    if v.vertical_slug:
        channels = [c for c in channels if c.vertical and c.vertical.slug == v.vertical_slug] or channels
    return render_template('admin/video_detail.html', video=v, channels=channels)

@app.route('/admin/video/<int:vid>/delete', methods=['POST'])
def admin_video_delete(vid):
    v = VideoProject.query.get_or_404(vid)
    db.session.delete(v)
    db.session.commit()
    flash('Da xoa video', 'success')
    return redirect(url_for('admin_tools', tab='video'))

@app.route('/admin/videos/bulk-delete', methods=['POST'])
def admin_videos_bulk_delete():
    """Bulk delete videos by IDs (JSON)"""
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify(ok=False, error='Khong co video nao'), 400
    deleted = VideoProject.query.filter(VideoProject.id.in_([int(i) for i in ids])).delete(synchronize_session=False)
    db.session.commit()
    return jsonify(ok=True, deleted=deleted)

@app.route('/admin/video/<int:vid>/publish', methods=['POST'])
def admin_video_publish(vid):
    v = VideoProject.query.get_or_404(vid)
    channel_ids = request.form.getlist('channel_ids')
    for cid in channel_ids:
        ch = SocialChannel.query.get(int(cid))
        if ch:
            vp = VideoPublish(
                video_id=v.id, channel_id=ch.id, platform=ch.platform,
                status='queued',
                views=random.randint(100,5000), likes=random.randint(10,500),
                shares=random.randint(5,100), comments=random.randint(2,50),
                click_throughs=random.randint(5,200)
            )
            db.session.add(vp)
    v.status = 'published'
    db.session.commit()
    flash(f'Da publish len {len(channel_ids)} channels', 'success')
    return redirect(url_for('admin_video_detail', vid=v.id))

@app.route('/admin/video/analytics')
def admin_video_analytics():
    publishes = VideoPublish.query.order_by(VideoPublish.id.desc()).all()
    channels = SocialChannel.query.all()
    total_views = db.session.query(db.func.sum(VideoPublish.views)).scalar() or 0
    total_clicks = db.session.query(db.func.sum(VideoPublish.click_throughs)).scalar() or 0
    return render_template('admin/video_analytics.html',
        publishes=publishes, channels=channels, total_views=total_views, total_clicks=total_clicks)

# =============================================
# ADMIN — HOTELS (Quản lý khách sạn)
# =============================================
@app.route('/admin/hotels')
def admin_hotels():
    """Redirect standalone hotels page to Hotels Hub"""
    return redirect(url_for('admin_hotels_hub'))

@app.route('/admin/hotel/new', methods=['GET','POST'])
def admin_hotel_new():
    if request.method == 'POST':
        h = Hotel(
            name=request.form['name'],
            slug=slugify(request.form['name'])[:60],
            destination=request.form['destination'],
            destination_name=request.form.get('destination_name',''),
            stars=int(request.form.get('stars',4)),
            district=request.form.get('district',''),
            description=request.form.get('description',''),
            amenities=request.form.get('amenities',''),
            rating=float(request.form.get('rating',8.0)),
            reviews_count=int(request.form.get('reviews_count',0)),
            price_from=float(request.form.get('price_from',0)),
            image_url=request.form.get('image_url',''),
            agoda_url=request.form.get('agoda_url',''),
            booking_url=request.form.get('booking_url',''),
            traveloka_url=request.form.get('traveloka_url',''),
            is_active='is_active' in request.form,
            is_featured='is_featured' in request.form
        )
        db.session.add(h); db.session.commit()
        flash(f'Da them: {h.name}', 'success')
        return redirect(url_for('admin_hotels_hub'))
    return render_template('admin/hotel_form.html', hotel=None)

@app.route('/admin/hotel/<int:hid>/edit', methods=['GET','POST'])
def admin_hotel_edit(hid):
    h = Hotel.query.get_or_404(hid)
    if request.method == 'POST':
        h.name = request.form['name']
        h.destination = request.form['destination']
        h.destination_name = request.form.get('destination_name','')
        h.stars = int(request.form.get('stars',4))
        h.district = request.form.get('district','')
        h.description = request.form.get('description','')
        h.amenities = request.form.get('amenities','')
        h.rating = float(request.form.get('rating',8.0))
        h.reviews_count = int(request.form.get('reviews_count',0))
        h.price_from = float(request.form.get('price_from',0))
        h.image_url = request.form.get('image_url','')
        h.agoda_url = request.form.get('agoda_url','')
        h.booking_url = request.form.get('booking_url','')
        h.traveloka_url = request.form.get('traveloka_url','')
        h.is_active = 'is_active' in request.form
        h.is_featured = 'is_featured' in request.form
        db.session.commit()
        flash(f'Da cap nhat: {h.name}', 'success')
        return redirect(url_for('admin_hotels_hub'))
    return render_template('admin/hotel_form.html', hotel=h)

@app.route('/admin/hotel/<int:hid>/toggle', methods=['POST'])
def admin_hotel_toggle(hid):
    h = Hotel.query.get_or_404(hid)
    h.is_active = not h.is_active; db.session.commit()
    return redirect(url_for('admin_hotels_hub'))

@app.route('/admin/hotel/<int:hid>/delete', methods=['POST'])
def admin_hotel_delete(hid):
    h = Hotel.query.get_or_404(hid)
    db.session.delete(h); db.session.commit()
    flash('Da xoa khach san', 'success')
    return redirect(url_for('admin_hotels_hub'))

@app.route('/admin/hotels/bulk-delete', methods=['POST'])
def admin_hotels_bulk_delete():
    ids = request.form.getlist('hotel_ids')
    if not ids:
        flash('Chua chon khach san nao', 'warning')
        return redirect(url_for('admin_hotels_hub'))
    count = Hotel.query.filter(Hotel.id.in_([int(i) for i in ids])).delete(synchronize_session=False)
    db.session.commit()
    flash(f'Da xoa {count} khach san', 'success')
    return redirect(url_for('admin_hotels_hub'))

@app.route('/admin/hotels/delete-all', methods=['POST'])
def admin_hotels_delete_all():
    count = Hotel.query.count()
    if count == 0:
        flash('Khong co khach san nao de xoa', 'warning')
        return redirect(url_for('admin_hotels_hub'))
    Hotel.query.delete(synchronize_session=False)
    db.session.commit()
    flash(f'Da xoa tat ca {count} khach san', 'success')
    return redirect(url_for('admin_hotels_hub'))

@app.route('/admin/hotels/bulk-toggle', methods=['POST'])
def admin_hotels_bulk_toggle():
    ids = request.form.getlist('hotel_ids')
    action = request.form.get('action', '')
    if not ids:
        flash('Chua chon khach san nao', 'warning')
        return redirect(url_for('admin_hotels_hub'))
    new_state = action == 'activate'
    hotels = Hotel.query.filter(Hotel.id.in_([int(i) for i in ids])).all()
    for h in hotels:
        h.is_active = new_state
    db.session.commit()
    label = 'bat' if new_state else 'tat'
    flash(f'Da {label} {len(hotels)} khach san', 'success')
    return redirect(url_for('admin_hotels_hub'))

# =============================================
# ADMIN — HOTEL SYNC (Agoda API)
# =============================================

@app.route('/admin/hotel-sync')
def admin_hotel_sync():
    """Dashboard for Agoda hotel sync"""
    from agoda_integration import get_agoda_api, AGODA_CITY_IDS, AGODA_CITY_NAMES

    # Auto-populate default Agoda credentials if not saved yet
    _default_cid = '1959245'
    _default_key = '1959245:5669c3b3-2865-4591-ba56-1b02a3c04082'
    if not SiteSettings.get('agoda_cid', ''):
        SiteSettings.set_val('agoda_cid', _default_cid, 'api')
        SiteSettings.set_val('agoda_api_key', _default_key, 'api')
        SiteSettings.set_val('agoda_enabled', '1', 'general')
        import agoda_integration
        agoda_integration._api_instance = None
        db.session.commit()

    api = get_agoda_api()
    api_connected = api is not None

    total_hotels = Hotel.query.count()
    total_agoda = Hotel.query.filter_by(source='agoda_api').count()
    total_manual = Hotel.query.filter(Hotel.source != 'agoda_api').count()
    total_active = Hotel.query.filter_by(is_active=True).count()

    # Credentials info
    cid = SiteSettings.get('agoda_cid', '')
    has_key = bool(SiteSettings.get('agoda_api_key', ''))

    # Full destination list: 34 provinces + Agoda tourism destinations
    from agoda_integration import VIETNAM_DESTINATIONS
    destinations = [
        {'slug': slug, 'name': name, 'city_id': cid, 'province_code': ''}
        for name, slug, cid in VIETNAM_DESTINATIONS
    ]

    # Fast Sync uses ALL destinations (not just AGODA_CITY_IDS)
    agoda_destinations = [
        {'slug': d['slug'], 'name': d['name'], 'city_id': d['city_id']}
        for d in destinations if d['city_id']
    ]

    # Paginated Agoda hotel list with image filter
    page = request.args.get('page', 1, type=int)
    per_page = 30
    f_image = request.args.get('image', '')
    rq = Hotel.query.filter_by(source='agoda_api')
    if f_image == 'missing':
        rq = rq.filter(db.or_(Hotel.image_url == '', Hotel.image_url == None))
    elif f_image == 'has':
        rq = rq.filter(Hotel.image_url != '', Hotel.image_url != None)
    agoda_no_image = Hotel.query.filter_by(source='agoda_api').filter(db.or_(Hotel.image_url == '', Hotel.image_url == None)).count()
    recent_total = rq.count()
    recent = rq.order_by(Hotel.id.desc()).offset((page - 1) * per_page).limit(per_page).all()
    recent_pages = (recent_total + per_page - 1) // per_page

    return render_template('admin/hotel_sync.html',
        api_connected=api_connected, cid=cid, has_key=has_key,
        total_hotels=total_hotels, total_agoda=total_agoda,
        total_manual=total_manual, total_active=total_active,
        destinations=destinations, agoda_destinations=agoda_destinations,
        recent=recent, page=page, recent_total=recent_total,
        recent_pages=recent_pages, per_page=per_page, f_image=f_image,
        agoda_no_image=agoda_no_image)


@app.route('/admin/hotel-sync/save-credentials', methods=['POST'])
def admin_hotel_sync_save_credentials():
    """Save Agoda API credentials"""
    cid = request.form.get('cid', '').strip()
    api_key = request.form.get('api_key', '').strip()
    # Don't overwrite API key with the masked placeholder
    if api_key == 'configured':
        api_key = SiteSettings.get('agoda_api_key', '')
    SiteSettings.set_val('agoda_cid', cid, 'api')
    SiteSettings.set_val('agoda_api_key', api_key, 'api')
    SiteSettings.set_val('agoda_enabled', '1' if cid and api_key else '0', 'general')
    # Update AffiliateNetwork table too
    net = AffiliateNetwork.query.filter_by(slug='agoda').first()
    if net:
        net.api_key = api_key
        net.status = 'connected' if cid and api_key else 'disconnected'
    db.session.commit()
    # Reset singleton
    import agoda_integration
    agoda_integration._api_instance = None
    flash(f'Da luu Agoda credentials (CID: {cid})', 'success')
    return redirect(url_for('admin_hotel_sync'))


@app.route('/admin/hotel-sync/test', methods=['POST'])
def admin_hotel_sync_test():
    """Test Agoda API connection"""
    from agoda_integration import get_agoda_api
    api = get_agoda_api()
    if not api:
        return jsonify({'connected': False, 'error': 'Chua cau hinh API credentials'})
    result = api.test_connection()
    return jsonify(result)


@app.route('/admin/hotel-sync/search', methods=['POST'])
def admin_hotel_sync_search():
    """Search hotels from Agoda API for a destination"""
    from agoda_integration import get_agoda_api
    api = get_agoda_api()
    if not api:
        return jsonify({'error': 'Agoda API chua cau hinh', 'hotels': []})

    data = request.get_json() or {}
    destination = data.get('destination', '')
    checkin = data.get('checkin', '')
    checkout = data.get('checkout', '')

    if not destination:
        return jsonify({'error': 'Chua chon destination', 'hotels': []})

    hotels = api.search_city_hotels(destination, checkin or None, checkout or None)

    # Mark which ones are already in DB
    existing_ids = set()
    for h in Hotel.query.filter_by(source='agoda_api').all():
        if h.agoda_url:
            # Extract hotel ID from agoda_url
            for part in h.agoda_url.split('&'):
                if part.startswith('hid='):
                    existing_ids.add(part.split('=')[1])

    for h in hotels:
        h['already_imported'] = str(h.get('agoda_id', '')) in existing_ids

    return jsonify({'hotels': hotels, 'total': len(hotels)})


@app.route('/admin/hotel-sync/import', methods=['POST'])
def admin_hotel_sync_import():
    """Import selected hotels from Agoda search results into DB"""
    data = request.get_json() or {}
    hotels_data = data.get('hotels', [])

    if not hotels_data:
        return jsonify({'error': 'Khong co hotel nao de import', 'imported': 0})

    imported = 0
    for h in hotels_data:
        agoda_id = str(h.get('agoda_id', ''))
        name = h.get('name', '').strip()
        if not name:
            continue

        # Check duplicate by agoda_url containing the hotel ID
        existing = Hotel.query.filter(Hotel.agoda_url.contains(f'hid={agoda_id}')).first()
        if existing:
            continue

        amenities_raw = h.get('amenities', '')
        if isinstance(amenities_raw, list):
            amenities_raw = ', '.join(str(a) for a in amenities_raw)

        hotel = Hotel(
            name=name,
            slug=slugify(name)[:60],
            destination=h.get('destination', ''),
            destination_name=h.get('destination_name', ''),
            stars=int(h.get('stars', 0)) or 4,
            district=h.get('district', '') or h.get('address', ''),
            description=h.get('description', ''),
            amenities=str(amenities_raw)[:500],
            rating=float(h.get('rating', 0)) or 8.0,
            reviews_count=int(h.get('reviews_count', 0)),
            price_from=float(h.get('price_from', 0)),
            image_url=h.get('image_url', ''),
            agoda_url=h.get('agoda_url', ''),
            latitude=float(h.get('latitude', 0)),
            longitude=float(h.get('longitude', 0)),
            address=h.get('address', '') or h.get('district', ''),
            source='agoda_api',
            is_active=True,
            is_featured=False
        )
        db.session.add(hotel)
        imported += 1

    db.session.commit()
    return jsonify({'imported': imported, 'message': f'Da import {imported} khach san'})


@app.route('/admin/hotel-sync/fast', methods=['POST'])
def admin_hotel_sync_fast():
    """Fast Sync: search + auto-import for a single destination.

    Called repeatedly from the frontend for each destination in the queue.
    Returns results for that one destination so the UI can show progress.
    """
    from agoda_integration import get_agoda_api, AGODA_CITY_NAMES
    api = get_agoda_api()
    if not api:
        return jsonify({'error': 'Agoda API chua cau hinh', 'imported': 0})

    data = request.get_json() or {}
    destination = data.get('destination', '')
    max_per_city = int(data.get('max_per_city', 30))

    if not destination:
        return jsonify({'error': 'Thieu destination', 'imported': 0})

    # Search hotels from Agoda
    hotels = api.search_city_hotels(destination)

    # Collect existing agoda hotel IDs from DB
    existing_ids = set()
    for h in Hotel.query.filter_by(source='agoda_api').all():
        if h.agoda_url:
            for part in h.agoda_url.split('&'):
                if part.startswith('hid='):
                    existing_ids.add(part.split('=')[1])

    imported = 0
    skipped_location = 0
    for h in hotels[:max_per_city]:
        agoda_id = str(h.get('agoda_id', ''))
        name = h.get('name', '').strip()
        if not name or agoda_id in existing_ids:
            continue
        # Double-check: skip non-Vietnam hotels that slipped through
        from agoda_integration import _is_vietnam_hotel
        if not _is_vietnam_hotel(name, h.get('address', ''), destination):
            skipped_location += 1
            continue

        amenities_raw = h.get('amenities', '')
        if isinstance(amenities_raw, list):
            amenities_raw = ', '.join(str(a) for a in amenities_raw)

        hotel = Hotel(
            name=name,
            slug=slugify(name)[:60],
            destination=h.get('destination', ''),
            destination_name=h.get('destination_name', '') or AGODA_CITY_NAMES.get(destination, destination),
            stars=int(h.get('stars', 0)) or 4,
            district=h.get('district', '') or h.get('address', ''),
            description=h.get('description', ''),
            amenities=str(amenities_raw)[:500],
            rating=float(h.get('rating', 0)) or 8.0,
            reviews_count=int(h.get('reviews_count', 0)),
            price_from=float(h.get('price_from', 0)),
            price_original=float(h.get('price_original', 0)),
            image_url=h.get('image_url', ''),
            agoda_url=h.get('agoda_url', ''),
            latitude=float(h.get('latitude', 0)),
            longitude=float(h.get('longitude', 0)),
            address=h.get('address', '') or h.get('district', ''),
            source='agoda_api',
            is_active=True,
            is_featured=False
        )
        db.session.add(hotel)
        existing_ids.add(agoda_id)
        imported += 1

    db.session.commit()
    return jsonify({
        'destination': destination,
        'found': len(hotels),
        'imported': imported,
        'skipped': len(hotels[:max_per_city]) - imported,
        'skipped_location': skipped_location
    })


@app.route('/admin/hotel-sync/fix-images', methods=['POST'])
def admin_hotel_fix_images():
    """Fix missing image_url for existing agoda_api hotels.

    Strategy 1: Re-search each destination via LT Search API → match by agoda_id
    Strategy 2: Content Feed API for batch image lookup
    Strategy 3: Generate SVG placeholder with hotel initial + destination gradient
    """
    from agoda_integration import get_agoda_api
    api = get_agoda_api()

    fixed_api = 0
    fixed_placeholder = 0
    cleared = 0

    # Clear broken image URLs (old pix6.agoda.net/{id}/0/{id}_1.jpg pattern)
    broken_pattern = 'agoda.net/hotelImages/'
    for h in Hotel.query.filter_by(source='agoda_api').all():
        agoda_id = None
        if h.agoda_url:
            for part in h.agoda_url.split('&'):
                if part.startswith('hid='):
                    agoda_id = part.split('=')[1]
                    break
        if h.image_url and broken_pattern in h.image_url and agoda_id and f'/{agoda_id}_' in h.image_url:
            h.image_url = ''
            cleared += 1

    # Collect hotels missing images, grouped by destination
    hotels_no_img = Hotel.query.filter(
        Hotel.source == 'agoda_api',
        db.or_(Hotel.image_url == '', Hotel.image_url == None)
    ).all()

    if not hotels_no_img and not cleared:
        return jsonify({'fixed': 0, 'fixed_api': 0, 'fixed_placeholder': 0, 'cleared': 0})

    by_dest = {}
    for h in hotels_no_img:
        by_dest.setdefault(h.destination, []).append(h)

    # Strategy 1: Re-search each destination via API and match by agoda_id
    if api:
        for dest, hotels in by_dest.items():
            try:
                results = api.search_city_hotels(dest)
                img_map = {}
                for r in results:
                    aid = str(r.get('agoda_id', ''))
                    img = r.get('image_url', '')
                    if aid and img:
                        img_map[aid] = img
                for h in hotels:
                    if h.image_url:
                        continue
                    agoda_id = None
                    if h.agoda_url:
                        for part in h.agoda_url.split('&'):
                            if part.startswith('hid='):
                                agoda_id = part.split('=')[1]
                                break
                    if agoda_id and agoda_id in img_map:
                        h.image_url = img_map[agoda_id]
                        fixed_api += 1
            except Exception:
                continue

    # Strategy 2: Generate SVG placeholder for remaining hotels without images
    still_missing = Hotel.query.filter(
        Hotel.source == 'agoda_api',
        db.or_(Hotel.image_url == '', Hotel.image_url == None)
    ).all()
    for h in still_missing:
        initial = h.name[0].upper() if h.name else 'H'
        hue = abs(hash(h.name)) % 360
        h.image_url = (
            f"data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='400' height='300'%3E"
            f"%3Cdefs%3E%3ClinearGradient id='g' x1='0%25' y1='0%25' x2='100%25' y2='100%25'%3E"
            f"%3Cstop offset='0%25' stop-color='hsl({hue},55%25,35%25)'/%3E"
            f"%3Cstop offset='100%25' stop-color='hsl({(hue+45)%360},45%25,22%25)'/%3E"
            f"%3C/linearGradient%3E%3C/defs%3E"
            f"%3Crect width='400' height='300' fill='url(%23g)'/%3E"
            f"%3Ctext x='200' y='135' text-anchor='middle' font-family='Arial,sans-serif' font-size='72' font-weight='700' fill='rgba(255,255,255,0.18)'%3E{initial}%3C/text%3E"
            f"%3Ctext x='200' y='185' text-anchor='middle' font-family='Arial,sans-serif' font-size='14' fill='rgba(255,255,255,0.5)'%3E"
            f"{'%E2%AD%90' * min(h.stars, 5)}"
            f"%3C/text%3E%3C/svg%3E"
        )
        fixed_placeholder += 1

    db.session.commit()
    return jsonify({
        'fixed': fixed_api + fixed_placeholder,
        'fixed_api': fixed_api,
        'fixed_placeholder': fixed_placeholder,
        'cleared': cleared
    })


@app.route('/admin/hotel-sync/fix-addresses', methods=['POST'])
def admin_hotel_fix_addresses():
    """Reverse geocode hotels using OpenStreetMap Nominatim to get street addresses."""
    import requests as req
    import time as _time

    headers = {'User-Agent': 'UniTravel/1.0 (hotel geocoding)'}
    hotels = Hotel.query.filter(Hotel.latitude != 0, Hotel.longitude != 0).all()
    updated = 0
    errors = 0

    for h in hotels:
        try:
            r = req.get('https://nominatim.openstreetmap.org/reverse',
                params={'lat': h.latitude, 'lon': h.longitude,
                        'format': 'json', 'addressdetails': 1,
                        'accept-language': 'vi', 'zoom': 18},
                headers=headers, timeout=10)
            data = r.json()
            addr = data.get('address', {})

            road = addr.get('road', '')
            house = addr.get('house_number', '')
            suburb = addr.get('suburb') or addr.get('city_district') or addr.get('quarter') or ''
            city = h.destination_name or ''

            parts = []
            if house and road:
                parts.append(f'{house} {road}')
            elif road:
                parts.append(road)
            if suburb:
                parts.append(suburb)
            if city and city not in (suburb, road):
                parts.append(city)

            new_addr = ', '.join(parts)
            if new_addr and len(new_addr) > len(h.address or ''):
                h.address = new_addr
                updated += 1
        except Exception:
            errors += 1

        _time.sleep(1.1)  # Nominatim rate limit: 1 req/sec

    db.session.commit()
    return jsonify({'updated': updated, 'total': len(hotels), 'errors': errors})


@app.route('/admin/api/agoda-search')
def admin_api_agoda_search_public():
    """Public API: search Agoda hotels for a destination (used by travel page)"""
    from agoda_integration import get_agoda_api
    api = get_agoda_api()
    if not api:
        return jsonify({'hotels': [], 'source': 'unavailable'})

    destination = request.args.get('destination', '')
    checkin = request.args.get('checkin', '')
    checkout = request.args.get('checkout', '')

    if not destination:
        return jsonify({'hotels': [], 'source': 'no_destination'})

    # Resolve province slug → city slug (e.g. lam-dong → da-lat)
    from agoda_integration import PROVINCE_TO_CITY_SLUG
    city_slug = PROVINCE_TO_CITY_SLUG.get(destination, destination)

    hotels = api.search_city_hotels(city_slug, checkin or None, checkout or None)
    return jsonify({'hotels': hotels, 'total': len(hotels), 'source': 'agoda_api'})


# =============================================
# ADMIN — ATTRACTIONS (Vé tham quan)
# =============================================
ATTRACTION_CATS = [
    ('zoo','🦁 Sở thú'),('aquarium','🐠 Thủy cung'),('cable_car','🚡 Cáp treo'),
    ('theme_park','🎢 Công viên'),('museum','🏛️ Bảo tàng'),('tour','🚌 Tour'),
    ('show','🎭 Show diễn'),('waterpark','🌊 Công viên nước')
]
ATTRACTION_CAT_MAP = dict(ATTRACTION_CATS)

@app.route('/admin/attractions')
def admin_attractions():
    f_dest = request.args.get('destination', '')
    f_cat = request.args.get('category', '')
    f_net = request.args.get('network', '')
    q = Attraction.query
    if f_dest:
        q = q.filter(Attraction.destination == f_dest)
    if f_cat:
        q = q.filter(Attraction.category == f_cat)
    if f_net:
        q = q.filter(Attraction.network == f_net)
    items = q.order_by(Attraction.is_featured.desc(), Attraction.clicks.desc()).all()
    destinations = db.session.query(Attraction.destination, Attraction.destination_name).distinct().all()
    total = Attraction.query.count()
    active = Attraction.query.filter_by(is_active=True).count()
    total_clicks = db.session.query(db.func.sum(Attraction.clicks)).scalar() or 0
    total_conv = db.session.query(db.func.sum(Attraction.conversions)).scalar() or 0
    return render_template('admin/attractions.html', items=items, destinations=destinations,
        cats=ATTRACTION_CATS, cat_map=ATTRACTION_CAT_MAP,
        f_dest=f_dest, f_cat=f_cat, f_net=f_net,
        total=total, active=active, total_clicks=total_clicks, total_conv=total_conv)

@app.route('/admin/attraction/new', methods=['GET','POST'])
def admin_attraction_new():
    if request.method == 'POST':
        a = Attraction(
            name=request.form['name'],
            slug=slugify(request.form['name'])[:60],
            destination=request.form['destination'],
            destination_name=request.form.get('destination_name',''),
            category=request.form.get('category','tour'),
            description=request.form.get('description',''),
            address=request.form.get('address',''),
            price_from=float(request.form.get('price_from',0)),
            price_original=float(request.form.get('price_original',0)),
            discount_pct=int(request.form.get('discount_pct',0)),
            image_url=request.form.get('image_url',''),
            network=request.form.get('network','klook'),
            affiliate_url=request.form.get('affiliate_url',''),
            is_active='is_active' in request.form,
            is_featured='is_featured' in request.form
        )
        db.session.add(a); db.session.commit()
        flash(f'Da them: {a.name}', 'success')
        return redirect(url_for('admin_attractions'))
    return render_template('admin/attraction_form.html', attraction=None, cats=ATTRACTION_CATS)

@app.route('/admin/attraction/<int:aid>/edit', methods=['GET','POST'])
def admin_attraction_edit(aid):
    a = Attraction.query.get_or_404(aid)
    if request.method == 'POST':
        a.name = request.form['name']
        a.destination = request.form['destination']
        a.destination_name = request.form.get('destination_name','')
        a.category = request.form.get('category','tour')
        a.description = request.form.get('description','')
        a.address = request.form.get('address','')
        a.price_from = float(request.form.get('price_from',0))
        a.price_original = float(request.form.get('price_original',0))
        a.discount_pct = int(request.form.get('discount_pct',0))
        a.image_url = request.form.get('image_url','')
        a.network = request.form.get('network','klook')
        a.affiliate_url = request.form.get('affiliate_url','')
        a.is_active = 'is_active' in request.form
        a.is_featured = 'is_featured' in request.form
        db.session.commit()
        flash(f'Da cap nhat: {a.name}', 'success')
        return redirect(url_for('admin_attractions'))
    return render_template('admin/attraction_form.html', attraction=a, cats=ATTRACTION_CATS)

@app.route('/admin/attraction/<int:aid>/toggle', methods=['POST'])
def admin_attraction_toggle(aid):
    a = Attraction.query.get_or_404(aid)
    a.is_active = not a.is_active; db.session.commit()
    return redirect(url_for('admin_attractions'))

@app.route('/admin/attraction/<int:aid>/delete', methods=['POST'])
def admin_attraction_delete(aid):
    a = Attraction.query.get_or_404(aid)
    db.session.delete(a); db.session.commit()
    flash('Da xoa ve tham quan', 'success')
    return redirect(url_for('admin_attractions'))

@app.route('/admin/attractions/bulk-delete', methods=['POST'])
def admin_attractions_bulk_delete():
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify(ok=False, error='No IDs'), 400
    deleted = Attraction.query.filter(Attraction.id.in_([int(i) for i in ids])).delete()
    db.session.commit()
    return jsonify(ok=True, deleted=deleted)

# =============================================
# ADMIN — VOUCHERS (Mã giảm giá)
# =============================================
VOUCHER_CATS = [
    ('food','🍔 Ăn uống'),('shopping','🛍️ Mua sắm'),('travel','✈️ Du lịch'),
    ('services','🔧 Dịch vụ'),('entertainment','🎬 Giải trí'),('tech','💻 Công nghệ'),
    ('health','💊 Sức khỏe'),('education','📚 Giáo dục')
]
VOUCHER_CAT_MAP = dict(VOUCHER_CATS)

VOUCHER_TYPES = [
    ('percentage','% Phần trăm'),('fixed_amount','₫ Số tiền'),('free_shipping','🚚 Freeship')
]
VOUCHER_TYPE_MAP = dict(VOUCHER_TYPES)

def _voucher_valid_filter(q):
    """Apply SQL-level is_valid() filter: active, within date range, usage not exceeded"""
    now = datetime.utcnow()
    q = q.filter(Voucher.is_active == True)
    q = q.filter(db.or_(Voucher.valid_from == None, Voucher.valid_from <= now))
    q = q.filter(db.or_(Voucher.valid_to == None, Voucher.valid_to >= now))
    q = q.filter(db.or_(
        Voucher.usage_limit == None,
        Voucher.usage_limit == 0,
        Voucher.usage_count < Voucher.usage_limit
    ))
    return q

@app.route('/admin/vouchers')
def admin_vouchers():
    f_cat = request.args.get('category', '')
    f_merchant = request.args.get('merchant', '')
    f_status = request.args.get('status', '')  # valid, expired, used_up
    q = Voucher.query
    if f_cat:
        q = q.filter_by(category=f_cat)
    if f_merchant:
        q = q.filter(Voucher.merchant.ilike(f'%{f_merchant}%'))

    # Filter by status in SQL instead of Python
    if f_status == 'valid':
        q = _voucher_valid_filter(q)
    elif f_status == 'expired':
        now = datetime.utcnow()
        q = q.filter(db.or_(
            Voucher.is_active == False,
            db.and_(Voucher.valid_to != None, Voucher.valid_to < now),
            db.and_(Voucher.usage_limit != None, Voucher.usage_limit > 0, Voucher.usage_count >= Voucher.usage_limit)
        ))

    items = q.order_by(Voucher.created_at.desc()).all()

    merchants = db.session.query(Voucher.merchant).distinct().all()
    merchants = sorted([m[0] for m in merchants])
    total = Voucher.query.count()
    active = _voucher_valid_filter(Voucher.query).count()
    total_clicks = db.session.query(db.func.sum(Voucher.clicks)).scalar() or 0
    total_conv = db.session.query(db.func.sum(Voucher.conversions)).scalar() or 0
    return render_template('admin/vouchers.html', items=items, merchants=merchants,
        cats=VOUCHER_CATS, cat_map=VOUCHER_CAT_MAP, types=VOUCHER_TYPES,
        f_cat=f_cat, f_merchant=f_merchant, f_status=f_status,
        total=total, active=active, total_clicks=total_clicks, total_conv=total_conv)

@app.route('/admin/voucher/new', methods=['GET','POST'])
def admin_voucher_new():
    if request.method == 'POST':
        from datetime import datetime
        v = Voucher(
            code=request.form['code'].upper().strip(),
            title=request.form['title'],
            description=request.form.get('description',''),
            merchant=request.form['merchant'],
            category=request.form.get('category','general'),
            discount_type=request.form.get('discount_type','percentage'),
            discount_value=float(request.form.get('discount_value',0)),
            min_order=float(request.form.get('min_order',0)),
            max_discount=float(request.form.get('max_discount',0)),
            valid_from=datetime.strptime(request.form.get('valid_from'), '%Y-%m-%d') if request.form.get('valid_from') else datetime.utcnow(),
            valid_to=datetime.strptime(request.form['valid_to'], '%Y-%m-%d'),
            usage_limit=int(request.form.get('usage_limit',0)),
            network=request.form.get('network','shopee'),
            affiliate_url=request.form.get('affiliate_url',''),
            terms=request.form.get('terms',''),
            icon=request.form.get('icon','🎫'),
            color=request.form.get('color','#e74c3c'),
            is_active=request.form.get('is_active') == 'on',
            is_featured=request.form.get('is_featured') == 'on',
            is_exclusive=request.form.get('is_exclusive') == 'on',
            sync_mode=request.form.get('sync_mode','manual')
        )
        db.session.add(v); db.session.commit()
        flash(f'Da them voucher: {v.code}', 'success')
        return redirect(url_for('admin_vouchers'))
    return render_template('admin/voucher_form.html', voucher=None, cats=VOUCHER_CATS, types=VOUCHER_TYPES)

@app.route('/admin/voucher/<int:vid>', methods=['GET','POST'])
def admin_voucher_edit(vid):
    v = Voucher.query.get_or_404(vid)
    if request.method == 'POST':
        from datetime import datetime
        v.code = request.form['code'].upper().strip()
        v.title = request.form['title']
        v.description = request.form.get('description','')
        v.merchant = request.form['merchant']
        v.category = request.form.get('category','general')
        v.discount_type = request.form.get('discount_type','percentage')
        v.discount_value = float(request.form.get('discount_value',0))
        v.min_order = float(request.form.get('min_order',0))
        v.max_discount = float(request.form.get('max_discount',0))
        v.valid_from = datetime.strptime(request.form.get('valid_from'), '%Y-%m-%d') if request.form.get('valid_from') else v.valid_from
        v.valid_to = datetime.strptime(request.form['valid_to'], '%Y-%m-%d')
        v.usage_limit = int(request.form.get('usage_limit',0))
        v.network = request.form.get('network','shopee')
        v.affiliate_url = request.form.get('affiliate_url','')
        v.terms = request.form.get('terms','')
        v.icon = request.form.get('icon','🎫')
        v.color = request.form.get('color','#e74c3c')
        v.is_active = request.form.get('is_active') == 'on'
        v.is_featured = request.form.get('is_featured') == 'on'
        v.is_exclusive = request.form.get('is_exclusive') == 'on'
        v.embed_code = request.form.get('embed_code', '')
        v.sync_mode = request.form.get('sync_mode', 'manual')
        v.updated_at = datetime.utcnow()
        db.session.commit()
        flash(f'Da cap nhat voucher: {v.code}', 'success')
        return redirect(url_for('admin_vouchers'))
    return render_template('admin/voucher_form.html', voucher=v, cats=VOUCHER_CATS, types=VOUCHER_TYPES)

@app.route('/admin/voucher/<int:vid>/delete', methods=['POST'])
def admin_voucher_delete(vid):
    v = Voucher.query.get_or_404(vid)
    db.session.delete(v); db.session.commit()
    flash('Da xoa voucher', 'success')
    return redirect(url_for('admin_vouchers'))

@app.route('/admin/vouchers/bulk-delete', methods=['POST'])
def admin_vouchers_bulk_delete():
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify(ok=False, error='No IDs'), 400
    deleted = Voucher.query.filter(Voucher.id.in_([int(i) for i in ids])).delete()
    db.session.commit()
    return jsonify(ok=True, deleted=deleted)

# =============================================
# ADMIN — VOUCHER WIDGETS (AccessTrade, etc.)
# =============================================
@app.route('/admin/voucher-widgets')
def admin_voucher_widgets():
    """Redirect standalone voucher widgets page to Vouchers Hub"""
    return redirect(url_for('admin_vouchers_hub', tab='widgets'))

@app.route('/admin/voucher-widget/new', methods=['GET','POST'])
def admin_voucher_widget_new():
    """Create new voucher widget"""
    if request.method == 'POST':
        w = VoucherWidget(
            name=request.form['name'],
            network=request.form.get('network', 'accesstrade'),
            embed_code=request.form['embed_code'],
            description=request.form.get('description', ''),
            placement=request.form.get('placement', 'voucher_page'),
            position=int(request.form.get('position', 1)),
            is_active=request.form.get('is_active') == 'on'
        )
        db.session.add(w)
        db.session.commit()
        flash(f'Đã thêm widget: {w.name}', 'success')
        return redirect(url_for('admin_vouchers_hub', tab='widgets'))
    return render_template('admin/voucher_widget_form.html', widget=None)

@app.route('/admin/voucher-widget/<int:wid>', methods=['GET','POST'])
def admin_voucher_widget_edit(wid):
    """Edit voucher widget"""
    w = VoucherWidget.query.get_or_404(wid)
    if request.method == 'POST':
        w.name = request.form['name']
        w.network = request.form.get('network', 'accesstrade')
        w.embed_code = request.form['embed_code']
        w.description = request.form.get('description', '')
        w.placement = request.form.get('placement', 'voucher_page')
        w.position = int(request.form.get('position', 1))
        w.is_active = request.form.get('is_active') == 'on'
        w.updated_at = datetime.utcnow()
        db.session.commit()
        flash(f'Đã cập nhật widget: {w.name}', 'success')
        return redirect(url_for('admin_vouchers_hub', tab='widgets'))
    return render_template('admin/voucher_widget_form.html', widget=w)

@app.route('/admin/voucher-widget/<int:wid>/toggle', methods=['POST'])
def admin_voucher_widget_toggle(wid):
    """Toggle widget active status"""
    w = VoucherWidget.query.get_or_404(wid)
    w.is_active = not w.is_active
    db.session.commit()
    flash(f'Widget {w.name}: {"Hiển thị" if w.is_active else "Ẩn"}', 'success')
    return redirect(url_for('admin_vouchers_hub', tab='widgets'))

@app.route('/admin/voucher-widget/<int:wid>/delete', methods=['POST'])
def admin_voucher_widget_delete(wid):
    """Delete voucher widget"""
    w = VoucherWidget.query.get_or_404(wid)
    name = w.name
    db.session.delete(w)
    db.session.commit()
    flash(f'Đã xóa widget: {name}', 'success')
    return redirect(url_for('admin_vouchers_hub', tab='widgets'))

@app.route('/admin/voucher-widgets/bulk-delete', methods=['POST'])
def admin_voucher_widgets_bulk_delete():
    """Bulk delete voucher widgets by IDs (JSON)"""
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify(ok=False, error='Khong co widget nao'), 400
    deleted = VoucherWidget.query.filter(VoucherWidget.id.in_([int(i) for i in ids])).delete(synchronize_session=False)
    db.session.commit()
    return jsonify(ok=True, deleted=deleted)

# =============================================
# ADMIN — HOT DEALS (Upload Excel)
# =============================================
@app.route('/admin/hotdeals')
def admin_hotdeals():
    """Redirect standalone hotdeals page to Vouchers Hub"""
    return redirect(url_for('admin_vouchers_hub', tab='hotdeals'))

@app.route('/admin/hotdeals/upload', methods=['GET', 'POST'])
def admin_hotdeals_upload():
    """Upload Hot_deal.xlsx and import data into HotDeal table"""
    if request.method == 'GET':
        return redirect(url_for('admin_vouchers_hub', tab='hotdeals'))
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('Vui long chon file Excel (.xlsx)', 'error')
        return redirect(url_for('admin_vouchers_hub', tab='hotdeals'))

    try:
        import openpyxl
    except ImportError:
        flash('Thieu thu vien openpyxl. Chay: pip install openpyxl', 'error')
        return redirect(url_for('admin_vouchers_hub', tab='hotdeals'))

    try:
        from io import BytesIO

        wb = openpyxl.load_workbook(BytesIO(file.read()))
        ws = wb.active

        # Read header row
        headers = [str(cell.value or '').strip() for cell in ws[1]]
        col_map = {}
        for i, h in enumerate(headers):
            hl = h.lower()
            if 'name' in hl: col_map['name'] = i
            elif 'campaign' in hl: col_map['campaign'] = i
            elif 'product' in hl and 'link' in hl: col_map['product_link'] = i
            elif 'start' in hl: col_map['start_date'] = i
            elif 'end' in hl: col_map['end_date'] = i
            elif hl == 'status': col_map['status'] = i
            elif 'hot' in hl and 'day' in hl: col_map['hot_day'] = i
            elif 'banner' in hl: col_map['banner'] = i
            elif 'detail' in hl: col_map['detail'] = i

        if 'name' not in col_map:
            flash('Khong tim thay cot "Name" trong file Excel', 'error')
            return redirect(url_for('admin_vouchers_hub', tab='hotdeals'))

        imported = 0
        skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=False):
            vals = [cell.value for cell in row]
            name = vals[col_map.get('name', 0)]
            if not name:
                continue

            # Parse dates
            start_date = vals[col_map.get('start_date', 3)] if 'start_date' in col_map else None
            end_date = vals[col_map.get('end_date', 4)] if 'end_date' in col_map else None

            if isinstance(start_date, str):
                try:
                    start_date = datetime.fromisoformat(start_date.replace('Z', '+00:00').replace('+00:00', ''))
                except:
                    start_date = datetime.utcnow()
            elif not isinstance(start_date, datetime):
                start_date = datetime.utcnow()

            if isinstance(end_date, str):
                try:
                    end_date = datetime.fromisoformat(end_date.replace('Z', '+00:00').replace('+00:00', ''))
                except:
                    end_date = datetime.utcnow() + timedelta(days=30)
            elif not isinstance(end_date, datetime):
                end_date = datetime.utcnow() + timedelta(days=30)

            campaign = str(vals[col_map.get('campaign', 1)] or '') if 'campaign' in col_map else ''
            product_link = str(vals[col_map.get('product_link', 2)] or '') if 'product_link' in col_map else ''
            status_val = str(vals[col_map.get('status', 5)] or '') if 'status' in col_map else ''
            hot_day = str(vals[col_map.get('hot_day', 6)] or '') if 'hot_day' in col_map else ''
            banner = str(vals[col_map.get('banner', 7)] or '') if 'banner' in col_map else ''
            detail = str(vals[col_map.get('detail', 8)] or '') if 'detail' in col_map else ''

            # Check for duplicate by name + campaign
            existing = HotDeal.query.filter_by(name=str(name), campaign=campaign).first()
            if existing:
                # Update existing
                existing.product_link = product_link
                existing.start_date = start_date
                existing.end_date = end_date
                existing.status = status_val
                existing.hot_day = hot_day
                existing.banner = banner
                existing.detail = detail
                existing.is_active = True
                existing.uploaded_at = datetime.utcnow()
                skipped += 1
            else:
                deal = HotDeal(
                    name=str(name),
                    campaign=campaign,
                    product_link=product_link,
                    start_date=start_date,
                    end_date=end_date,
                    status=status_val,
                    hot_day=hot_day,
                    banner=banner,
                    detail=detail,
                    is_active=True,
                )
                db.session.add(deal)
                imported += 1

        db.session.commit()
        flash(f'Import thanh cong: {imported} deal moi, {skipped} deal cap nhat', 'success')
    except Exception as e:
        flash(f'Loi import: {str(e)}', 'error')

    return redirect(url_for('admin_vouchers_hub', tab='hotdeals'))

@app.route('/admin/hotdeals/<int:deal_id>/toggle', methods=['POST'])
def admin_hotdeal_toggle(deal_id):
    """Toggle hotdeal active status"""
    deal = HotDeal.query.get_or_404(deal_id)
    deal.is_active = not deal.is_active
    db.session.commit()
    flash(f'{"Bat" if deal.is_active else "Tat"} deal: {deal.campaign}', 'success')
    return redirect(url_for('admin_vouchers_hub', tab='hotdeals'))

@app.route('/admin/hotdeals/<int:deal_id>/delete', methods=['POST'])
def admin_hotdeal_delete(deal_id):
    """Delete a hotdeal"""
    deal = HotDeal.query.get_or_404(deal_id)
    name = deal.campaign or deal.name[:30]
    db.session.delete(deal)
    db.session.commit()
    flash(f'Da xoa deal: {name}', 'success')
    return redirect(url_for('admin_vouchers_hub', tab='hotdeals'))

@app.route('/admin/hotdeals/delete-all', methods=['POST'])
def admin_hotdeals_delete_all():
    """Delete all hotdeals"""
    count = HotDeal.query.count()
    HotDeal.query.delete()
    db.session.commit()
    flash(f'Da xoa {count} deal', 'success')
    return redirect(url_for('admin_vouchers_hub', tab='hotdeals'))

@app.route('/admin/hotdeals/bulk-delete-selected', methods=['POST'])
def admin_hotdeals_bulk_delete_selected():
    """Delete selected hotdeals by IDs"""
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify(ok=False, error='No IDs'), 400
    deleted = HotDeal.query.filter(HotDeal.id.in_([int(i) for i in ids])).delete()
    db.session.commit()
    return jsonify(ok=True, deleted=deleted)

# =============================================
# ADMIN — ACCESSTRADE BANNERS (Auto-pull)
# =============================================

@app.route('/admin/at-banners')
def admin_at_banners():
    """Redirect standalone AT banners page to Vouchers Hub"""
    return redirect(url_for('admin_vouchers_hub', tab='banners'))

def _do_banner_sync():
    """Core banner sync logic — returns (imported, updated, total, error)"""
    from accesstrade_integration import get_accesstrade_api
    api = get_accesstrade_api()
    if not api:
        return 0, 0, 0, 'Chưa cấu hình AccessTrade API key.'

    offers = api.get_offers(limit=100, status=1)
    coupons = api.get_coupons_hot(limit=50)

    imported = 0
    updated = 0
    all_items = []

    for raw in offers:
        item = api._parse_offer(raw)
        item['_source'] = 'offer'
        all_items.append(item)
    for raw in coupons:
        item = api._parse_offer(raw)
        item['_source'] = 'coupon'
        all_items.append(item)

    for item in all_items:
        offer_id = item.get('offer_id', '')
        if not offer_id:
            continue

        discount_parts = []
        if item.get('discount_percentage'):
            discount_parts.append(str(item['discount_percentage']) + '%')
        if item.get('discount_value'):
            val = item['discount_value']
            if isinstance(val, (int, float)) and val >= 1000:
                discount_parts.append(str(int(val / 1000)) + 'K')
            elif val:
                discount_parts.append(str(val))
        discount_text = ' | '.join(discount_parts) if discount_parts else ''

        start_dt = None
        end_dt = None
        for date_str in [item.get('start_date', '')]:
            if date_str:
                try:
                    start_dt = datetime.fromisoformat(str(date_str).replace('Z', '+00:00').replace('+00:00', ''))
                except Exception:
                    pass
        for date_str in [item.get('end_date', '')]:
            if date_str:
                try:
                    end_dt = datetime.fromisoformat(str(date_str).replace('Z', '+00:00').replace('+00:00', ''))
                except Exception:
                    pass

        image_url = ''
        raw_img = item.get('merchant_logo', '')
        if raw_img and isinstance(raw_img, str) and raw_img.startswith('http'):
            image_url = raw_img
        for raw in offers + coupons:
            if str(raw.get('id', '')) == offer_id:
                direct_img = raw.get('image', '') or ''
                if direct_img and isinstance(direct_img, str) and direct_img.startswith('http'):
                    image_url = direct_img
                break

        existing = AccessTradeBanner.query.filter_by(offer_id=offer_id).first()
        if existing:
            existing.offer_name = item.get('offer_name', existing.offer_name)
            existing.description = item.get('description', '')[:500]
            existing.merchant = item.get('merchant', '')
            existing.merchant_logo = item.get('merchant_logo', '')
            existing.category = item.get('category', '')
            existing.aff_link = item.get('aff_link', '')
            existing.discount_text = discount_text
            if image_url:
                existing.image_url = image_url
            if start_dt:
                existing.start_date = start_dt
            if end_dt:
                existing.end_date = end_dt
            existing.synced_at = datetime.utcnow()
            updated += 1
        else:
            banner = AccessTradeBanner(
                offer_id=offer_id,
                offer_name=item.get('offer_name', 'Unknown'),
                description=item.get('description', '')[:500],
                merchant=item.get('merchant', ''),
                merchant_logo=item.get('merchant_logo', ''),
                category=item.get('category', ''),
                image_url=image_url,
                aff_link=item.get('aff_link', ''),
                start_date=start_dt,
                end_date=end_dt,
                discount_text=discount_text,
                placement='both',
                is_active=True,
            )
            db.session.add(banner)
            imported += 1

    db.session.commit()
    return imported, updated, len(all_items), None


@app.route('/admin/at-banners/sync', methods=['POST'])
def admin_at_banners_sync():
    """Pull offers/coupons from AccessTrade API and save as banners"""
    try:
        imported, updated, total, err = _do_banner_sync()
        if err:
            flash(err, 'error')
        else:
            flash(f'Sync thành công: {imported} banner mới, {updated} cập nhật. Tổng {total} offers/coupons.', 'success')
    except Exception as e:
        flash(f'Lỗi sync: {str(e)}', 'error')

    return redirect(url_for('admin_vouchers_hub', tab='banners'))

@app.route('/admin/at-banners/<int:banner_id>/toggle', methods=['POST'])
def admin_at_banner_toggle(banner_id):
    """Toggle banner active status"""
    b = AccessTradeBanner.query.get_or_404(banner_id)
    b.is_active = not b.is_active
    db.session.commit()
    flash(f'{"Bat" if b.is_active else "Tat"} banner: {b.merchant}', 'success')
    return redirect(url_for('admin_vouchers_hub', tab='banners'))

@app.route('/admin/at-banners/<int:banner_id>/placement', methods=['POST'])
def admin_at_banner_placement(banner_id):
    """Change banner placement"""
    b = AccessTradeBanner.query.get_or_404(banner_id)
    new_placement = request.form.get('placement', 'hotdeal')
    if new_placement in ('hotdeal', 'sidebar', 'both'):
        b.placement = new_placement
        db.session.commit()
        flash(f'Da doi vi tri: {b.merchant} → {new_placement}', 'success')
    return redirect(url_for('admin_vouchers_hub', tab='banners'))

@app.route('/admin/at-banners/<int:banner_id>/delete', methods=['POST'])
def admin_at_banner_delete(banner_id):
    """Delete a banner"""
    b = AccessTradeBanner.query.get_or_404(banner_id)
    name = b.merchant or b.offer_name[:30]
    db.session.delete(b)
    db.session.commit()
    flash(f'Da xoa banner: {name}', 'success')
    return redirect(url_for('admin_vouchers_hub', tab='banners'))

@app.route('/admin/at-banners/delete-all', methods=['POST'])
def admin_at_banners_delete_all():
    """Delete all banners"""
    count = AccessTradeBanner.query.count()
    AccessTradeBanner.query.delete()
    db.session.commit()
    flash(f'Da xoa {count} banner', 'success')
    return redirect(url_for('admin_vouchers_hub', tab='banners'))

@app.route('/admin/at-banners/bulk-delete-selected', methods=['POST'])
def admin_at_banners_bulk_delete_selected():
    """Delete selected banners by IDs"""
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify(ok=False, error='No IDs'), 400
    deleted = AccessTradeBanner.query.filter(AccessTradeBanner.id.in_([int(i) for i in ids])).delete()
    db.session.commit()
    return jsonify(ok=True, deleted=deleted)

@app.route('/admin/at-banners/schedule', methods=['POST'])
def admin_at_banners_schedule():
    """Save banner auto-sync schedule settings"""
    data = request.get_json(silent=True) or {}
    enabled = bool(data.get('enabled', False))
    sync_time = data.get('sync_time', '03:00')  # HH:MM
    SiteSettings.set_val('banner_auto_sync', 'on' if enabled else 'off', 'banner')
    SiteSettings.set_val('banner_sync_time', sync_time, 'banner')
    db.session.commit()
    # Restart the scheduler thread with new settings
    _start_banner_scheduler()
    return jsonify(status='ok', enabled=enabled, sync_time=sync_time)

@app.route('/admin/at-banners/schedule/status')
def admin_at_banners_schedule_status():
    """Get current schedule status and last run info"""
    return jsonify(
        enabled=SiteSettings.get('banner_auto_sync', 'on') == 'on',
        sync_time=SiteSettings.get('banner_sync_time', '03:00'),
        last_run=SiteSettings.get('banner_last_auto_sync', ''),
        last_result=SiteSettings.get('banner_last_auto_sync_result', ''),
    )

@app.route('/admin/at-banners/click/<int:banner_id>')
def admin_at_banner_click(banner_id):
    """Track banner click and redirect to affiliate link"""
    b = AccessTradeBanner.query.get(banner_id)
    if b and b.aff_link:
        b.clicks = (b.clicks or 0) + 1
        db.session.commit()
        return redirect(b.aff_link)
    return redirect(url_for('index'))

# =============================================
# ADMIN — WARD/COMMUNE (Phường/Xã)
# =============================================

# Province coordinates for map centering
PROVINCE_COORDS = {
    '01': {'lat': 21.0285, 'lng': 105.8542, 'name': 'Hà Nội'},
    '04': {'lat': 22.6666, 'lng': 106.2640, 'name': 'Cao Bằng'},
    '08': {'lat': 21.8237, 'lng': 105.2140, 'name': 'Tuyên Quang'},
    '11': {'lat': 21.3860, 'lng': 103.0230, 'name': 'Điện Biên'},
    '12': {'lat': 22.3686, 'lng': 103.4700, 'name': 'Lai Châu'},
    '14': {'lat': 21.3270, 'lng': 103.9144, 'name': 'Sơn La'},
    '15': {'lat': 22.3380, 'lng': 104.1487, 'name': 'Lào Cai'},
    '19': {'lat': 21.5928, 'lng': 105.8442, 'name': 'Thái Nguyên'},
    '20': {'lat': 21.8460, 'lng': 106.7610, 'name': 'Lạng Sơn'},
    '22': {'lat': 20.9530, 'lng': 107.0750, 'name': 'Quảng Ninh'},
    '24': {'lat': 21.1861, 'lng': 106.0763, 'name': 'Bắc Ninh'},
    '25': {'lat': 21.4225, 'lng': 105.2290, 'name': 'Phú Thọ'},
    '31': {'lat': 20.8449, 'lng': 106.6881, 'name': 'Hải Phòng'},
    '33': {'lat': 20.6530, 'lng': 106.0510, 'name': 'Hưng Yên'},
    '37': {'lat': 20.2510, 'lng': 105.9750, 'name': 'Ninh Bình'},
    '38': {'lat': 19.8070, 'lng': 105.7760, 'name': 'Thanh Hóa'},
    '40': {'lat': 18.6790, 'lng': 105.6813, 'name': 'Nghệ An'},
    '42': {'lat': 18.3430, 'lng': 105.9058, 'name': 'Hà Tĩnh'},
    '44': {'lat': 16.7500, 'lng': 107.1860, 'name': 'Quảng Trị'},
    '46': {'lat': 16.4637, 'lng': 107.5909, 'name': 'Huế'},
    '48': {'lat': 16.0544, 'lng': 108.2022, 'name': 'Đà Nẵng'},
    '51': {'lat': 15.1214, 'lng': 108.8044, 'name': 'Quảng Ngãi'},
    '52': {'lat': 13.9833, 'lng': 108.0000, 'name': 'Gia Lai'},
    '56': {'lat': 12.2388, 'lng': 109.1967, 'name': 'Khánh Hòa'},
    '66': {'lat': 12.7100, 'lng': 108.2378, 'name': 'Đắk Lắk'},
    '68': {'lat': 11.9404, 'lng': 108.4583, 'name': 'Lâm Đồng'},
    '75': {'lat': 10.9453, 'lng': 106.8243, 'name': 'Đồng Nai'},
    '79': {'lat': 10.8231, 'lng': 106.6297, 'name': 'Hồ Chí Minh'},
    '80': {'lat': 11.3352, 'lng': 106.0980, 'name': 'Tây Ninh'},
    '82': {'lat': 10.4524, 'lng': 105.6322, 'name': 'Đồng Tháp'},
    '86': {'lat': 10.2530, 'lng': 105.9720, 'name': 'Vĩnh Long'},
    '91': {'lat': 10.3860, 'lng': 105.4350, 'name': 'An Giang'},
    '92': {'lat': 10.0452, 'lng': 105.7469, 'name': 'Cần Thơ'},
    '96': {'lat': 9.1770, 'lng': 105.1500, 'name': 'Cà Mau'},
}

# Map province_code → destination slug (used on hotel page)
PROVINCE_TO_SLUG = {
    '01': 'ha-noi', '04': 'cao-bang', '08': 'tuyen-quang', '11': 'dien-bien',
    '12': 'lai-chau', '14': 'son-la', '15': 'lao-cai', '19': 'thai-nguyen',
    '20': 'lang-son', '22': 'quang-ninh', '24': 'bac-ninh', '25': 'phu-tho',
    '31': 'hai-phong', '33': 'hung-yen', '37': 'ninh-binh', '38': 'thanh-hoa',
    '40': 'nghe-an', '42': 'ha-tinh', '44': 'quang-tri', '46': 'hue',
    '48': 'da-nang', '51': 'quang-ngai', '52': 'gia-lai', '56': 'khanh-hoa',
    '66': 'dak-lak', '68': 'lam-dong', '75': 'dong-nai', '79': 'ho-chi-minh',
    '80': 'tay-ninh', '82': 'dong-thap', '86': 'vinh-long', '91': 'an-giang',
    '92': 'can-tho', '96': 'ca-mau',
}

def _auto_seed_wards():
    """Seed WardCommune from wards_default.json if table is empty."""
    if WardCommune.query.count() > 0:
        return 0
    import json as _json
    _path = os.path.join(os.path.dirname(__file__), 'wards_default.json')
    if not os.path.exists(_path):
        return 0
    try:
        with open(_path, 'r', encoding='utf-8') as f:
            rows = _json.load(f)
        for code, name, level, resolution, province_code, province_name in rows:
            db.session.add(WardCommune(
                code=code, name=name, level=level,
                resolution=resolution,
                province_code=province_code,
                province_name=province_name
            ))
        db.session.commit()
        return len(rows)
    except Exception:
        db.session.rollback()
        return 0

@app.route('/admin/wards')
def admin_wards():
    """Manage ward/commune data"""
    seeded = _auto_seed_wards()
    if seeded:
        flash(f'Tu dong nap {seeded} phuong/xa mac dinh', 'success')

    wards = WardCommune.query.order_by(WardCommune.province_code, WardCommune.name).all()
    # Group by province
    provinces = {}
    for w in wards:
        if w.province_code not in provinces:
            provinces[w.province_code] = {
                'name': w.province_name,
                'code': w.province_code,
                'wards': [],
                'coords': PROVINCE_COORDS.get(w.province_code, {'lat': 16.0, 'lng': 108.0})
            }
        provinces[w.province_code]['wards'].append(w)

    f_province = request.args.get('province', '')
    f_level = request.args.get('level', '')
    f_search = request.args.get('q', '')

    filtered_wards = wards
    if f_province:
        filtered_wards = [w for w in filtered_wards if w.province_code == f_province]
    if f_level:
        filtered_wards = [w for w in filtered_wards if w.level == f_level]
    if f_search:
        q = f_search.lower()
        filtered_wards = [w for w in filtered_wards if q in w.name.lower() or q in w.code]

    return render_template('admin/wards.html',
        wards=filtered_wards,
        provinces=provinces,
        province_coords=PROVINCE_COORDS,
        total=len(wards),
        f_province=f_province,
        f_level=f_level,
        f_search=f_search
    )

@app.route('/admin/wards/add', methods=['POST'])
def admin_wards_add():
    """Add a new ward"""
    code = request.form.get('code', '').strip()
    name = request.form.get('name', '').strip()
    level = request.form.get('level', '').strip()
    province_code = request.form.get('province_code', '').strip()
    province_name = request.form.get('province_name', '').strip()
    if not code or not name:
        flash('Ma va ten phuong/xa la bat buoc', 'error')
        return redirect(url_for('admin_wards'))
    if WardCommune.query.filter_by(code=code).first():
        flash(f'Ma {code} da ton tai', 'error')
        return redirect(url_for('admin_wards'))
    db.session.add(WardCommune(code=code, name=name, level=level,
                                province_code=province_code, province_name=province_name))
    db.session.commit()
    flash(f'Da them: {name}', 'success')
    return redirect(url_for('admin_wards'))

@app.route('/admin/wards/edit/<int:ward_id>', methods=['POST'])
def admin_wards_edit(ward_id):
    """Edit a ward"""
    w = WardCommune.query.get_or_404(ward_id)
    w.code = request.form.get('code', w.code).strip()
    w.name = request.form.get('name', w.name).strip()
    w.level = request.form.get('level', w.level).strip()
    w.province_code = request.form.get('province_code', w.province_code).strip()
    w.province_name = request.form.get('province_name', w.province_name).strip()
    db.session.commit()
    flash(f'Da cap nhat: {w.name}', 'success')
    return redirect(url_for('admin_wards'))

@app.route('/admin/wards/delete/<int:ward_id>', methods=['POST'])
def admin_wards_delete(ward_id):
    """Delete a single ward"""
    w = WardCommune.query.get_or_404(ward_id)
    name = w.name
    db.session.delete(w)
    db.session.commit()
    flash(f'Da xoa: {name}', 'success')
    return redirect(url_for('admin_wards'))

@app.route('/admin/wards/bulk-delete', methods=['POST'])
def admin_wards_bulk_delete():
    """Bulk delete wards by IDs (JSON)"""
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify(ok=False, error='Khong co ward nao'), 400
    deleted = WardCommune.query.filter(WardCommune.id.in_([int(i) for i in ids])).delete(synchronize_session=False)
    db.session.commit()
    return jsonify(ok=True, deleted=deleted)

@app.route('/admin/wards/reset', methods=['POST'])
def admin_wards_reset():
    """Reset to default data from wards_default.json"""
    WardCommune.query.delete()
    db.session.commit()
    seeded = _auto_seed_wards()
    flash(f'Da reset ve mac dinh: {seeded} phuong/xa', 'success')
    return redirect(url_for('admin_wards'))

@app.route('/api/wards')
def api_wards():
    """API endpoint to get wards filtered by province"""
    province_code = request.args.get('province', '')
    level = request.args.get('level', '')
    q = request.args.get('q', '')

    query = WardCommune.query
    if province_code:
        query = query.filter_by(province_code=province_code)
    if level:
        query = query.filter_by(level=level)
    if q:
        query = query.filter(WardCommune.name.ilike(f'%{q}%'))

    wards = query.order_by(WardCommune.name).all()
    return jsonify({
        'wards': [{
            'code': w.code,
            'name': w.name,
            'level': w.level,
            'province_code': w.province_code,
            'province_name': w.province_name,
            'destination_slug': slugify(w.province_name)
        } for w in wards],
        'total': len(wards)
    })

@app.route('/api/wards/by-destination')
def api_wards_by_destination():
    """Get wards for an Agoda destination slug (for cascading dropdown)."""
    _auto_seed_wards()  # ensure data exists
    slug = request.args.get('slug', '').strip()
    if not slug:
        return jsonify({'wards': [], 'province': ''})

    from agoda_integration import AGODA_TO_PROVINCE_NAMES
    province_names = AGODA_TO_PROVINCE_NAMES.get(slug, [])
    if not province_names:
        return jsonify({'wards': [], 'province': ''})

    # Search WardCommune by province name variants
    wards = []
    matched_province = ''
    for pname in province_names:
        results = WardCommune.query.filter(
            WardCommune.province_name.ilike(f'%{pname}%')
        ).order_by(WardCommune.name).all()
        if results:
            wards = [{'code': w.code, 'name': w.name, 'level': w.level} for w in results]
            matched_province = pname
            break

    return jsonify({'wards': wards, 'province': matched_province, 'total': len(wards)})

@app.route('/api/wards/search')
def api_wards_search():
    """Search provinces + wards from WardCommune DB for hotel autocomplete.
    No hardcoded province lists — everything comes from the database."""
    q = request.args.get('q', '').strip()
    if not q or len(q) < 2:
        return jsonify({'results': []})

    results = []
    ql = q.lower()

    # 1. Search provinces dynamically from WardCommune table
    all_provinces = db.session.query(
        WardCommune.province_code, WardCommune.province_name
    ).distinct().order_by(WardCommune.province_name).all()

    for pcode, pname in all_provinces:
        if ql in pname.lower():
            results.append({
                'type': 'province',
                'name': pname,
                'destination_slug': slugify(pname),
                'province_code': pcode,
                'province_name': pname
            })

    # 2. Search wards/phuong by name (limit 30)
    wards = WardCommune.query.filter(
        WardCommune.name.ilike(f'%{q}%')
    ).order_by(WardCommune.name).limit(30).all()

    for w in wards:
        results.append({
            'type': 'ward',
            'name': w.name,
            'level': w.level,
            'destination_slug': slugify(w.province_name),
            'province_code': w.province_code,
            'province_name': w.province_name
        })

    return jsonify({'results': results})

@app.route('/api/province-coords')
def api_province_coords():
    """API endpoint to get province coordinates for map"""
    code = request.args.get('code', '')
    if code and code in PROVINCE_COORDS:
        return jsonify(PROVINCE_COORDS[code])
    return jsonify(PROVINCE_COORDS)

@app.route('/api/top-products')
def api_top_products():
    """API endpoint — top bestselling products from AccessTrade"""
    merchant = request.args.get('merchant', '')
    try:
        from accesstrade_integration import get_accesstrade_api
        api = get_accesstrade_api()
        if api:
            result = api.get_top_products(merchant=merchant or None)
            return jsonify({'ok': True, 'data': result.get('data', []), 'total': result.get('total', 0)})
        return jsonify({'ok': False, 'error': 'API not configured', 'data': []})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'data': []})

# =============================================
# ADMIN — VOUCHER SYNC (AccessTrade Auto-Import)
# =============================================

# Merchant → category mapping for auto-categorization
MERCHANT_CAT_MAP = {
    'shopee': 'shopping', 'lazada': 'shopping', 'tiki': 'shopping', 'sendo': 'shopping',
    'tiktok': 'shopping', 'yes24': 'shopping', 'jd': 'shopping',
    'grab': 'food', 'grabfood': 'food', 'shopeefood': 'food', 'baemin': 'food',
    'gojek': 'food', 'loship': 'food',
    'traveloka': 'travel', 'agoda': 'travel', 'klook': 'travel', 'booking': 'travel',
    'vietnam airlines': 'travel', 'vietjet': 'travel', 'trip.com': 'travel', 'vntrip': 'travel',
    'fpt shop': 'tech', 'cellphones': 'tech', 'dien may xanh': 'tech', 'phong vu': 'tech',
    'the gioi di dong': 'tech', 'gearvn': 'tech',
    'guardian': 'health', 'hasaki': 'health', 'pharmacity': 'health',
    'cgv': 'entertainment', 'galaxy': 'entertainment', 'lotte cinema': 'entertainment',
}

# Merchant → icon mapping
MERCHANT_ICON_MAP = {
    'shopee': '🟠', 'lazada': '🔵', 'tiki': '🔷', 'sendo': '🔴',
    'grab': '🟢', 'grabfood': '🟢', 'traveloka': '🔵', 'agoda': '🏨',
    'klook': '🟠', 'fpt shop': '💻', 'cellphones': '📱', 'dien may xanh': '💚',
    'tiktok': '🎵', 'cgv': '🎬', 'momo': '💜',
}

# Merchant → color mapping
MERCHANT_COLOR_MAP = {
    'shopee': '#ee4d2d', 'lazada': '#0f146d', 'tiki': '#1a94ff', 'sendo': '#ee2624',
    'grab': '#00b14f', 'traveloka': '#0064d2', 'agoda': '#5392f9', 'klook': '#ff5722',
    'fpt shop': '#d70018', 'cellphones': '#d70018', 'tiktok': '#000000',
}

def _guess_category(merchant_name):
    """Guess voucher category from merchant name"""
    name_lower = (merchant_name or '').lower()
    for key, cat in MERCHANT_CAT_MAP.items():
        if key in name_lower:
            return cat
    return 'shopping'

def _guess_icon(merchant_name):
    name_lower = (merchant_name or '').lower()
    for key, icon in MERCHANT_ICON_MAP.items():
        if key in name_lower:
            return icon
    return '🎫'

def _guess_color(merchant_name):
    name_lower = (merchant_name or '').lower()
    for key, color in MERCHANT_COLOR_MAP.items():
        if key in name_lower:
            return color
    return '#e74c3c'

def _parse_discount(text):
    """Try to parse discount type & value from offer text"""
    import re
    text = (text or '').lower()
    # Match "giảm 50%" or "50%" or "discount 30%"
    m = re.search(r'(\d+)\s*%', text)
    if m:
        return 'percentage', float(m.group(1))
    # Match "giảm 50k" or "giảm 50.000đ" or "50,000đ"
    m = re.search(r'(\d[\d.,]*)\s*(k|đ|d|vnd)', text)
    if m:
        val_str = m.group(1).replace('.', '').replace(',', '')
        val = float(val_str)
        if m.group(2) == 'k':
            val *= 1000
        return 'fixed_amount', val
    # Match "freeship" or "free ship"
    if 'freeship' in text or 'free ship' in text or 'miễn phí' in text:
        return 'free_shipping', 0
    return 'percentage', 0

def _parse_datetime(date_str):
    """Parse various date formats from AccessTrade"""
    if not date_str:
        return None
    from datetime import datetime as dt
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return dt.strptime(date_str.strip()[:19], fmt)
        except:
            continue
    return None

@app.route('/admin/voucher-sync')
def admin_voucher_sync():
    """Dashboard for AccessTrade voucher auto-sync"""
    from accesstrade_integration import get_accesstrade_api
    api = get_accesstrade_api()
    api_connected = api is not None

    # Stats
    total_synced = Voucher.query.filter_by(sync_mode='api').count()
    total_active = _voucher_valid_filter(Voucher.query.filter_by(sync_mode='api')).count()
    total_manual = Voucher.query.filter_by(sync_mode='manual').count()

    # Sync settings from SiteSettings
    auto_sync_enabled = SiteSettings.get('voucher_auto_sync', 'off') == 'on'
    sync_interval = SiteSettings.get('voucher_sync_interval', '6')
    last_sync = SiteSettings.get('voucher_last_sync', '')
    last_sync_count = SiteSettings.get('voucher_last_sync_count', '0')
    last_sync_error = SiteSettings.get('voucher_last_sync_error', '')

    # Recent API-synced vouchers
    recent = Voucher.query.filter_by(sync_mode='api').order_by(Voucher.created_at.desc()).limit(20).all()

    return render_template('admin/voucher_sync.html',
        api_connected=api_connected,
        total_synced=total_synced, total_active=total_active, total_manual=total_manual,
        auto_sync_enabled=auto_sync_enabled, sync_interval=sync_interval,
        last_sync=last_sync, last_sync_count=last_sync_count, last_sync_error=last_sync_error,
        recent=recent)

@app.route('/admin/voucher-sync/merchants')
def admin_voucher_sync_merchants():
    """Get merchant list from AccessTrade API for dropdown filter"""
    from accesstrade_integration import get_accesstrade_api
    api = get_accesstrade_api()
    if not api:
        return {'merchants': []}
    try:
        merchants = api.get_merchant_list()
        return {'merchants': merchants}
    except Exception:
        return {'merchants': []}

@app.route('/admin/voucher-sync/fetch', methods=['POST'])
def admin_voucher_sync_fetch():
    """Fetch offers from AccessTrade API (preview before import).

    Supports multiple sources via JSON body:
      source: 'all' (default) | 'hot' | 'expiring' | 'url'
      merchant: filter by merchant (for source='all')
      period: 1=weekly, 2=monthly (for source='hot')
      product_url: product link (for source='url')
    """
    from accesstrade_integration import get_accesstrade_api
    api = get_accesstrade_api()
    if not api:
        return {'error': 'AccessTrade API chưa cấu hình. Vào Settings → API để nhập key.'}, 400

    try:
        data = request.get_json(silent=True) or {}
        source = data.get('source', 'all')
        merchant = data.get('merchant', '')

        if source == 'hot':
            period = data.get('period', 1)
            raw = api.get_coupons_hot(limit=50, date=period)
            offers = [api._parse_offer(o) for o in raw]
        elif source == 'expiring':
            offers = api.get_offers_detailed(limit=50, scope='expiring')
        elif source == 'url':
            product_url = data.get('product_url', '').strip()
            if not product_url:
                return {'error': 'Vui lòng nhập link sản phẩm'}, 400
            raw = api.search_coupons_by_url(product_url)
            offers = [api._parse_offer(o) for o in raw]
        else:
            offers = api.get_offers_detailed(limit=100, merchant=merchant or None)

        # Count total coupons
        total_coupons = sum(len(o.get('coupons', [])) for o in offers)
        # Mark which ones already exist
        existing_ids = set()
        for v in Voucher.query.filter(Voucher.accesstrade_offer_id != '', Voucher.accesstrade_offer_id != None).all():
            existing_ids.add(v.accesstrade_offer_id)
        existing_codes = set(v.code.upper() for v in Voucher.query.all())

        for o in offers:
            oid = o['offer_id']
            o['already_imported'] = oid in existing_ids
            for c in o.get('coupons', []):
                c['already_imported'] = (c.get('code', '') or '').upper() in existing_codes

        return {
            'offers': offers,
            'total_offers': len(offers),
            'total_coupons': total_coupons,
            'existing_count': len(existing_ids)
        }
    except Exception as e:
        return {'error': f'Lỗi khi lấy dữ liệu: {str(e)}'}, 500

@app.route('/admin/voucher-sync/import', methods=['POST'])
def admin_voucher_sync_import():
    """Import selected offers as vouchers"""
    from accesstrade_integration import get_accesstrade_api
    data = request.get_json() or {}
    selected = data.get('selected', [])  # list of {offer_id, code, merchant, ...}

    if not selected:
        return {'error': 'Chưa chọn voucher nào để import'}, 400

    api = get_accesstrade_api()
    if not api:
        return {'error': 'AccessTrade API chưa cấu hình'}, 400

    imported = 0
    skipped = 0
    errors = []

    for item in selected:
        try:
            code = (item.get('code') or '').upper().strip()
            if not code:
                code = f"AT-{item.get('offer_id', 'X')}-{imported+1}"

            # Check duplicate
            existing = Voucher.query.filter(
                db.or_(Voucher.code == code,
                       db.and_(Voucher.accesstrade_offer_id == item.get('offer_id', ''),
                               Voucher.accesstrade_offer_id != ''))
            ).first()

            if existing:
                skipped += 1
                continue

            merchant = item.get('merchant', 'N/A')
            title = item.get('title') or item.get('offer_name') or f'Ưu đãi {merchant}'
            desc = item.get('description', '')

            # Use API discount fields if available, fallback to parsing title
            api_disc_val = item.get('discount_value')
            api_disc_pct = item.get('discount_percentage')
            api_min_spend = item.get('min_spend')
            api_max_value = item.get('max_value')

            if api_disc_pct and float(api_disc_pct) > 0:
                d_type = 'percentage'
                d_val = float(api_disc_pct)
            elif api_disc_val and float(api_disc_val) > 0:
                d_type = 'fixed_amount'
                d_val = float(api_disc_val)
            else:
                d_type, d_val = _parse_discount(title + ' ' + desc)

            # Parse dates
            valid_from = _parse_datetime(item.get('start_date')) or datetime.utcnow()
            valid_to = _parse_datetime(item.get('end_date'))
            if not valid_to:
                valid_to = datetime.utcnow() + timedelta(days=30)

            v = Voucher(
                code=code,
                title=title[:300],
                description=desc[:2000] if desc else '',
                merchant=merchant,
                category=_guess_category(merchant),
                discount_type=d_type,
                discount_value=d_val,
                min_order=float(api_min_spend) if api_min_spend else 0,
                max_discount=float(api_max_value) if api_max_value else 0,
                valid_from=valid_from,
                valid_to=valid_to,
                network='accesstrade',
                affiliate_url=item.get('aff_link', ''),
                icon=_guess_icon(merchant),
                color=_guess_color(merchant),
                is_active=True,
                sync_mode='api',
                accesstrade_offer_id=item.get('offer_id', ''),
            )
            db.session.add(v)
            imported += 1
        except Exception as e:
            errors.append(f"{item.get('code','?')}: {str(e)}")

    db.session.commit()

    # Update sync stats
    SiteSettings.set_val('voucher_last_sync', datetime.utcnow().strftime('%Y-%m-%d %H:%M'))
    SiteSettings.set_val('voucher_last_sync_count', str(imported))

    return {
        'imported': imported,
        'skipped': skipped,
        'errors': errors,
        'message': f'Đã import {imported} voucher, bỏ qua {skipped} trùng lặp'
    }

@app.route('/admin/voucher-sync/auto', methods=['POST'])
def admin_voucher_sync_auto():
    """Toggle auto-sync and save settings"""
    data = request.get_json() or {}
    enabled = data.get('enabled', False)
    interval = data.get('interval', '6')

    SiteSettings.set_val('voucher_auto_sync', 'on' if enabled else 'off')
    SiteSettings.set_val('voucher_sync_interval', str(interval))
    db.session.commit()

    return {'status': 'ok', 'auto_sync': enabled, 'interval': interval}

@app.route('/admin/voucher-sync/run', methods=['POST'])
def admin_voucher_sync_run():
    """Run sync now — fetch all offers and auto-import new ones"""
    from accesstrade_integration import get_accesstrade_api
    api = get_accesstrade_api()
    if not api:
        return {'error': 'AccessTrade API chưa cấu hình'}, 400

    try:
        offers = api.get_offers_detailed(limit=100)
        existing_codes = set(v.code for v in Voucher.query.all())
        existing_at_ids = set()
        for v in Voucher.query.filter(Voucher.accesstrade_offer_id != '', Voucher.accesstrade_offer_id != None).all():
            existing_at_ids.add(v.accesstrade_offer_id)

        imported = 0
        for offer in offers:
            oid = offer['offer_id']
            if oid in existing_at_ids:
                continue
            merchant = offer.get('merchant', 'N/A')
            # Import offer-level voucher
            coupons = offer.get('coupons', [])
            # Extract API discount fields
            api_disc_pct = offer.get('discount_percentage')
            api_disc_val = offer.get('discount_value')
            api_min_spend = offer.get('min_spend')
            api_max_value = offer.get('max_value')

            if coupons:
                for c in coupons:
                    code = (c.get('code', '') or '').upper().strip()
                    if not code or code in existing_codes:
                        continue
                    title = c.get('description') or offer.get('offer_name') or f'Ưu đãi {merchant}'

                    if api_disc_pct and float(api_disc_pct) > 0:
                        d_type, d_val = 'percentage', float(api_disc_pct)
                    elif api_disc_val and float(api_disc_val) > 0:
                        d_type, d_val = 'fixed_amount', float(api_disc_val)
                    else:
                        d_type, d_val = _parse_discount(title)

                    valid_from = _parse_datetime(c.get('start_date') or offer.get('start_date')) or datetime.utcnow()
                    valid_to = _parse_datetime(c.get('end_date') or offer.get('end_date'))
                    if not valid_to:
                        valid_to = datetime.utcnow() + timedelta(days=30)

                    v = Voucher(
                        code=code, title=title[:300],
                        description=offer.get('description', '')[:2000],
                        merchant=merchant, category=_guess_category(merchant),
                        discount_type=d_type, discount_value=d_val,
                        min_order=float(api_min_spend) if api_min_spend else 0,
                        max_discount=float(api_max_value) if api_max_value else 0,
                        valid_from=valid_from, valid_to=valid_to,
                        network='accesstrade', affiliate_url=offer.get('aff_link', ''),
                        icon=_guess_icon(merchant), color=_guess_color(merchant),
                        is_active=True, sync_mode='api',
                        accesstrade_offer_id=oid,
                    )
                    db.session.add(v)
                    existing_codes.add(code)
                    imported += 1
            else:
                # Offer without coupon codes — create as deal
                code = f"AT-{oid}"
                if code in existing_codes:
                    continue
                title = offer.get('offer_name') or f'Ưu đãi {merchant}'

                if api_disc_pct and float(api_disc_pct) > 0:
                    d_type, d_val = 'percentage', float(api_disc_pct)
                elif api_disc_val and float(api_disc_val) > 0:
                    d_type, d_val = 'fixed_amount', float(api_disc_val)
                else:
                    d_type, d_val = _parse_discount(title + ' ' + offer.get('description', ''))

                valid_from = _parse_datetime(offer.get('start_date')) or datetime.utcnow()
                valid_to = _parse_datetime(offer.get('end_date'))
                if not valid_to:
                    valid_to = datetime.utcnow() + timedelta(days=30)
                v = Voucher(
                    code=code, title=title[:300],
                    description=offer.get('description', '')[:2000],
                    merchant=merchant, category=_guess_category(merchant),
                    discount_type=d_type, discount_value=d_val,
                    min_order=float(api_min_spend) if api_min_spend else 0,
                    max_discount=float(api_max_value) if api_max_value else 0,
                    valid_from=valid_from, valid_to=valid_to,
                    network='accesstrade', affiliate_url=offer.get('aff_link', ''),
                    icon=_guess_icon(merchant), color=_guess_color(merchant),
                    is_active=True, sync_mode='api',
                    accesstrade_offer_id=oid,
                )
                db.session.add(v)
                existing_codes.add(code)
                imported += 1

        # Auto-deactivate expired API vouchers
        expired_count = 0
        for v in Voucher.query.filter_by(sync_mode='api', is_active=True).all():
            if v.valid_to and v.valid_to < datetime.utcnow():
                v.is_active = False
                expired_count += 1

        db.session.commit()

        SiteSettings.set_val('voucher_last_sync', datetime.utcnow().strftime('%Y-%m-%d %H:%M'))
        SiteSettings.set_val('voucher_last_sync_count', str(imported))
        SiteSettings.set_val('voucher_last_sync_error', '')

        return {
            'imported': imported, 'expired': expired_count,
            'message': f'Đã import {imported} voucher mới, vô hiệu {expired_count} voucher hết hạn'
        }
    except Exception as e:
        SiteSettings.set_val('voucher_last_sync_error', str(e))
        db.session.commit()
        return {'error': str(e)}, 500

@app.route('/admin/voucher-sync/cleanup', methods=['POST'])
def admin_voucher_sync_cleanup():
    """Remove all expired API-synced vouchers"""
    expired = Voucher.query.filter(
        Voucher.sync_mode == 'api',
        Voucher.valid_to < datetime.utcnow()
    ).all()
    count = len(expired)
    for v in expired:
        db.session.delete(v)
    db.session.commit()
    return {'deleted': count, 'message': f'Đã xóa {count} voucher hết hạn'}

# =============================================
# AI AUTO-CONTENT ENGINE
# =============================================

@app.route('/admin/content-calendar')
def admin_content_calendar():
    """Content Calendar - monthly overview with events and scheduled content"""
    import calendar as cal_mod
    year = request.args.get('year', date.today().year, type=int)
    month = request.args.get('month', date.today().month, type=int)

    # Get calendar data
    first_day = date(year, month, 1)
    days_in_month = cal_mod.monthrange(year, month)[1]
    last_day = date(year, month, days_in_month)

    # Get events for this month
    events = ContentEvent.query.filter(
        ContentEvent.start_date <= last_day,
        db.or_(ContentEvent.end_date >= first_day, ContentEvent.end_date.is_(None)),
        ContentEvent.is_active == True
    ).all()

    # Get scheduled queue items
    queue_items = ContentQueue.query.filter(
        ContentQueue.scheduled_at >= datetime(year, month, 1),
        ContentQueue.scheduled_at <= datetime(year, month, days_in_month, 23, 59, 59)
    ).all()

    # Build calendar grid
    cal = cal_mod.Calendar(firstweekday=0)
    weeks = cal.monthdatescalendar(year, month)

    # Map events and queue items to dates
    date_events = {}
    for e in events:
        d = e.start_date
        end = e.end_date or e.start_date
        while d <= end and d <= last_day:
            if d >= first_day:
                date_events.setdefault(d, []).append({'type': 'event', 'obj': e})
            d += timedelta(days=1)

    for q in queue_items:
        if q.scheduled_at:
            qd = q.scheduled_at.date()
            date_events.setdefault(qd, []).append({'type': 'queue', 'obj': q})

    # Navigation
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1

    verticals = Vertical.query.filter_by(status='published').all()

    return render_template('admin/content_calendar.html',
        year=year, month=month, weeks=weeks, date_events=date_events,
        events=events, queue_items=queue_items,
        prev_year=prev_year, prev_month=prev_month,
        next_year=next_year, next_month=next_month,
        month_name=cal_mod.month_name[month],
        today=date.today(), verticals=verticals)


@app.route('/admin/content-calendar/event', methods=['POST'])
def admin_content_event_save():
    """Save a content event"""
    event_id = request.form.get('event_id', type=int)
    if event_id:
        event = ContentEvent.query.get_or_404(event_id)
    else:
        event = ContentEvent(name='', start_date=date.today())
        db.session.add(event)

    event.name = request.form.get('name', '')
    event.event_type = request.form.get('event_type', 'holiday')
    event.start_date = datetime.strptime(request.form.get('start_date', ''), '%Y-%m-%d').date()
    end = request.form.get('end_date', '')
    event.end_date = datetime.strptime(end, '%Y-%m-%d').date() if end else None
    event.keywords = request.form.get('keywords', '')
    event.verticals = request.form.get('verticals', 'all')
    event.icon = request.form.get('icon', '')
    event.recurrence = request.form.get('recurrence', 'yearly')
    event.auto_generate = 'auto_generate' in request.form
    event.is_active = True
    db.session.commit()
    flash(f'Event saved: {event.name}', 'success')
    return redirect(url_for('admin_ai_engine', tab='calendar'))


@app.route('/admin/content-calendar/event/<int:eid>/delete', methods=['POST'])
def admin_content_event_delete(eid):
    """Delete a content event"""
    e = ContentEvent.query.get_or_404(eid)
    db.session.delete(e)
    db.session.commit()
    flash('Event deleted', 'success')
    return redirect(url_for('admin_ai_engine', tab='calendar'))


@app.route('/admin/auto-rules')
def admin_auto_rules():
    """Auto content rules - configure frequency/tone/layer per vertical"""
    verticals = Vertical.query.filter_by(status='published').order_by(Vertical.name).all()
    rules = AutoContentRule.query.all()
    rules_map = {r.vertical_id: r for r in rules}
    return render_template('admin/auto_rules.html', verticals=verticals, rules_map=rules_map)


@app.route('/admin/auto-rules/save', methods=['POST'])
def admin_auto_rules_save():
    """Save auto content rules for a vertical"""
    vid = request.form.get('vertical_id', type=int)
    rule = AutoContentRule.query.filter_by(vertical_id=vid).first()
    if not rule:
        rule = AutoContentRule(vertical_id=vid)
        db.session.add(rule)

    rule.frequency = request.form.get('frequency', 'daily')
    rule.articles_per_day = request.form.get('articles_per_day', 1, type=int)
    rule.knowledge_layer = request.form.get('knowledge_layer', 'auto')
    rule.tone = request.form.get('tone', 'seo')
    rule.auto_publish = 'auto_publish' in request.form
    rule.is_active = 'is_active' in request.form
    rule.topics_to_avoid = request.form.get('topics_to_avoid', '')
    rule.focus_keywords = request.form.get('focus_keywords', '')
    db.session.commit()
    flash(f'Rules saved for vertical', 'success')
    return redirect(url_for('admin_ai_engine', tab='rules'))


@app.route('/admin/content-queue')
def admin_content_queue():
    """Content Queue - review AI-suggested topics"""
    status_filter = request.args.get('status', 'all')
    vertical_filter = request.args.get('vertical', 'all')

    query = ContentQueue.query
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    if vertical_filter != 'all':
        v = Vertical.query.filter_by(slug=vertical_filter).first()
        if v:
            query = query.filter_by(vertical_id=v.id)

    queue = query.order_by(ContentQueue.created_at.desc()).all()
    verticals = Vertical.query.filter_by(status='published').all()

    stats = {
        'pending': ContentQueue.query.filter_by(status='pending').count(),
        'review': ContentQueue.query.filter_by(status='review').count(),
        'published': ContentQueue.query.filter_by(status='published').count(),
        'skipped': ContentQueue.query.filter_by(status='skipped').count(),
    }

    return render_template('admin/content_queue.html',
        queue=queue, verticals=verticals,
        status_filter=status_filter, vertical_filter=vertical_filter,
        stats=stats)


@app.route('/admin/content-queue/add', methods=['POST'])
def admin_content_queue_add():
    """Add item to content queue"""
    item = ContentQueue(
        vertical_id=request.form.get('vertical_id', type=int),
        topic=request.form.get('topic', ''),
        keywords=request.form.get('keywords', ''),
        knowledge_layer=request.form.get('knowledge_layer', 'L1'),
        source_type='manual',
        status='pending'
    )
    scheduled = request.form.get('scheduled_at', '')
    if scheduled:
        item.scheduled_at = datetime.strptime(scheduled, '%Y-%m-%dT%H:%M')
    db.session.add(item)
    db.session.commit()
    flash(f'Added to queue: {item.topic}', 'success')
    return redirect(url_for('admin_ai_engine', tab='pipeline', step='queue'))


@app.route('/admin/content-queue/<int:qid>/action', methods=['POST'])
def admin_content_queue_action(qid):
    """Approve/Skip/Edit/Generate/Publish queue item"""
    item = ContentQueue.query.get_or_404(qid)
    action = request.form.get('action', '')
    if action == 'approve':
        item.status = 'review'
    elif action == 'skip':
        item.status = 'skipped'
    elif action == 'generate':
        # Generate AI content for this item
        from ai_service import generate_from_queue_item
        try:
            result = generate_from_queue_item(item)
            flash(f'Generated: {result["title"]} ({result["word_count"]} words)', 'success')
        except Exception as e:
            flash(f'Generation failed: {str(e)}', 'danger')
        return redirect(url_for('admin_ai_engine', tab='pipeline', step='queue'))
    elif action == 'publish':
        # Publish as Article
        if item.generated_content:
            from ai_service import publish_queue_item
            try:
                article = publish_queue_item(item)
                flash(f'Published: {article.title}', 'success')
            except Exception as e:
                flash(f'Publish failed: {str(e)}', 'danger')
            return redirect(url_for('admin_ai_engine', tab='pipeline', step='queue'))
        else:
            item.status = 'published'
            item.published_at = datetime.utcnow()
    elif action == 'edit':
        item.topic = request.form.get('topic', item.topic)
        item.keywords = request.form.get('keywords', item.keywords)
        item.knowledge_layer = request.form.get('knowledge_layer', item.knowledge_layer)
    db.session.commit()
    return redirect(url_for('admin_ai_engine', tab='pipeline', step='queue'))


@app.route('/admin/content-queue/<int:qid>/delete', methods=['POST'])
def admin_content_queue_delete(qid):
    """Delete queue item"""
    item = ContentQueue.query.get_or_404(qid)
    db.session.delete(item)
    db.session.commit()
    flash('Queue item deleted', 'success')
    return redirect(url_for('admin_ai_engine', tab='pipeline', step='queue'))


@app.route('/admin/gap-analysis')
def admin_gap_analysis():
    """Unified AI Pipeline — Scan gaps, view queue, generate, publish — all in one page"""
    verticals = Vertical.query.filter_by(status='published').order_by(Vertical.name).all()

    # ── Gap analysis per vertical ──
    analysis = []
    for v in verticals:
        segments = Segment.query.filter_by(vertical_id=v.id).all()
        total_zones = 0
        zones_with_content = 0
        empty_zones = []

        for seg in segments:
            zones = Zone.query.filter_by(segment_id=seg.id).all()
            for z in zones:
                total_zones += 1
                parts_count = Part.query.filter_by(zone_id=z.id).count()
                has_seo = bool(z.seo_content and len(z.seo_content.strip()) > 50)
                if parts_count > 0 or has_seo:
                    zones_with_content += 1
                else:
                    empty_zones.append({'segment': seg.name, 'zone': z.name, 'zone_id': z.id})

        articles_count = Article.query.filter_by(vertical_slug=v.slug, status='published').count()
        queue_pending = ContentQueue.query.filter_by(vertical_id=v.id, status='pending').count()
        queue_review = ContentQueue.query.filter_by(vertical_id=v.id, status='review').count()
        rule = AutoContentRule.query.filter_by(vertical_id=v.id).first()

        coverage = round(zones_with_content / total_zones * 100) if total_zones > 0 else 0

        analysis.append({
            'vertical': v,
            'total_zones': total_zones,
            'zones_with_content': zones_with_content,
            'coverage': coverage,
            'empty_zones': empty_zones[:5],
            'articles_count': articles_count,
            'queue_pending': queue_pending,
            'queue_review': queue_review,
            'has_rule': rule is not None and rule.is_active if rule else False,
            'suggestions': _generate_gap_suggestions(v, coverage, empty_zones, articles_count)
        })

    # ── Queue items (for pipeline panel) ──
    queue_items = ContentQueue.query.order_by(ContentQueue.created_at.desc()).limit(50).all()
    queue_stats = {
        'pending': ContentQueue.query.filter_by(status='pending').count(),
        'review': ContentQueue.query.filter_by(status='review').count(),
        'published': ContentQueue.query.filter_by(status='published').count(),
        'total': ContentQueue.query.count(),
    }

    return render_template('admin/gap_analysis.html',
        analysis=analysis, verticals=verticals,
        queue_items=queue_items, queue_stats=queue_stats)


def _generate_gap_suggestions(vertical, coverage, empty_zones, articles_count):
    """Generate content gap suggestions based on analysis"""
    suggestions = []
    if coverage < 30:
        suggestions.append({'priority': 'high', 'text': f'Coverage only {coverage}% - need bulk content for {vertical.name}', 'action': 'auto_fill'})
    elif coverage < 70:
        suggestions.append({'priority': 'medium', 'text': f'{len(empty_zones)} zones without content in {vertical.name}', 'action': 'fill_zones'})

    if articles_count < 3:
        suggestions.append({'priority': 'high', 'text': f'Only {articles_count} articles - need knowledge base content', 'action': 'create_articles'})
    elif articles_count < 10:
        suggestions.append({'priority': 'medium', 'text': f'Add more L2/L3 depth articles for {vertical.name}', 'action': 'create_articles'})

    if not suggestions:
        suggestions.append({'priority': 'low', 'text': f'{vertical.name} looking good! Consider seasonal content.', 'action': 'seasonal'})

    return suggestions


# =============================================
# AI CONTENT PIPELINE — Scan, Generate, Publish
# =============================================

@app.route('/admin/ai/scan-gaps')
def admin_ai_scan_gaps():
    """Enhanced gap analysis with cross-vertical comparison and per-tier breakdown"""
    from ai_service import scan_all_gaps
    results = scan_all_gaps()

    # Overall stats
    total_possible = sum(r['total_possible'] for r in results)
    total_existing = sum(r['existing_articles'] for r in results)
    overall_coverage = round(total_existing / total_possible * 100, 1) if total_possible > 0 else 0

    # Per-tier totals across all verticals
    tier_totals = {'nganh': {'total': 0, 'existing': 0}, 'chung': {'total': 0, 'existing': 0}, 'chi-tiet': {'total': 0, 'existing': 0}}
    for r in results:
        for tier_name, tier_data in r['tier_breakdown'].items():
            tier_totals[tier_name]['total'] += tier_data['total']
            tier_totals[tier_name]['existing'] += tier_data['existing']

    return render_template('admin/ai_gap_scan.html',
        results=results,
        overall_coverage=overall_coverage,
        total_possible=total_possible,
        total_existing=total_existing,
        tier_totals=tier_totals)


@app.route('/admin/ai/auto-fill', methods=['POST'])
def admin_ai_auto_fill():
    """Auto-populate ContentQueue with gaps from scan"""
    from ai_service import auto_fill_queue
    vertical_id = request.form.get('vertical_id', type=int)
    max_items = request.form.get('max_items', 10, type=int)
    created = auto_fill_queue(vertical_id=vertical_id, max_items=max_items)
    flash(f'Added {len(created)} items to content queue', 'success')
    return redirect(url_for('admin_ai_engine', tab='pipeline', step='queue'))


@app.route('/admin/ai/generate/<int:qid>', methods=['POST'])
def admin_ai_generate(qid):
    """Generate article content for a specific queue item"""
    from ai_service import generate_from_queue_item
    item = ContentQueue.query.get_or_404(qid)
    try:
        result = generate_from_queue_item(item)
        flash(f'Generated: {result["title"]} ({result["word_count"]} words)', 'success')
    except Exception as e:
        flash(f'Generation failed: {str(e)}', 'danger')
    return redirect(url_for('admin_ai_engine', tab='pipeline', step='queue'))


@app.route('/admin/ai/publish/<int:qid>', methods=['POST'])
def admin_ai_publish(qid):
    """Publish a generated queue item as an Article"""
    from ai_service import publish_queue_item
    item = ContentQueue.query.get_or_404(qid)
    try:
        article = publish_queue_item(item)
        flash(f'Published article: {article.title}', 'success')
    except Exception as e:
        flash(f'Publish failed: {str(e)}', 'danger')
    return redirect(url_for('admin_ai_engine', tab='pipeline', step='queue'))


@app.route('/admin/ai/batch-generate', methods=['POST'])
def admin_ai_batch_generate():
    """Generate content for all pending/approved queue items"""
    from ai_service import generate_from_queue_item
    items = ContentQueue.query.filter(ContentQueue.status.in_(['pending', 'review'])).limit(5).all()
    success = 0
    errors = 0
    for item in items:
        if item.generated_content:
            continue  # Skip already generated
        try:
            generate_from_queue_item(item)
            success += 1
        except Exception as e:
            errors += 1
    flash(f'Batch generate: {success} success, {errors} errors', 'success' if errors == 0 else 'warning')
    return redirect(url_for('admin_ai_engine', tab='pipeline', step='queue'))


@app.route('/admin/ai/auto-run', methods=['POST'])
def admin_ai_auto_run():
    """Full pipeline: scan gaps → fill queue → generate → optionally publish"""
    from ai_service import auto_fill_queue, generate_from_queue_item, publish_queue_item
    vertical_id = request.form.get('vertical_id', type=int)
    auto_publish = request.form.get('auto_publish') == '1'
    max_items = request.form.get('max_items', 5, type=int)

    # Step 1: Fill queue
    created = auto_fill_queue(vertical_id=vertical_id, max_items=max_items)

    # Step 2: Generate content
    generated = 0
    published = 0
    errors = 0
    for item in created:
        try:
            generate_from_queue_item(item)
            generated += 1
            # Step 3: Auto-publish if enabled
            if auto_publish:
                rule = AutoContentRule.query.filter_by(vertical_id=item.vertical_id, is_active=True).first()
                if rule and rule.auto_publish:
                    publish_queue_item(item)
                    published += 1
        except Exception as e:
            errors += 1

    flash(f'Auto-run: {len(created)} queued, {generated} generated, {published} published, {errors} errors',
          'success' if errors == 0 else 'warning')
    return redirect(url_for('admin_ai_engine', tab='pipeline'))


# =============================================
# UNILAB HOMEPAGE — Article aggregation from all verticals
# =============================================
@app.route('/')
def unilab_home():
    """Homepage: aggregate articles from all verticals with WordPress Newspaper theme"""
    # Get all active verticals
    verticals = Vertical.query.filter_by(status='published').order_by(Vertical.name).all()

    # Featured article (most recent)
    featured = Article.query.filter_by(status='published').order_by(Article.created_at.desc()).first()

    # Recent articles from all verticals (exclude featured)
    recent_query = Article.query.filter_by(status='published')
    if featured:
        recent_query = recent_query.filter(Article.id != featured.id)
    recent_articles = recent_query.order_by(Article.created_at.desc()).limit(12).all()

    # Popular articles (by views)
    popular_articles = Article.query.filter_by(status='published').order_by(Article.views.desc()).limit(6).all()

    # Articles by vertical (for category sections)
    vertical_articles = {}
    for v in verticals:
        articles = Article.query.filter_by(vertical_slug=v.slug, status='published').order_by(Article.created_at.desc()).limit(4).all()
        if articles:
            vertical_articles[v.slug] = {'vertical': v, 'articles': articles}

    return render_template('unilab_home.html',
        featured=featured,
        recent_articles=recent_articles,
        popular_articles=popular_articles,
        verticals=verticals,
        vertical_articles=vertical_articles)

@app.route('/bai-viet/<slug>')
def unilab_article(slug):
    """Article detail page on homepage - read articles without leaving UniLab"""
    article = Article.query.filter_by(slug=slug, status='published').first_or_404()

    # Increment views
    article.views += 1
    db.session.commit()

    # Get vertical info
    vertical = Vertical.query.filter_by(slug=article.vertical_slug).first()

    # Get related articles (same vertical, exclude current)
    related_articles = Article.query.filter_by(
        vertical_slug=article.vertical_slug,
        status='published'
    ).filter(Article.id != article.id).order_by(Article.views.desc()).limit(6).all()

    # Popular articles from all verticals
    popular_articles = Article.query.filter_by(status='published')\
        .order_by(Article.views.desc()).limit(5).all()

    return render_template('unilab_article.html',
        article=article,
        vertical=vertical,
        related_articles=related_articles,
        popular_articles=popular_articles)

@app.route('/robots.txt')
def robots_txt():
    """Serve robots.txt to control search engine indexing"""
    from flask import send_from_directory
    return send_from_directory(app.root_path, 'robots.txt', mimetype='text/plain')

# =============================================
# =============================================
# OLD ROUTES REMOVED
# All car/pet/travel common routes now use shared routes:
# /<vertical_slug>, /<vertical_slug>/kien-thuc, etc.
# Only travel special routes kept below:
# =============================================

@app.route('/travel/khach-san')
def travel_hotels():
    v = Vertical.query.filter_by(slug='travel').first_or_404()
    destination = request.args.get('destination', '')
    stars = request.args.get('stars', '')
    price_min = request.args.get('price_min', '')
    price_max = request.args.get('price_max', '')

    agoda_enabled = SiteSettings.get('agoda_enabled', '0') == '1'
    api_status = 'configured' if agoda_enabled else 'local_db'

    # Resolve province slug (sidebar) → city slug (DB/Agoda)
    # e.g. 'lam-dong' → 'da-lat', 'khanh-hoa' → 'nha-trang'
    from agoda_integration import PROVINCE_TO_CITY_SLUG
    city_slug = PROVINCE_TO_CITY_SLUG.get(destination, destination)

    hotels = []
    agoda_hotels = []
    if destination:
        # Always get local DB hotels (try both province and city slug)
        q = Hotel.query.filter(
            Hotel.is_active == True,
            Hotel.destination.in_([destination, city_slug])
        )
        if stars:
            q = q.filter(Hotel.stars == int(stars))
        if price_min:
            q = q.filter(Hotel.price_from >= int(price_min))
        if price_max:
            q = q.filter(Hotel.price_from <= int(price_max))
        hotels = q.order_by(Hotel.is_featured.desc(), Hotel.rating.desc()).all()

    # Popular destinations from DB + Agoda city list
    dest_counts = db.session.query(Hotel.destination, Hotel.destination_name, db.func.count(Hotel.id)
        ).filter(Hotel.is_active == True).group_by(Hotel.destination, Hotel.destination_name).all()
    dest_icons = {'da-nang':'🏖️','phu-quoc':'🌴','nha-trang':'🌊','ha-noi':'🏯','ho-chi-minh':'🏙️','da-lat':'🌸','hoi-an':'🏮','sa-pa':'🏔️','vung-tau':'🏖️','quy-nhon':'🏖️','hue':'🏯','hai-phong':'⚓','can-tho':'🌾','ninh-binh':'🏯'}
    popular = [{'slug': d[0], 'name': d[1], 'icon': dest_icons.get(d[0],'📍'), 'count': d[2]} for d in dest_counts]

    # Add Agoda cities that aren't in DB yet
    if agoda_enabled:
        from agoda_integration import AGODA_CITY_IDS, AGODA_CITY_NAMES
        db_slugs = set(d[0] for d in dest_counts)
        for slug, city_id in AGODA_CITY_IDS.items():
            if slug not in db_slugs:
                popular.append({
                    'slug': slug,
                    'name': AGODA_CITY_NAMES.get(slug, slug),
                    'icon': dest_icons.get(slug, '📍'),
                    'count': 0
                })

    # Build hotel markers JSON for Leaflet map
    from agoda_integration import PROVINCE_CENTERS, _DEST_SLUG_TO_NAME
    destination_display = _DEST_SLUG_TO_NAME.get(destination, destination.replace('-', ' ').title()) if destination else ''
    hotel_markers = []
    marker_src = hotels
    if not destination:
        # No destination selected → show ALL active hotels on the Vietnam-wide map
        marker_src = Hotel.query.filter(Hotel.is_active == True, Hotel.latitude != 0, Hotel.longitude != 0).all()
    for h in marker_src:
        lat = h.latitude or 0
        lng = h.longitude or 0
        if lat and lng:
            hotel_markers.append({
                'name': h.name, 'lat': lat, 'lng': lng,
                'stars': h.stars, 'rating': h.rating,
                'price': h.price_from, 'image': h.image_url or '',
                'url': h.agoda_url or h.booking_url or '#',
                'address': h.address or h.district or '',
            })
    province_center = PROVINCE_CENTERS.get(destination, PROVINCE_CENTERS.get(city_slug, (16.0, 108.0)))

    return render_template('travel/hotels.html', vertical=v, hotels=hotels,
        destination=destination, destination_display=destination_display, city_slug=city_slug,
        stars=stars, price_min=price_min, price_max=price_max,
        api_status=api_status, agoda_enabled=agoda_enabled, popular=popular,
        hotel_markers=hotel_markers, province_center=province_center)


@app.route('/travel/ve-tham-quan')
def travel_attractions():
    v = Vertical.query.filter_by(slug='travel').first_or_404()
    f_dest = request.args.get('destination', '')
    f_cat = request.args.get('category', '')
    f_net = request.args.get('network', '')

    q = Attraction.query.filter_by(is_active=True)
    if f_dest:
        q = q.filter(Attraction.destination == f_dest)
    if f_cat:
        q = q.filter(Attraction.category == f_cat)
    if f_net:
        q = q.filter(Attraction.network == f_net)
    items = q.order_by(Attraction.is_featured.desc(), Attraction.rating.desc()).all()

    destinations = db.session.query(Attraction.destination, Attraction.destination_name
        ).filter(Attraction.is_active == True).distinct().all()
    cat_icons = {'zoo':'🦁','aquarium':'🐠','cable_car':'🚡','theme_park':'🎢','museum':'🏛️','tour':'🚌','show':'🎭','waterpark':'🌊'}
    cat_names = {'zoo':'Sở thú','aquarium':'Thủy cung','cable_car':'Cáp treo','theme_park':'Công viên','museum':'Bảo tàng','tour':'Tour','show':'Show diễn','waterpark':'Công viên nước'}
    cats_available = db.session.query(Attraction.category, db.func.count(Attraction.id)
        ).filter(Attraction.is_active == True).group_by(Attraction.category).all()

    return render_template('travel/attractions.html', vertical=v, items=items,
        destinations=destinations, cats_available=cats_available,
        cat_icons=cat_icons, cat_names=cat_names,
        f_dest=f_dest, f_cat=f_cat, f_net=f_net)


# SHARED VERTICAL ROUTES (Dynamic)
# =============================================
def get_template_path(vertical_slug, template_name):
    """Get the correct template path for a vertical.
    All verticals use shared templates. Extra nav (hotels, attractions)
    is handled via conditional block in shared/base.html."""
    return f'shared/{template_name}'

def get_vertical_config(vertical_slug):
    """Get vertical-specific customization config"""
    configs = {
        'car': {
            'hero_title': 'Kiến thức ô tô',
            'hero_subtitle': 'từ tổng thể đến từng bu-lông',
            'hero_desc': 'Tìm hiểu chi tiết về phụ tùng, hệ thống xe, bảo dưỡng và nâng cấp. Tra cứu mã OEM, so sánh giá, và mua đúng sản phẩm.',
            'hero_cta1': 'Khám phá phân khúc',
            'segments_label': 'phân khúc xe',
            'segments_heading': 'Phân khúc xe',
            'tier1_desc': 'Thị trường, xu hướng, OEM vs aftermarket, chi phí bảo dưỡng tổng quan.',
            'tier2_desc': 'Hệ thống treo, phanh, điện, động cơ — cấu tạo và nguyên lý hoạt động.',
            'tier3_desc': 'Bài viết sâu về từng phụ tùng: khi nào thay, chọn loại nào, mẹo tiết kiệm.',
            'cta_title': 'Không tìm thấy phụ tùng cần tra cứu?',
            'cta_desc': 'Hệ thống kiến thức ô tô đang được mở rộng mỗi ngày. Chọn phân khúc xe để xem chi tiết từng hệ thống.',
            'cta_button': 'Chọn phân khúc →',
            'products_title': 'Sản phẩm phụ tùng ô tô',
            'products_subtitle': 'Giá tốt nhất từ các sàn TMĐT — Shopee, Lazada, Tiki',
            'products_icon': '🛒',
            'parts_label': 'phụ tùng',
            'parts_heading': 'Phụ tùng',
        },
        'pet': {
            'hero_title': 'Kiến thức thú cưng',
            'hero_subtitle': 'chăm sóc đúng cách từ ngày đầu',
            'hero_desc': 'Tìm hiểu dinh dưỡng, sức khỏe, huấn luyện cho chó, mèo, cá cảnh. Review sản phẩm, so sánh giá.',
            'hero_cta1': 'Khám phá',
            'segments_label': 'danh mục',
            'segments_heading': 'Danh mục',
            'tier1_desc': 'Tổng quan thị trường, xu hướng, chi phí.',
            'tier2_desc': 'Hệ thống, quy trình, phương pháp.',
            'tier3_desc': 'Review sản phẩm, so sánh, hướng dẫn cụ thể.',
            'cta_title': 'Khám phá thêm kiến thức',
            'cta_desc': 'Hệ thống kiến thức đang được mở rộng mỗi ngày.',
            'cta_button': 'Khám phá ngay →',
            'products_title': 'Sản phẩm thú cưng',
            'products_subtitle': 'Review chi tiết, giá tốt từ các sàn TMĐT',
            'products_icon': '🛒',
            'parts_label': 'sản phẩm',
            'parts_heading': 'Sản phẩm',
        },
        'travel': {
            'hero_title': 'Khám phá du lịch',
            'hero_subtitle': 'trải nghiệm đáng nhớ từng điểm đến',
            'hero_desc': 'Khách sạn, tour, vé tham quan, hướng dẫn chi tiết. Giá tốt từ Agoda, Klook, Traveloka.',
            'hero_cta1': 'Khám phá điểm đến',
            'segments_label': 'điểm đến',
            'segments_heading': 'Điểm đến',
            'tier1_desc': 'Xu hướng du lịch, chi phí tổng quan.',
            'tier2_desc': 'Hướng dẫn, kinh nghiệm, tips du lịch.',
            'tier3_desc': 'Review khách sạn, tour, vé tham quan cụ thể.',
            'cta_title': 'Lên kế hoạch chuyến đi',
            'cta_desc': 'Khám phá các điểm đến và trải nghiệm tuyệt vời.',
            'cta_button': 'Xem điểm đến →',
            'products_title': 'Khách sạn & Vé tham quan',
            'products_subtitle': 'Giá tốt nhất từ Agoda, Klook, Traveloka',
            'products_icon': '🏨',
            'parts_label': 'trải nghiệm',
            'parts_heading': 'Trải nghiệm',
        },
        'bike': {
            'hero_title': 'Kiến thức xe đạp',
            'hero_subtitle': 'từ chọn xe đến nâng cấp chi tiết',
            'hero_desc': 'Road bike, MTB, touring — chọn xe phù hợp, bảo dưỡng đúng cách, nâng cấp hiệu quả. Tư vấn groupset, bánh xe, phụ kiện.',
            'hero_cta1': 'Khám phá loại xe',
            'segments_label': 'loại xe',
            'segments_heading': 'Loại xe đạp',
            'tier1_desc': 'Thị trường xe đạp, xu hướng, chi phí tổng quan, chọn xe phù hợp.',
            'tier2_desc': 'Hệ thống truyền động, phanh, bánh xe — cấu tạo và nguyên lý.',
            'tier3_desc': 'Review chi tiết từng bộ phận: groupset, vành, lốp, yên, tay lái.',
            'cta_title': 'Bắt đầu hành trình đạp xe?',
            'cta_desc': 'Khám phá kiến thức từ chọn xe đến nâng cấp, bảo dưỡng chuyên sâu.',
            'cta_button': 'Chọn loại xe →',
            'products_title': 'Phụ tùng & Phụ kiện xe đạp',
            'products_subtitle': 'Groupset, vành, lốp, phụ kiện — giá tốt từ Shopee, Lazada, Tiki',
            'products_icon': '🚴',
            'parts_label': 'phụ tùng',
            'parts_heading': 'Phụ tùng xe đạp',
        },
        'beauty': {
            'hero_title': 'Kiến thức làm đẹp',
            'hero_subtitle': 'skincare, makeup & chăm sóc bản thân',
            'hero_desc': 'Review mỹ phẩm, hướng dẫn skincare routine, so sánh thành phần. Serum, kem dưỡng, chống nắng — chọn đúng cho da bạn.',
            'hero_cta1': 'Khám phá danh mục',
            'segments_label': 'danh mục',
            'segments_heading': 'Danh mục làm đẹp',
            'tier1_desc': 'Thị trường mỹ phẩm, xu hướng K-beauty, clean beauty.',
            'tier2_desc': 'Quy trình skincare, phương pháp chăm sóc da theo loại da.',
            'tier3_desc': 'Review chi tiết sản phẩm: thành phần, cách dùng, so sánh giá.',
            'cta_title': 'Chưa tìm thấy sản phẩm phù hợp?',
            'cta_desc': 'Hệ thống review mỹ phẩm đang mở rộng mỗi ngày. Khám phá thêm theo danh mục.',
            'cta_button': 'Xem danh mục →',
            'products_title': 'Mỹ phẩm & Skincare',
            'products_subtitle': 'Review chi tiết, giá tốt từ Shopee, Lazada, Tiki',
            'products_icon': '💄',
            'parts_label': 'sản phẩm',
            'parts_heading': 'Sản phẩm',
        },
        'tech': {
            'hero_title': 'Kiến thức công nghệ',
            'hero_subtitle': 'smartphone, tai nghe & gadgets',
            'hero_desc': 'So sánh chip, camera, màn hình. Review điện thoại, tai nghe, phụ kiện. Chọn thiết bị phù hợp nhu cầu và ngân sách.',
            'hero_cta1': 'Khám phá danh mục',
            'segments_label': 'danh mục',
            'segments_heading': 'Danh mục thiết bị',
            'tier1_desc': 'Thị trường smartphone, xu hướng AI, chip mới nhất.',
            'tier2_desc': 'Hệ thống camera, màn hình, pin — cấu tạo và so sánh.',
            'tier3_desc': 'Review chi tiết: benchmark, camera test, pin thực tế.',
            'cta_title': 'Cần tư vấn chọn thiết bị?',
            'cta_desc': 'So sánh chi tiết giữa các thiết bị, phụ kiện theo ngân sách.',
            'cta_button': 'Xem danh mục →',
            'products_title': 'Điện thoại & Phụ kiện',
            'products_subtitle': 'Giá tốt nhất từ Shopee, Lazada, Tiki, CellphoneS',
            'products_icon': '📱',
            'parts_label': 'thiết bị',
            'parts_heading': 'Thiết bị',
        },
        'sport': {
            'hero_title': 'Kiến thức thể thao',
            'hero_subtitle': 'từ tập luyện đến thi đấu chuyên nghiệp',
            'hero_desc': 'Giày chạy, đồ gym, dinh dưỡng thể thao, thiết bị tập luyện. Review chi tiết, so sánh giá, hướng dẫn chọn đúng.',
            'hero_cta1': 'Khám phá bộ môn',
            'segments_label': 'bộ môn',
            'segments_heading': 'Bộ môn thể thao',
            'tier1_desc': 'Thị trường thể thao, xu hướng fitness, chi phí tập luyện.',
            'tier2_desc': 'Kỹ thuật tập luyện, phương pháp, chương trình training.',
            'tier3_desc': 'Review chi tiết gear: giày, đồ tập, thiết bị, dinh dưỡng.',
            'cta_title': 'Bắt đầu hành trình thể thao?',
            'cta_desc': 'Khám phá kiến thức từ cơ bản đến nâng cao cho mọi bộ môn.',
            'cta_button': 'Chọn bộ môn →',
            'products_title': 'Đồ thể thao & Thiết bị',
            'products_subtitle': 'Giày, quần áo, gear — giá tốt từ Shopee, Lazada, Decathlon',
            'products_icon': '🏋️',
            'parts_label': 'sản phẩm',
            'parts_heading': 'Sản phẩm',
        },
        'garden': {
            'hero_title': 'Kiến thức làm vườn',
            'hero_subtitle': 'chăm sóc cây trồng từ hạt giống đến thu hoạch',
            'hero_desc': 'Trồng rau sạch, hoa, cây ăn quả, cây cảnh. Hướng dẫn chi tiết, review dụng cụ, phân bón, giá thể.',
            'hero_cta1': 'Khám phá danh mục',
            'segments_label': 'danh mục',
            'segments_heading': 'Danh mục cây trồng',
            'tier1_desc': 'Tổng quan thị trường, xu hướng làm vườn đô thị, nông nghiệp sạch.',
            'tier2_desc': 'Kỹ thuật trồng, chăm sóc, phòng bệnh cho cây.',
            'tier3_desc': 'Hướng dẫn chi tiết từng loại cây, review sản phẩm vật tư.',
            'cta_title': 'Bắt đầu hành trình làm vườn?',
            'cta_desc': 'Khám phá kiến thức từ chọn giống đến thu hoạch, phù hợp mọi không gian.',
            'cta_button': 'Xem danh mục →',
            'products_title': 'Dụng cụ & Vật tư làm vườn',
            'products_subtitle': 'Đất, phân bón, chậu, hạt giống — giá tốt từ Shopee, Lazada, Tiki',
            'products_icon': '🌱',
            'parts_label': 'loại cây / sản phẩm',
            'parts_heading': 'Loại cây & Sản phẩm',
        },
    }
    return configs.get(vertical_slug, {})

@app.route('/<vertical_slug>')
def vertical_index(vertical_slug):
    v = Vertical.query.filter_by(slug=vertical_slug).first_or_404()
    articles_nganh = Article.query.filter_by(vertical_slug=vertical_slug, tier='nganh', status='published').order_by(Article.created_at.desc()).limit(4).all()
    articles_chung = Article.query.filter_by(vertical_slug=vertical_slug, tier='chung', status='published').order_by(Article.created_at.desc()).limit(6).all()
    articles_chitiet = Article.query.filter_by(vertical_slug=vertical_slug, tier='chi-tiet', status='published').order_by(Article.created_at.desc()).limit(6).all()
    recent_articles = Article.query.filter_by(vertical_slug=vertical_slug, status='published').order_by(Article.created_at.desc()).limit(8).all()
    config = get_vertical_config(vertical_slug)
    template = get_template_path(vertical_slug, 'index.html')
    return render_template(template, vertical=v, articles_nganh=articles_nganh,
        articles_chung=articles_chung, articles_chitiet=articles_chitiet, recent_articles=recent_articles, **config)

@app.route('/<vertical_slug>/kien-thuc')
def vertical_knowledge(vertical_slug):
    v = Vertical.query.filter_by(slug=vertical_slug).first_or_404()
    articles_nganh = Article.query.filter_by(vertical_slug=vertical_slug, tier='nganh', status='published').order_by(Article.created_at.desc()).all()
    articles_chung = Article.query.filter_by(vertical_slug=vertical_slug, tier='chung', status='published').order_by(Article.created_at.desc()).all()
    articles_chitiet = Article.query.filter_by(vertical_slug=vertical_slug, tier='chi-tiet', status='published').order_by(Article.created_at.desc()).all()
    template = get_template_path(vertical_slug, 'knowledge.html')
    return render_template(template, vertical=v, articles_nganh=articles_nganh,
        articles_chung=articles_chung, articles_chitiet=articles_chitiet)

@app.route('/<vertical_slug>/san-pham')
def vertical_products(vertical_slug):
    v = Vertical.query.filter_by(slug=vertical_slug).first_or_404()
    f_zone = request.args.get('zone', '')
    f_network = request.args.get('network', '')
    page = request.args.get('page', 1, type=int)
    per_page = 24
    q = db.session.query(AffiliateLink, Part, Zone, Segment).join(
        Part, AffiliateLink.part_id == Part.id
    ).join(Zone, Part.zone_id == Zone.id
    ).join(Segment, Zone.segment_id == Segment.id
    ).filter(Segment.vertical_id == v.id, AffiliateLink.is_active == True)
    if f_zone:
        q = q.filter(Zone.slug == f_zone)
    if f_network:
        q = q.filter(AffiliateLink.network == f_network)
    total_count = q.count()
    pagination = q.order_by(AffiliateLink.price.desc()).paginate(page=page, per_page=per_page, error_out=False)
    products = pagination.items
    segments = Segment.query.filter_by(vertical_id=v.id).all()
    zone_counts = dict(db.session.query(Zone.slug, db.func.count(AffiliateLink.id)).join(
        Part, Zone.id == Part.zone_id).join(AffiliateLink, Part.id == AffiliateLink.part_id
    ).join(Segment, Zone.segment_id == Segment.id).filter(
        Segment.vertical_id == v.id, AffiliateLink.is_active == True
    ).group_by(Zone.slug).all())
    networks = db.session.query(AffiliateLink.network).join(Part).join(Zone).join(Segment).filter(
        Segment.vertical_id == v.id).distinct().all()
    config = get_vertical_config(vertical_slug)
    template = get_template_path(vertical_slug, 'products.html')
    return render_template(template, vertical=v, products=products,
        segments=segments, zone_counts=zone_counts, networks=[n[0] for n in networks],
        f_zone=f_zone, f_network=f_network, product_url='vertical_products',
        pagination=pagination, total_count=total_count, page=page, **config)

@app.route('/<vertical_slug>/bai-viet/<slug>')
def vertical_article(vertical_slug, slug):
    v = Vertical.query.filter_by(slug=vertical_slug).first_or_404()
    a = Article.query.filter_by(vertical_slug=vertical_slug, slug=slug, status='published').first_or_404()
    a.views += 1
    db.session.commit()

    # Related articles (same category or tier)
    related = Article.query.filter(Article.id != a.id, Article.vertical_slug==vertical_slug, Article.status=='published',
        db.or_(Article.category==a.category, Article.tier==a.tier)
    ).order_by(Article.views.desc()).limit(4).all()

    # Featured articles (top by views, different from current)
    featured = Article.query.filter(Article.id != a.id, Article.vertical_slug==vertical_slug, Article.status=='published'
    ).order_by(Article.views.desc()).limit(5).all()

    # Related parts for product carousel — flatten to individual affiliate cards
    carousel_items = []
    # Check if embed_code uses product_ids format (selected products from admin UI)
    embed_code_val = (a.embed_code or '').strip()
    if embed_code_val.startswith('product_ids:'):
        try:
            pids = [int(x) for x in embed_code_val.replace('product_ids:', '').split(',') if x.strip()]
            if pids:
                carousel_items = AffiliateLink.query.filter(AffiliateLink.id.in_(pids), AffiliateLink.is_active==True).all()
        except (ValueError, TypeError):
            pass
    elif a.related_zone_slug:
        z = Zone.query.filter_by(slug=a.related_zone_slug).first()
        if z:
            carousel_limit = int(SiteSettings.get('carousel_product_limit', '3'))
            parts = Part.query.filter_by(zone_id=z.id, status='published').all()
            for p in parts:
                for al in p.affiliate_links:
                    if al.is_active:
                        carousel_items.append(al)
                        if len(carousel_items) >= carousel_limit:
                            break
                if len(carousel_items) >= carousel_limit:
                    break

    # Banners for sidebar (vertical-specific or global)
    banners = Banner.query.filter(
        Banner.is_active==True,
        Banner.placement=='sidebar',
        db.or_(Banner.vertical_slug=='', Banner.vertical_slug==vertical_slug)
    ).order_by(Banner.position).all()

    # Inject SEO backlinks into article content
    article_content = _inject_backlinks(a.content or '', vertical_slug, 'article', a.id, a.slug)

    # Outtext backlinks (related articles from backlink system, shown in sidebar)
    outtext_links = BacklinkInstance.query.join(BacklinkKeyword).filter(
        BacklinkInstance.source_type == 'article',
        BacklinkInstance.source_id == a.id,
        BacklinkInstance.link_type == 'outtext',
        BacklinkInstance.status == 'active',
        BacklinkKeyword.is_active == True
    ).limit(5).all()

    # Inline suggestions (quote-style, inserted at 1/3 of article)
    suggest_instances = BacklinkInstance.query.filter_by(
        source_type='article', source_id=a.id,
        link_type='suggest', status='active'
    ).limit(3).all()
    # Resolve to actual Article objects
    inline_suggests = []
    for si in suggest_instances:
        if si.target_type == 'article':
            sa = Article.query.filter_by(slug=si.target_slug, vertical_slug=vertical_slug, status='published').first()
            if sa:
                inline_suggests.append(sa)

    return render_template('shared/article.html', vertical=v, article=a, related=related, featured=featured,
        carousel_items=carousel_items, banners=banners,
        article_content_with_backlinks=article_content, outtext_links=outtext_links,
        inline_suggests=inline_suggests)

@app.route('/article/<int:article_id>/feedback', methods=['POST'])
def submit_article_feedback(article_id):
    """Submit feedback on article accuracy"""
    article = Article.query.get_or_404(article_id)

    feedback = ArticleFeedback(
        article_id=article_id,
        feedback_type=request.form.get('feedback_type', 'other'),
        description=request.form.get('description', ''),
        user_email=request.form.get('user_email', ''),
        status='pending'
    )

    db.session.add(feedback)
    db.session.commit()

    flash('Cảm ơn phản hồi của bạn! Chúng tôi sẽ xem xét và cải thiện nội dung.', 'success')
    return redirect(request.referrer or url_for('vertical_article', vertical_slug=article.vertical_slug, slug=article.slug))

@app.route('/<vertical_slug>/<segment_slug>')
def vertical_segment(vertical_slug, segment_slug):
    v = Vertical.query.filter_by(slug=vertical_slug).first_or_404()
    s = Segment.query.filter_by(vertical_id=v.id, slug=segment_slug).first_or_404()
    articles = Article.query.filter_by(vertical_slug=vertical_slug, related_segment_slug=segment_slug, status='published').order_by(Article.created_at.desc()).limit(4).all()
    config = get_vertical_config(vertical_slug)
    return render_template('shared/segment.html', vertical=v, segment=s, articles=articles, **config)

@app.route('/<vertical_slug>/<segment_slug>/<zone_slug>')
def vertical_zone(vertical_slug, segment_slug, zone_slug):
    v = Vertical.query.filter_by(slug=vertical_slug).first_or_404()
    s = Segment.query.filter_by(vertical_id=v.id, slug=segment_slug).first_or_404()
    z = Zone.query.filter_by(segment_id=s.id, slug=zone_slug).first_or_404()
    articles = Article.query.filter_by(vertical_slug=vertical_slug, related_zone_slug=zone_slug, status='published').order_by(Article.created_at.desc()).limit(4).all()
    if not articles:
        articles = Article.query.filter_by(vertical_slug=vertical_slug, category=zone_slug, status='published').limit(4).all()
    config = get_vertical_config(vertical_slug)
    return render_template('shared/zone.html', vertical=v, segment=s, zone=z, articles=articles, **config)

@app.route('/<vertical_slug>/<segment_slug>/<zone_slug>/<part_slug>')
def vertical_part(vertical_slug, segment_slug, zone_slug, part_slug):
    v = Vertical.query.filter_by(slug=vertical_slug).first_or_404()
    s = Segment.query.filter_by(vertical_id=v.id, slug=segment_slug).first_or_404()
    z = Zone.query.filter_by(segment_id=s.id, slug=zone_slug).first_or_404()
    p = Part.query.filter_by(zone_id=z.id, slug=part_slug).first_or_404()
    related_articles = Article.query.filter_by(vertical_slug=vertical_slug, related_zone_slug=zone_slug, status='published').limit(3).all()
    if not related_articles:
        related_articles = Article.query.filter_by(vertical_slug=vertical_slug, tier='chi-tiet', status='published').limit(3).all()
    related_parts = Part.query.filter(Part.zone_id==z.id, Part.id!=p.id, Part.status=='published').limit(4).all()
    zone_parts = Part.query.filter_by(zone_id=z.id, status='published').order_by(Part.order).all()
    config = get_vertical_config(vertical_slug)
    # Inject SEO backlinks into part content
    part_content = _inject_backlinks(p.content or '', vertical_slug, 'part', p.id, p.slug)
    return render_template('shared/part.html', vertical=v, segment=s, zone=z, part=p,
        related_articles=related_articles, related_parts=related_parts, zone_parts=zone_parts,
        part_content_with_backlinks=part_content, **config)

# =============================================
# SHOP ROUTES (Standalone e-commerce aggregator)
# =============================================
@app.route('/shop')
def shop_index():
    """UniShop — aggregated product listing"""
    # Shop display mode from settings
    shop_mode = SiteSettings.get('shop_display_mode', 'hub')  # hub (default) or vertical_only

    # Filters
    f_vertical = request.args.get('vertical', '')
    f_zone = request.args.get('zone', '')
    f_network = request.args.get('network', '')
    f_q = request.args.get('q', '').strip()
    f_sort = request.args.get('sort', 'popular')
    f_price_min = request.args.get('price_min', '', type=str)
    f_price_max = request.args.get('price_max', '', type=str)
    page = request.args.get('page', 1, type=int)
    per_page = 24

    # Base query: join all the way to Vertical
    q = db.session.query(AffiliateLink, Part, Zone, Segment, Vertical).join(
        Part, AffiliateLink.part_id == Part.id
    ).join(Zone, Part.zone_id == Zone.id
    ).join(Segment, Zone.segment_id == Segment.id
    ).join(Vertical, Segment.vertical_id == Vertical.id
    ).filter(AffiliateLink.is_active == True)

    # vertical_only mode: only show products from published verticals
    if shop_mode == 'vertical_only':
        q = q.filter(Vertical.status == 'published')

    # Apply filters
    if f_vertical:
        q = q.filter(Vertical.slug == f_vertical)
    if f_zone:
        q = q.filter(Zone.slug == f_zone)
    if f_network:
        q = q.filter(AffiliateLink.network == f_network)
    if f_q:
        search = f'%{f_q}%'
        q = q.filter(db.or_(
            AffiliateLink.product_name.ilike(search),
            Part.name_vi.ilike(search),
            Part.tags.ilike(search),
        ))
    if f_price_min:
        try:
            q = q.filter(AffiliateLink.price >= float(f_price_min))
        except ValueError:
            pass
    if f_price_max:
        try:
            q = q.filter(AffiliateLink.price <= float(f_price_max))
        except ValueError:
            pass

    # Sorting
    if f_sort == 'price_asc':
        q = q.order_by(AffiliateLink.price.asc())
    elif f_sort == 'price_desc':
        q = q.order_by(AffiliateLink.price.desc())
    elif f_sort == 'newest':
        q = q.order_by(AffiliateLink.id.desc())
    else:  # popular
        q = q.order_by(AffiliateLink.clicks.desc(), AffiliateLink.price.desc())

    # Paginate (already includes total count internally)
    pagination = q.paginate(page=page, per_page=per_page, error_out=False)
    products = pagination.items
    filtered_count = pagination.total

    # All verticals for nav (in vertical_only mode, only published non-hub verticals)
    if shop_mode == 'vertical_only':
        verticals = Vertical.query.filter(Vertical.status == 'published', Vertical.slug != 'hub').order_by(Vertical.name).all()
    else:
        verticals = Vertical.query.order_by(Vertical.name).all()

    # Per-vertical counts + total in one query
    vc_rows = db.session.query(Vertical.slug, db.func.count(AffiliateLink.id)).join(
        Segment, Vertical.id == Segment.vertical_id
    ).join(Zone, Segment.id == Zone.segment_id
    ).join(Part, Zone.id == Part.zone_id
    ).join(AffiliateLink, Part.id == AffiliateLink.part_id
    ).filter(AffiliateLink.is_active == True
    ).group_by(Vertical.slug).all()
    vertical_counts = dict(vc_rows)
    total_products = sum(c for _, c in vc_rows)

    # Sidebar: single query for all zone counts (replaces N+1)
    zone_count_q = db.session.query(
        Zone.id, Zone.slug, Zone.name, Zone.icon, Zone.color, Zone.segment_id, Zone.order,
        db.func.count(AffiliateLink.id).label('cnt')
    ).join(Part, Zone.id == Part.zone_id
    ).join(AffiliateLink, Part.id == AffiliateLink.part_id
    ).filter(AffiliateLink.is_active == True)
    if f_vertical:
        zone_count_q = zone_count_q.join(Segment, Zone.segment_id == Segment.id
        ).join(Vertical, Segment.vertical_id == Vertical.id
        ).filter(Vertical.slug == f_vertical)
    zone_count_q = zone_count_q.group_by(Zone.id).all()

    # Build zone lookup: segment_id -> [(slug, name, icon, color, count)]
    zone_by_seg = {}
    active_seg_ids = set()
    for zid, zslug, zname, zicon, zcolor, zsid, zorder, zcnt in zone_count_q:
        if zcnt > 0:
            zone_by_seg.setdefault(zsid, []).append((zorder, zslug, zname, zicon, zcolor, zcnt))
            active_seg_ids.add(zsid)

    # Get segments that have active zones
    sidebar_segments = []
    if active_seg_ids:
        seg_q = Segment.query.filter(Segment.id.in_(active_seg_ids)).order_by(Segment.order)
        for seg in seg_q.all():
            zones_data = [(s, n, i, c, cnt) for _, s, n, i, c, cnt in sorted(zone_by_seg.get(seg.id, []))]
            if zones_data:
                sidebar_segments.append((seg.name, seg.icon, zones_data))

    # Network stats
    net_q = db.session.query(
        AffiliateLink.network, db.func.count(AffiliateLink.id)
    ).filter(AffiliateLink.is_active == True)
    if f_vertical:
        net_q = net_q.join(Part).join(Zone).join(Segment).join(Vertical).filter(Vertical.slug == f_vertical)
    network_stats = net_q.group_by(AffiliateLink.network).order_by(
        db.func.count(AffiliateLink.id).desc()).all()

    # Active names for display
    active_vertical_name = ''
    active_zone_name = ''
    if f_vertical:
        vt_obj = Vertical.query.filter_by(slug=f_vertical).first()
        if vt_obj:
            active_vertical_name = vt_obj.name
    if f_zone:
        z_obj = Zone.query.filter_by(slug=f_zone).first()
        if z_obj:
            active_zone_name = z_obj.name

    return render_template('shop/index.html',
        products=products,
        pagination=pagination,
        verticals=verticals,
        total_products=total_products,
        total_verticals=len(verticals),
        filtered_count=filtered_count,
        vertical_counts=vertical_counts,
        sidebar_segments=sidebar_segments,
        network_stats=network_stats,
        shop_mode=shop_mode,
        f_vertical=f_vertical,
        f_zone=f_zone,
        f_network=f_network,
        f_q=f_q,
        f_sort=f_sort,
        f_price_min=f_price_min,
        f_price_max=f_price_max,
        active_vertical_name=active_vertical_name,
        active_zone_name=active_zone_name,
    )


# =============================================
# VOUCHER ROUTES (Standalone feature)
# =============================================
@app.route('/voucher')
def voucher_index():
    """Voucher listing page with filters"""
    from sqlalchemy import func
    f_category = request.args.get('category', '')
    f_merchant = request.args.get('merchant', '')
    f_type = request.args.get('type', '')  # percentage, fixed_amount, free_shipping
    f_platform = request.args.get('platform', '')  # shopee, lazada, grab...
    f_sort = request.args.get('sort', 'newest')  # newest, discount, expiring
    page = request.args.get('page', 1, type=int)
    per_page = 24

    # Build query
    q = Voucher.query.filter_by(is_active=True)
    if f_category:
        q = q.filter_by(category=f_category)
    if f_merchant:
        q = q.filter(Voucher.merchant.ilike(f'%{f_merchant}%'))
    if f_type:
        q = q.filter_by(discount_type=f_type)
    if f_platform:
        q = q.filter(db.or_(
            Voucher.merchant.ilike(f'%{f_platform}%'),
            Voucher.network.ilike(f'%{f_platform}%')
        ))

    # Get valid vouchers only
    all_vouchers = q.all()
    vouchers = [v for v in all_vouchers if v.is_valid()]

    # Sort
    if f_sort == 'discount':
        vouchers.sort(key=lambda v: v.discount_value or 0, reverse=True)
    elif f_sort == 'expiring':
        vouchers.sort(key=lambda v: v.valid_to or datetime.max)
    else:
        vouchers.sort(key=lambda v: v.created_at or datetime.min, reverse=True)

    # Pagination
    total_vouchers = len(vouchers)
    total_pages = max(1, (total_vouchers + per_page - 1) // per_page)
    page = min(page, total_pages)
    vouchers_paginated = vouchers[(page - 1) * per_page : page * per_page]

    # Featured: top discount vouchers as featured if none marked
    featured = Voucher.query.filter_by(is_active=True, is_featured=True).limit(6).all()
    featured = [v for v in featured if v.is_valid()]
    if not featured:
        # Auto-feature: highest discount % vouchers
        featured = sorted(
            [v for v in Voucher.query.filter_by(is_active=True, discount_type='percentage').all() if v.is_valid()],
            key=lambda v: v.discount_value or 0, reverse=True
        )[:6]

    # Get category counts
    cat_counts = db.session.query(Voucher.category, func.count(Voucher.id)).filter_by(is_active=True).group_by(Voucher.category).all()
    categories = dict(cat_counts)

    # Get merchant list with counts
    merchant_counts = db.session.query(Voucher.merchant, func.count(Voucher.id)).filter_by(is_active=True).group_by(Voucher.merchant).order_by(func.count(Voucher.id).desc()).all()
    merchants = [m[0] for m in merchant_counts]
    merchant_count_map = dict(merchant_counts)

    # Platform stats (group by network/domain)
    platform_stats = {}
    for v in Voucher.query.filter_by(is_active=True).all():
        # Determine platform from merchant name or network
        name = (v.merchant or '').lower()
        if 'shopee' in name:
            p = 'Shopee'
        elif 'lazada' in name:
            p = 'Lazada'
        elif 'tiki' in name:
            p = 'Tiki'
        elif 'grab' in name:
            p = 'Grab'
        elif 'traveloka' in name:
            p = 'Traveloka'
        else:
            p = 'Khác'
        platform_stats[p] = platform_stats.get(p, 0) + 1

    # Expiring soon (within 3 days)
    expiring_soon = sorted(
        [v for v in vouchers if v.valid_to and (v.valid_to - datetime.utcnow()).days <= 3],
        key=lambda v: v.valid_to
    )[:6]

    # Get active voucher widgets for display
    widgets = VoucherWidget.query.filter_by(is_active=True, placement='voucher_page').order_by(VoucherWidget.position).all()

    return render_template('voucher/index.html',
        vouchers=vouchers_paginated, total_vouchers=total_vouchers,
        page=page, total_pages=total_pages, per_page=per_page,
        featured=featured, categories=categories,
        merchants=merchants, merchant_count_map=merchant_count_map,
        platform_stats=platform_stats, expiring_soon=expiring_soon,
        widgets=widgets, f_category=f_category, f_merchant=f_merchant,
        f_type=f_type, f_platform=f_platform, f_sort=f_sort,
        now=datetime.utcnow())

@app.route('/voucher/nhan-hang')
def voucher_brands():
    """Brands page — show all AT banners, one per brand"""
    import re as _re
    now_utc = datetime.utcnow()
    all_banners = AccessTradeBanner.query.filter(
        AccessTradeBanner.is_active == True
    ).order_by(AccessTradeBanner.synced_at.desc()).all()

    # Dedup: each brand (merchant + sub-brand) appears only once
    seen = set()
    brands = []
    for ab in all_banners:
        if ab.end_date and ab.end_date < now_utc:
            continue
        merchant = (ab.merchant or '').strip().lower()
        offer = ab.offer_name or ''
        m = _re.search(r'\[([^\]]+)\]', offer)
        brand_key = f"{merchant}_{m.group(1).strip().lower()}" if m else merchant
        if brand_key and brand_key not in seen:
            seen.add(brand_key)
            brands.append(ab)

    return render_template('voucher/brands.html', brands=brands)

@app.route('/voucher/nhan-hang/click/<int:banner_id>', methods=['POST'])
def voucher_brand_click(banner_id):
    """Track brand click"""
    ab = AccessTradeBanner.query.get(banner_id)
    if ab:
        ab.clicks = (ab.clicks or 0) + 1
        db.session.commit()
    return '', 204

@app.route('/voucher/<code>')
def voucher_detail(code):
    """Voucher detail page"""
    v = Voucher.query.filter_by(code=code).first_or_404()

    # Track click
    v.clicks = (v.clicks or 0) + 1
    db.session.commit()

    # Related vouchers (same category or merchant)
    related = Voucher.query.filter(
        Voucher.id != v.id,
        Voucher.is_active == True,
        db.or_(Voucher.category == v.category, Voucher.merchant == v.merchant)
    ).limit(6).all()
    related = [r for r in related if r.is_valid()]

    return render_template('voucher/detail.html', voucher=v, related=related)

@app.route('/voucher/<int:vid>/use', methods=['POST'])
def voucher_use(vid):
    """Track voucher usage"""
    v = Voucher.query.get_or_404(vid)
    v.usage_count = (v.usage_count or 0) + 1
    v.conversions = (v.conversions or 0) + 1
    db.session.commit()
    return {'status': 'ok', 'usage_count': v.usage_count}

# =============================================
# INIT
# =============================================
def _get_table_columns(table_name):
    """Get set of column names for a table using PRAGMA (single query)."""
    try:
        rows = db.session.execute(db.text(f"PRAGMA table_info({table_name})")).fetchall()
        return {row[1] for row in rows}
    except:
        return set()

def _table_exists(table_name):
    """Check if a table exists."""
    try:
        rows = db.session.execute(db.text(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=:t"
        ), {'t': table_name}).fetchall()
        return len(rows) > 0
    except:
        return False

def _ensure_column(table_name, col_name, col_def, existing_cols):
    """Add column if missing. Returns True if added."""
    if col_name not in existing_cols:
        try:
            print(f'  [+] Adding {table_name}.{col_name}')
            db.session.execute(db.text(f"ALTER TABLE {table_name} ADD COLUMN {col_name} {col_def}"))
            return True
        except Exception as e:
            print(f'  [!] Failed to add {table_name}.{col_name}: {e}')
            return False
    return False

def _run_schema_migration():
    """Fast schema migration — uses PRAGMA to batch-check columns.
    Resilient: each migration step is independent, failures don't block others."""
    import time
    t0 = time.time()
    print('[*] Checking database schema...')

    changed = False

    # --- Zone table ---
    zone_cols = _get_table_columns('zone')
    if _ensure_column('zone', 'seo_content', "TEXT DEFAULT ''", zone_cols):
        changed = True

    # --- Banner table ---
    if not _table_exists('banner'):
        print('  [+] Creating banner table')
        db.session.execute(db.text("""
            CREATE TABLE IF NOT EXISTS banner (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name VARCHAR(200) NOT NULL,
                placement VARCHAR(50) DEFAULT 'sidebar',
                vertical_slug VARCHAR(50) DEFAULT '',
                image_url VARCHAR(500) DEFAULT '',
                link_url VARCHAR(500) DEFAULT '',
                html_code TEXT DEFAULT '',
                position INTEGER DEFAULT 1,
                is_active BOOLEAN DEFAULT 1,
                impressions INTEGER DEFAULT 0,
                clicks INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """))
        changed = True

    # --- Vertical table ---
    vert_cols = _get_table_columns('vertical')
    if _ensure_column('vertical', 'template', "VARCHAR(20) DEFAULT 'general'", vert_cols):
        changed = True
    if _ensure_column('vertical', 'style', "VARCHAR(20) DEFAULT 'classic'", vert_cols):
        changed = True
    if _ensure_column('vertical', 'shop_link', "VARCHAR(500) DEFAULT ''", vert_cols):
        changed = True

    # --- Performance indexes for shop page ---
    try:
        db.session.execute(db.text("CREATE INDEX IF NOT EXISTS ix_affiliate_link_part_id ON affiliate_link (part_id)"))
        db.session.execute(db.text("CREATE INDEX IF NOT EXISTS ix_affiliate_link_is_active ON affiliate_link (is_active)"))
        db.session.execute(db.text("CREATE INDEX IF NOT EXISTS ix_affiliate_link_network ON affiliate_link (network)"))
        db.session.execute(db.text("CREATE INDEX IF NOT EXISTS ix_part_zone_id ON part (zone_id)"))
        db.session.execute(db.text("CREATE INDEX IF NOT EXISTS ix_zone_segment_id ON zone (segment_id)"))
        db.session.execute(db.text("CREATE INDEX IF NOT EXISTS ix_segment_vertical_id ON segment (vertical_id)"))
    except Exception:
        pass

    # --- AffiliateLink table ---
    al_cols = _get_table_columns('affiliate_link')
    if _ensure_column('affiliate_link', 'category', "VARCHAR(100) DEFAULT ''", al_cols):
        changed = True

    # --- Affiliate Network table ---
    an_cols = _get_table_columns('affiliate_network')
    if _ensure_column('affiliate_network', 'deeplink_template', "VARCHAR(1000) DEFAULT ''", an_cols):
        changed = True

    # --- Voucher table ---
    v_cols = _get_table_columns('voucher')
    if _ensure_column('voucher', 'embed_code', "TEXT DEFAULT ''", v_cols):
        changed = True
    if _ensure_column('voucher', 'sync_mode', "VARCHAR(20) DEFAULT 'manual'", v_cols):
        changed = True
    if _ensure_column('voucher', 'accesstrade_offer_id', "VARCHAR(100) DEFAULT ''", v_cols):
        changed = True

    # --- Scheduled CSV Import table ---
    if not _table_exists('scheduled_csv_import'):
        print('  [+] Creating scheduled_csv_import table')
        db.session.execute(db.text("""
            CREATE TABLE IF NOT EXISTS scheduled_csv_import (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name VARCHAR(200) NOT NULL,
                csv_url VARCHAR(1000) NOT NULL,
                part_id INTEGER NOT NULL,
                network VARCHAR(50) DEFAULT 'shopee',
                apply_deeplink BOOLEAN DEFAULT 0,
                update_interval_days INTEGER DEFAULT 7,
                is_active BOOLEAN DEFAULT 1,
                last_import_at DATETIME,
                next_import_at DATETIME,
                last_import_count INTEGER DEFAULT 0,
                last_import_status VARCHAR(20) DEFAULT 'pending',
                last_error TEXT DEFAULT '',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (part_id) REFERENCES part(id)
            )
        """))
        changed = True

    # --- Voucher Widget table ---
    if not _table_exists('voucher_widget'):
        print('  [+] Creating voucher_widget table')
        db.session.execute(db.text("""
            CREATE TABLE IF NOT EXISTS voucher_widget (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name VARCHAR(200) NOT NULL,
                network VARCHAR(50) DEFAULT 'accesstrade',
                embed_code TEXT NOT NULL,
                description TEXT DEFAULT '',
                placement VARCHAR(50) DEFAULT 'voucher_page',
                position INTEGER DEFAULT 1,
                is_active BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """))
        changed = True

    # --- Hot Deal table ---
    if not _table_exists('hot_deal'):
        print('  [+] Creating hot_deal table')
        db.session.execute(db.text("""
            CREATE TABLE IF NOT EXISTS hot_deal (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name VARCHAR(500) NOT NULL,
                campaign VARCHAR(200) DEFAULT '',
                product_link VARCHAR(1000) DEFAULT '',
                start_date DATETIME NOT NULL,
                end_date DATETIME NOT NULL,
                status VARCHAR(50) DEFAULT '',
                hot_day VARCHAR(100) DEFAULT '',
                banner TEXT DEFAULT '',
                detail TEXT DEFAULT '',
                is_active BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """))
        changed = True

    # --- Hotel table ---
    hotel_cols = _get_table_columns('hotel')
    if hotel_cols:
        if _ensure_column('hotel', 'price_original', "FLOAT DEFAULT 0", hotel_cols):
            changed = True
        if _ensure_column('hotel', 'latitude', "FLOAT DEFAULT 0", hotel_cols):
            changed = True
        if _ensure_column('hotel', 'longitude', "FLOAT DEFAULT 0", hotel_cols):
            changed = True
        if _ensure_column('hotel', 'address', "VARCHAR(500) DEFAULT ''", hotel_cols):
            changed = True

    # --- Attraction table ---
    attraction_cols = _get_table_columns('attraction')
    if attraction_cols:
        if _ensure_column('attraction', 'price_original', "FLOAT DEFAULT 0", attraction_cols):
            changed = True

    # --- Backlink Keyword table ---
    if not _table_exists('backlink_keyword'):
        print('  [+] Creating backlink_keyword table')
        db.session.execute(db.text("""
            CREATE TABLE IF NOT EXISTS backlink_keyword (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                vertical_slug VARCHAR(50) DEFAULT '',
                keyword VARCHAR(200) NOT NULL,
                target_type VARCHAR(20) NOT NULL,
                target_slug VARCHAR(200) NOT NULL,
                target_title VARCHAR(300) DEFAULT '',
                anchor_text VARCHAR(200) DEFAULT '',
                priority INTEGER DEFAULT 5,
                max_per_page INTEGER DEFAULT 1,
                is_active BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """))
        db.session.execute(db.text("CREATE INDEX IF NOT EXISTS idx_bk_vertical ON backlink_keyword(vertical_slug)"))
        db.session.execute(db.text("CREATE INDEX IF NOT EXISTS idx_bk_active ON backlink_keyword(is_active)"))
        changed = True

    # --- Backlink Instance table ---
    if not _table_exists('backlink_instance'):
        print('  [+] Creating backlink_instance table')
        db.session.execute(db.text("""
            CREATE TABLE IF NOT EXISTS backlink_instance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                keyword_id INTEGER NOT NULL,
                source_type VARCHAR(20) NOT NULL,
                source_id INTEGER NOT NULL,
                source_slug VARCHAR(200) DEFAULT '',
                source_title VARCHAR(300) DEFAULT '',
                target_type VARCHAR(20) NOT NULL,
                target_slug VARCHAR(200) NOT NULL,
                link_type VARCHAR(10) DEFAULT 'intext',
                anchor_text VARCHAR(200) DEFAULT '',
                status VARCHAR(15) DEFAULT 'active',
                clicks INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (keyword_id) REFERENCES backlink_keyword(id)
            )
        """))
        db.session.execute(db.text("CREATE INDEX IF NOT EXISTS idx_bi_keyword ON backlink_instance(keyword_id)"))
        db.session.execute(db.text("CREATE INDEX IF NOT EXISTS idx_bi_source ON backlink_instance(source_type, source_id)"))
        changed = True

    if changed:
        db.session.commit()

    # --- Auto-seed WardCommune from JSON if table is empty ---
    try:
        seeded = _auto_seed_wards()
        if seeded:
            print(f'  [+] Seeded {seeded} wards from wards_default.json')
    except Exception:
        pass  # WardCommune table might not exist yet

    elapsed = round((time.time() - t0) * 1000)
    print(f'[*] Schema check done ({elapsed}ms)')


# =============================================
# BACKGROUND SCHEDULER — Banner Auto Delete + Sync
# =============================================
import threading

_banner_stop_flag = threading.Event()

def _banner_scheduler_loop():
    """Background loop: waits until configured time, then delete-all + sync. Repeats daily."""
    import time as _time
    while not _banner_stop_flag.is_set():
        try:
            with app.app_context():
                if SiteSettings.get('banner_auto_sync', 'on') != 'on':
                    break
                sync_time = SiteSettings.get('banner_sync_time', '03:00')
                try:
                    target_h, target_m = int(sync_time.split(':')[0]), int(sync_time.split(':')[1])
                except Exception:
                    target_h, target_m = 3, 0

            now = datetime.now()
            target = now.replace(hour=target_h, minute=target_m, second=0, microsecond=0)
            if target <= now:
                target += timedelta(days=1)
            wait_secs = (target - now).total_seconds()

            # Sleep in 30s increments to allow stop flag check
            if _banner_stop_flag.wait(timeout=wait_secs):
                return  # Stopped

            # Execute: delete all → sync
            with app.app_context():
                if SiteSettings.get('banner_auto_sync', 'on') != 'on':
                    break
                try:
                    count = AccessTradeBanner.query.count()
                    AccessTradeBanner.query.delete()
                    db.session.commit()
                    imported, updated, total, err = _do_banner_sync()
                    if err:
                        result = f'Xóa {count}, lỗi sync: {err}'
                    else:
                        result = f'Xóa {count}, sync {imported} mới, {updated} cập nhật'
                except Exception as e:
                    result = f'Lỗi: {str(e)}'
                SiteSettings.set_val('banner_last_auto_sync', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'banner')
                SiteSettings.set_val('banner_last_auto_sync_result', result, 'banner')
                db.session.commit()
        except Exception:
            _time.sleep(60)


def _start_banner_scheduler():
    """Start (or restart) the banner auto-sync background thread."""
    global _banner_stop_flag
    _banner_stop_flag.set()  # Signal old thread to stop
    _banner_stop_flag = threading.Event()  # Fresh flag for new thread

    with app.app_context():
        if SiteSettings.get('banner_auto_sync', 'on') != 'on':
            return

    t = threading.Thread(target=_banner_scheduler_loop, daemon=True)
    t.start()


@app.errorhandler(404)
def page_not_found(e):
    target = SiteSettings.get('redirect_404_target', 'home')
    target_map = {
        'home': '/',
        'shop': '/shop',
        'voucher': '/voucher',
    }
    dest = target_map.get(target, '/')
    return redirect(dest)


if __name__ == '__main__':
    import os, shutil, time, gc

    def _force_close_db():
        """Force-close all SQLAlchemy connections and release file handles."""
        db.session.remove()
        db.engine.dispose()
        gc.collect()
        time.sleep(0.5)  # Give OS time to release file locks (Windows)

    def _remove_db_file(db_path):
        """Remove DB file with retry for Windows file locking."""
        for attempt in range(4):
            try:
                os.remove(db_path)
                return True
            except PermissionError:
                gc.collect()
                time.sleep(1 * (attempt + 1))
        # Fallback: rename instead of delete
        try:
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            os.rename(db_path, db_path + f'.old_{ts}')
            return True
        except Exception:
            return False

    def _safe_init_db():
        """Initialize database with auto-recovery on schema errors."""
        db_path = os.path.join(app.instance_path, 'unilab.db')

        # Step 1: If DB exists, run integrity check + auto-repair
        if os.path.exists(db_path):
            try:
                from db_backup import startup_check
                health = startup_check(app)
                if health['action'] != 'none':
                    print(f'[DB] {health["details"]}')
                if not health['healthy']:
                    print(f'[!] DB unhealthy but continuing — will try create_all...')
            except Exception as e:
                print(f'[!] Startup check error: {e}')

        # Step 2: Normal schema init
        try:
            db.create_all()
            _run_schema_migration()
            print('[*] Database ready.')
        except Exception as e:
            err_msg = str(e)
            if 'malformed' in err_msg or 'no such table' in err_msg or 'OperationalError' in err_msg:
                print(f'[!] Database schema error: {err_msg}')
                print('[*] Attempting dump-rebuild repair...')
                _force_close_db()
                try:
                    from db_backup import repair_database, create_backup
                    repair = repair_database(app)
                    if repair['success']:
                        print(f'[*] Repair OK via {repair["method"]}: {repair["tables_recovered"]} tables, {repair["rows_recovered"]} rows')
                        db.create_all()
                        _run_schema_migration()
                        print('[*] Database ready after repair.')
                    else:
                        print(f'[!] Repair failed: {repair["errors"]}')
                        print('[*] Creating fresh database (corrupted DB saved in backups/)...')
                        _force_close_db()
                        if os.path.exists(db_path):
                            if not _remove_db_file(db_path):
                                print(f'[!] Cannot remove locked DB. Close other apps using it and restart.')
                                return
                        db.create_all()
                        _run_schema_migration()
                        print('[*] Fresh database created. Run seed_data to restore content.')
                except Exception as e2:
                    print(f'[!] Full recovery failed: {e2}')
                    # Last resort: fresh DB
                    _force_close_db()
                    if os.path.exists(db_path):
                        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                        try:
                            shutil.copy2(db_path, db_path + f'.crashed_{ts}')
                        except Exception:
                            pass
                        if not _remove_db_file(db_path):
                            print(f'[!] Cannot remove locked DB. Close other apps using it and restart.')
                            return
                    db.create_all()
                    _run_schema_migration()
                    print('[*] Fresh database created (emergency).')
            else:
                raise

    # Only init DB once (skip on Werkzeug reloader child process)
    if not os.environ.get('WERKZEUG_RUN_MAIN'):
        with app.app_context():
            _safe_init_db()

        # Start banner auto-sync scheduler
        _start_banner_scheduler()

        # Auto-open browser after server is ready (only once, not on reload)
        import webbrowser
        threading.Timer(1.5, webbrowser.open, args=['http://localhost:7000/admin']).start()

    app.run(host='0.0.0.0', port=7000, debug=True)
